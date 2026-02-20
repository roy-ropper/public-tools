#!/usr/bin/env python3
"""
pcap_to_drawio.py  v4.0  — Pentester Edition
----------------------------------------------
Converts a .pcap / .pcapng file into:

  1. draw.io diagram  (.drawio)
       - Subnet swim-lane containers, Cisco icons, role colours
       - Hostname / IP / MAC on each node (passively resolved)
       - Pentest flags on nodes (cleartext protocols, lateral movement)
       - Gateway / router detection from ARP and TTL evidence

  2. Excel workbook   (.xlsx)  — 10 sheets, all with AutoFilter + freeze pane:
       • Connections         — every conversation: src→dst, proto, port, resource
       • Node Summary        — per-host: role, OS guess (TTL), open ports, risk flags
       • Pentest Findings    — CRITICAL/HIGH/MEDIUM findings with recommendations
       • Protocol Summary    — proto breakdown, cleartext flags
       • Port Inventory      — passive open-port map per host
       • Cleartext Intercepts— credentials extracted: FTP, HTTP Auth, SMTP, LDAP,
                               SNMP community strings, NetBIOS hostnames (decoded)
       • Banner Intel        — server banners / version strings from 15+ protocols
       • TLS Sessions        — SNI, cipher suites, weak/deprecated version flagging
       • DNS Events          — all queries/responses: A/PTR/MX/CNAME/mDNS/NXDOMAIN
       • Wi-Fi Networks      — APs (BSSID/SSID/enc/WPS), clients, deauth events

Input formats supported:
    .pcap   — little-endian, big-endian, nanosecond variants (auto-detected)
    .pcapng — standard PCAPng (multi-interface captures from Wireshark/tshark)

Link types supported:
    1   Ethernet II            (tcpdump -i eth0)
    101 Raw IPv4               (tunnel / tun0 interfaces)
    105 Raw 802.11             (some Wi-Fi adapters / older airmon-ng)
    113 Linux cooked (SLL)     (tcpdump -i any)
    127 Radiotap + 802.11      (tcpdump -i wlan0mon, monitor mode)
    228 Raw IPv4 (alt)         (some BSD/macOS captures)

Usage:
    python3 pcap_to_drawio.py capture.pcap
    python3 pcap_to_drawio.py capture.pcap -o out.drawio --xlsx out.xlsx
    python3 pcap_to_drawio.py capture.pcap --min-packets 3 --collapse-external
    python3 pcap_to_drawio.py capture.pcap --hostname-file hosts.txt
    python3 pcap_to_drawio.py capture.pcap --internal-networks 20.16.0.0/14

    hosts.txt format (IP<whitespace>hostname, one per line, # = comment):
        192.168.1.1   core-router
        10.0.1.10     dc-01.corp.local

Dependencies:
    Python 3.6+  (stdlib only — struct, socket, ipaddress, argparse, xml)
    openpyxl     (pip install openpyxl) — required for .xlsx output only

Tests:
    python3 test_pcap_tool.py   — 203 tests, 100% pass rate
"""

import struct, socket, ipaddress, argparse, sys, math, os, re
from collections import defaultdict
from xml.dom import minidom
import xml.etree.ElementTree as ET

# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────

PCAP_MAGIC_LE    = 0xA1B2C3D4
PCAP_MAGIC_BE    = 0xD4C3B2A1
PCAP_MAGIC_NS_LE = 0xA1B23C4D
PCAP_MAGIC_NS_BE = 0x4D3CB2A1
PCAPNG_MAGIC     = 0x0A0D0D0A
ETH_TYPE_IP      = 0x0800
ETH_TYPE_IP6     = 0x86DD
ETH_TYPE_ARP     = 0x0806
PROTO_TCP        = 6
PROTO_UDP        = 17
PROTO_ICMP       = 1
PROTO_ICMP6      = 58

WELL_KNOWN = {
    ("TCP",20):"FTP-data", ("TCP",21):"FTP",    ("TCP",22):"SSH",
    ("TCP",23):"Telnet",   ("TCP",25):"SMTP",   ("TCP",53):"DNS",
    ("TCP",80):"HTTP",     ("TCP",110):"POP3",  ("TCP",143):"IMAP",
    ("TCP",389):"LDAP",    ("TCP",443):"HTTPS", ("TCP",445):"SMB",
    ("TCP",636):"LDAPS",   ("TCP",993):"IMAPS", ("TCP",995):"POP3S",
    ("TCP",1433):"MSSQL",  ("TCP",3306):"MySQL",("TCP",3389):"RDP",
    ("TCP",5432):"PostgreSQL",("TCP",5900):"VNC",("TCP",6379):"Redis",
    ("TCP",8080):"HTTP-alt",  ("TCP",8443):"HTTPS-alt",
    ("TCP",27017):"MongoDB",  ("TCP",4444):"Metasploit?",
    ("TCP",4445):"Metasploit?",("TCP",31337):"BackOrifice?",
    ("UDP",53):"DNS",   ("UDP",67):"DHCP",  ("UDP",68):"DHCP",
    ("UDP",69):"TFTP",  ("UDP",123):"NTP",  ("UDP",137):"NetBIOS",
    ("UDP",138):"NetBIOS",("UDP",161):"SNMP",("UDP",162):"SNMP-trap",
    ("UDP",500):"IKE",  ("UDP",514):"Syslog",("UDP",1900):"SSDP",
    ("UDP",4500):"IKE-NAT",("UDP",5353):"mDNS",
}

HTTP_PORTS  = {80, 8080, 8000, 8008}
HTTPS_PORTS = {443, 8443, 4443}

# Pentest: protocols that send credentials in cleartext
CLEARTEXT_PROTOS = {"Telnet","FTP","FTP-data","HTTP","POP3","IMAP",
                    "SMTP","LDAP","SNMP","NetBIOS","TFTP","Syslog"}

# Pentest: interesting/suspicious ports
SUSPICIOUS_PORTS = {
    4444:"Metasploit default", 4445:"Metasploit alt",
    1337:"Leet shell?", 31337:"Back Orifice",
    1234:"Generic backdoor?", 9001:"Tor?", 9050:"Tor proxy",
    6667:"IRC/C2", 6666:"IRC/C2", 6668:"IRC/C2",
    8888:"Alt HTTP / Jupyter", 2222:"Alt SSH",
}

# Pentest: lateral movement indicators
LATERAL_PROTOS = {"SMB","RDP","VNC","NetBIOS","LDAP"}

# TTL → OS guess (coarse)
def _os_from_ttl(ttl):
    if ttl is None:     return "Unknown"
    if ttl <= 64:       return "Linux/Mac"
    if ttl <= 128:      return "Windows"
    if ttl <= 255:      return "Network Device"
    return "Unknown"


# ─────────────────────────────────────────────────────────────────────────────
# PCAP / PCAPng parser
# ─────────────────────────────────────────────────────────────────────────────

def parse_pcap(path):
    """Yield enriched packet dicts."""
    with open(path, "rb") as f:
        magic_bytes = f.read(4)
        if len(magic_bytes) < 4:
            return

        # Detect byte order from raw magic bytes (RFC-correct approach).
        # A correctly-formed PCAP file starts with:
        #   A1 B2 C3 D4  → native/little-endian
        #   A1 B2 3C 4D  → native/little-endian, nanosecond timestamps
        #   D4 C3 B2 A1  → big-endian  (or malformed LE written with wrong pack fmt)
        #   4D 3C B2 A1  → big-endian, nanosecond timestamps
        #   0A 0D 0D 0A  → PCAPng
        _LE_MAGIC    = b"\xa1\xb2\xc3\xd4"
        _LE_NS_MAGIC = b"\xa1\xb2\x3c\x4d"
        _BE_MAGIC    = b"\xd4\xc3\xb2\xa1"
        _BE_NS_MAGIC = b"\x4d\x3c\xb2\xa1"
        _NG_MAGIC    = b"\x0a\x0d\x0d\x0a"

        if magic_bytes in (_LE_MAGIC, _LE_NS_MAGIC):
            endian = "<"
        elif magic_bytes == _NG_MAGIC:
            yield from _parse_pcapng(path)
            return
        elif magic_bytes in (_BE_MAGIC, _BE_NS_MAGIC):
            # Validate by checking version field: PCAP version is always 2.4.
            # Some tools incorrectly write the BE magic bytes but store all
            # fields in little-endian order (a common libpcap bug on LE hosts).
            # If the version reads as (2, 4) in LE but nonsense in BE, treat as LE.
            rest4 = f.read(4)       # version major + minor (2+2 bytes)
            if len(rest4) >= 4:
                vmaj_le, vmin_le = struct.unpack("<HH", rest4)
                vmaj_be, vmin_be = struct.unpack(">HH", rest4)
                if vmaj_le in (1, 2) and vmin_le in (0, 4):
                    endian = "<"    # version sane as LE → file is actually LE
                elif vmaj_be in (1, 2) and vmin_be in (0, 4):
                    endian = ">"    # version sane as BE → genuinely big-endian
                else:
                    endian = "<"    # fallback: assume LE (most common)
                # Re-open to rewind — simpler than tracking position
                f.seek(4)          # skip magic, re-read from after magic
            else:
                endian = "<"
        else:
            raise ValueError(f"Unrecognised PCAP magic: {magic_bytes.hex()}")
        rest = f.read(20)
        # Link type is at bytes 20-23 of global header (index 16-19 of this 20-byte chunk)
        ltype = 1   # default: Ethernet
        if len(rest) >= 20:
            try:
                ltype = struct.unpack(endian + "I", rest[16:20])[0]
            except Exception:
                pass
        while True:
            rec = f.read(16)
            if len(rec) < 16:
                break
            _, ts_us, incl_len, _ = struct.unpack(endian + "IIII", rec)
            raw = f.read(incl_len)
            r = _dissect(raw, ltype, ts_us)
            if r:
                yield r


def _parse_pcapng(path):
    with open(path, "rb") as f:
        endian = "<"
        ltypes = {}
        while True:
            hdr = f.read(8)
            if len(hdr) < 8:
                break
            btype, blen = struct.unpack(endian + "II", hdr)
            body = f.read(blen - 12)
            f.read(4)
            if btype == 0x0A0D0D0A:
                endian = "<" if struct.unpack("<I", body[:4])[0] == 0x1A2B3C4D else ">"
            elif btype == 0x00000001:
                ltypes[len(ltypes)] = struct.unpack(endian + "H", body[:2])[0]
            elif btype == 0x00000006:
                iid = struct.unpack(endian + "I", body[:4])[0]
                ts_hi, ts_lo, cap_len, _ = struct.unpack(endian + "IIII", body[4:20])
                r = _dissect(body[20:20+cap_len], ltypes.get(iid,1), ts_lo)
                if r: yield r
            elif btype == 0x00000003:
                cap_len = struct.unpack(endian+"I", body[:4])[0]
                r = _dissect(body[4:4+cap_len], ltypes.get(0,1), 0)
                if r: yield r


def _mac(b):
    return ":".join(f"{x:02x}" for x in b)


def _dissect(raw, ltype, ts_us=0):
    try:
        smac = dmac = "ff:ff:ff:ff:ff:ff"
        vlan_id = None  # populated for 802.1Q tagged Ethernet frames

        # ── 802.11 / Wi-Fi link types ─────────────────────────────────────
        # ltype 127 = LINKTYPE_IEEE802_11_RADIOTAP
        # ltype 105 = LINKTYPE_IEEE802_11 (raw, no radiotap)
        if ltype in (105, 127):
            # Must have at least some bytes to be meaningful
            if not raw:
                return None
            # Return a synthetic "wifi" packet that extract_wifi_events can read
            return dict(src_ip="", dst_ip="", src_mac="", dst_mac="",
                        src_port=0, dst_port=0, proto="WiFi-802.11",
                        length=len(raw), resource="", ttl=None, ts_us=ts_us,
                        win_size=0, app_payload=b"",
                        arp_sender_mac=None, arp_sender_ip=None,
                        _raw_frame=raw, _wifi=True)

        if ltype == 1:
            if len(raw) < 14: return None
            dmac  = _mac(raw[0:6])
            smac  = _mac(raw[6:12])
            etype = struct.unpack(">H", raw[12:14])[0]
            off   = 14
            vlan_id = None
            while etype in (0x8100, 0x88A8, 0x9100) and off+4 <= len(raw):
                # 802.1Q / QinQ: TCI[0:12] = VLAN ID (12 bits, lower 12 of 16-bit TCI)
                tci = struct.unpack(">H", raw[off:off+2])[0]
                if vlan_id is None:     # keep outermost (first) VLAN tag only
                    vlan_id = tci & 0x0FFF
                etype = struct.unpack(">H", raw[off+2:off+4])[0]; off += 4
            payload = raw[off:]
        elif ltype in (101,228): etype = ETH_TYPE_IP;  payload = raw
        elif ltype == 229:       return None  # raw IPv6 — excluded
        elif ltype == 113:
            if len(raw) < 16: return None
            etype = struct.unpack(">H", raw[14:16])[0]; payload = raw[16:]
        else:
            return None

        if   etype == ETH_TYPE_IP:  r = _ipv4(payload, ts_us)
        elif etype == ETH_TYPE_IP6: return None   # IPv6 excluded — IPv4 only
        elif etype == ETH_TYPE_ARP: r = _arp(payload)
        else: return None

        if r:
            r["src_mac"] = smac
            r["dst_mac"] = dmac
            r["_raw_frame"] = raw   # store for WiFi correlation if needed
            if vlan_id is not None:
                r["vlan_id"] = vlan_id
        return r
    except Exception:
        return None


def _ipv4(d, ts_us):
    if len(d) < 20: return None
    ihl = (d[0] & 0x0F) * 4
    ttl = d[8]
    r = _transport(socket.inet_ntoa(d[12:16]), socket.inet_ntoa(d[16:20]),
                   d[9], d[ihl:], len(d), ts_us)
    if r: r["ttl"] = ttl
    return r


def _ipv6(d, ts_us):
    if len(d) < 40: return None
    r = _transport(socket.inet_ntop(socket.AF_INET6, d[8:24]),
                   socket.inet_ntop(socket.AF_INET6, d[24:40]),
                   d[6], d[40:], len(d), ts_us)
    if r: r["ttl"] = d[7]  # hop limit
    return r


def _arp(d):
    if len(d) < 28: return None
    return dict(src_ip=socket.inet_ntoa(d[14:18]),
                dst_ip=socket.inet_ntoa(d[24:28]),
                src_mac="", dst_mac="",
                src_port=0, dst_port=0, proto="ARP",
                length=len(d), resource="", ttl=None,
                ts_us=0, win_size=0, app_payload=b"",
                # ARP: capture sender MAC for ARP anomaly detection
                arp_sender_mac=_mac(d[8:14]),
                arp_sender_ip=socket.inet_ntoa(d[14:18]))


def _transport(src, dst, proto, data, pkt_len, ts_us):
    sp = dp = win = 0
    resource = ""
    app_payload = b""
    if proto == PROTO_TCP:
        name = "TCP"
        if len(data) >= 4:
            sp, dp = struct.unpack(">HH", data[:4])
            name = WELL_KNOWN.get(("TCP",dp)) or WELL_KNOWN.get(("TCP",sp)) or "TCP"
        if len(data) >= 14:
            win = struct.unpack(">H", data[14:16])[0]
        tcp_hdr_len = ((data[12] >> 4) * 4) if len(data) > 12 else 20
        app_payload = data[tcp_hdr_len:] if len(data) > tcp_hdr_len else b""
        if dp in HTTP_PORTS or sp in HTTP_PORTS:
            resource = _http_host(app_payload)
    elif proto == PROTO_UDP:
        name = "UDP"
        if len(data) >= 4:
            sp, dp = struct.unpack(">HH", data[:4])
            name = WELL_KNOWN.get(("UDP",dp)) or WELL_KNOWN.get(("UDP",sp)) or "UDP"
        app_payload = data[8:] if len(data) > 8 else b""
    elif proto == PROTO_ICMP:
        name = "ICMP"
        app_payload = data   # full ICMP datagram (type+code+checksum+body)
    elif proto == PROTO_ICMP6: name = "ICMPv6"
    else:                      name = f"IP/{proto}"
    return dict(src_ip=src, dst_ip=dst, src_mac="", dst_mac="",
                src_port=sp, dst_port=dp, proto=name,
                length=pkt_len, resource=resource,
                ttl=None, ts_us=ts_us, win_size=win,
                app_payload=app_payload,
                arp_sender_mac=None, arp_sender_ip=None)


def _http_host(payload):
    try:
        text = payload.decode("ascii", errors="ignore")
        for line in text.split("\r\n"):
            if line.lower().startswith("host:"):
                return line.split(":",1)[1].strip()
    except Exception:
        pass
    return ""


# ─────────────────────────────────────────────────────────────────────────────
# Cleartext data / credential extractor
# ─────────────────────────────────────────────────────────────────────────────

import base64 as _b64, re as _re

def extract_cleartext(pkt):
    """
    Inspect application payload for credentials and sensitive data.
    Returns list of dicts: {protocol, type, value, context, src_ip, dst_ip, src_port, dst_port}
    """
    payload = pkt.get("app_payload", b"")
    if not payload:
        return []

    proto  = pkt.get("proto","")
    src_ip = pkt.get("src_ip","")
    dst_ip = pkt.get("dst_ip","")
    sp     = pkt.get("src_port", 0)
    dp     = pkt.get("dst_port", 0)
    found  = []

    try:
        text = payload.decode("utf-8", errors="replace")
    except Exception:
        text = ""

    def hit(typ, value, context=""):
        found.append(dict(protocol=proto, type=typ,
                          value=str(value)[:250], context=str(context)[:350],
                          src_ip=src_ip, dst_ip=dst_ip,
                          src_port=sp, dst_port=dp))

    # ── FTP ────────────────────────────────────────────────────────────────
    if proto in ("FTP", "FTP-data"):
        for line in text.splitlines():
            l = line.strip()
            if _re.match(r"(?i)^USER\s+\S+", l):
                hit("FTP Username", l.split(None,1)[-1], l)
            elif _re.match(r"(?i)^PASS(\s+|$)", l):
                hit("FTP Password", l.split(None,1)[-1] if len(l.split()) > 1 else "(empty)", l)

    # ── Telnet ─────────────────────────────────────────────────────────────
    if proto == "Telnet":
        printable = "".join(c for c in text if c.isprintable() or c in "\r\n")
        clean = printable.strip()
        if clean:
            hit("Telnet Keystrokes/Data", clean[:300], f"{src_ip} -> {dst_ip}")

    # ── HTTP ───────────────────────────────────────────────────────────────
    if proto in ("HTTP", "HTTP-alt"):
        lines = text.splitlines()
        req_line = lines[0].strip() if lines else ""

        for line in lines:
            # Basic Auth
            m = _re.match(r"(?i)^Authorization:\s+Basic\s+(\S+)", line)
            if m:
                b64 = m.group(1)
                try:
                    decoded = _b64.b64decode(b64 + "==").decode("utf-8","replace")
                    hit("HTTP Basic Auth (decoded)", decoded, req_line)
                except Exception:
                    hit("HTTP Basic Auth (raw b64)", b64, req_line)

            # Bearer token
            m = _re.match(r"(?i)^Authorization:\s+Bearer\s+(\S+)", line)
            if m:
                hit("HTTP Bearer Token", m.group(1)[:120], req_line)

            # Cookie
            if _re.match(r"(?i)^Cookie:\s+", line):
                hit("HTTP Cookie", line.split(":",1)[-1].strip()[:250], req_line)

            # Proxy-Auth
            if _re.match(r"(?i)^Proxy-Authorization:", line):
                hit("HTTP Proxy Auth", line.split(":",1)[-1].strip(), line.strip())

        # POST body credential fields
        body = text.split("\r\n\r\n", 1)[-1] if "\r\n\r\n" in text else ""
        if body:
            for field, val in _re.findall(
                    r"(?i)(password|passwd|pwd|pass|secret|token|apikey|api_key|"
                    r"auth|credential|session)=([^&\s]{1,120})", body):
                hit("HTTP POST Credential", f"{field}={val}", req_line)
            for field, val in _re.findall(
                    r'(?i)"(password|passwd|pwd|secret|token|api_?key|auth)"\s*:\s*"([^"]{1,120})"', body):
                hit("HTTP JSON Credential", f"{field}: {val}", req_line)

    # ── SMTP ───────────────────────────────────────────────────────────────
    if proto == "SMTP":
        for line in text.splitlines():
            l = line.strip()
            if _re.match(r"(?i)^AUTH\s+(LOGIN|PLAIN|CRAM)", l):
                hit("SMTP Auth Command", l, f"{src_ip} -> {dst_ip}")
            elif _re.match(r"(?i)^(MAIL FROM|RCPT TO):", l):
                hit("SMTP Email Address", l, "")
            elif _re.match(r"^[A-Za-z0-9+/=]{8,}$", l.strip()):
                try:
                    decoded = _b64.b64decode(l + "==").decode("utf-8","replace")
                    if decoded.isprintable() and len(decoded) >= 3:
                        hit("SMTP Auth (b64 decoded)", decoded, f"raw: {l}")
                except Exception:
                    pass

    # ── POP3 ───────────────────────────────────────────────────────────────
    if proto == "POP3":
        for line in text.splitlines():
            l = line.strip()
            if _re.match(r"(?i)^USER\s+\S+", l):
                hit("POP3 Username", l.split(None,1)[-1], l)
            elif _re.match(r"(?i)^PASS\s+", l):
                hit("POP3 Password", l.split(None,1)[-1] if len(l.split()) > 1 else "(empty)", l)

    # ── IMAP ───────────────────────────────────────────────────────────────
    if proto == "IMAP":
        for line in text.splitlines():
            l = line.strip()
            m = _re.search(r"(?i)LOGIN\s+(\S+)\s+(\S+)", l)
            if m:
                hit("IMAP Login", f"user={m.group(1)}  pass={m.group(2)}", l)

    # ── LDAP simple bind ───────────────────────────────────────────────────
    if proto == "LDAP":
        runs = _re.findall(rb"[\x20-\x7e]{4,}", payload)
        for run in runs:
            s = run.decode("ascii","replace")
            if any(k in s.lower() for k in ("cn=","dc=","ou=","uid=","password","pass")):
                hit("LDAP Bind Data", s, f"{src_ip} -> {dst_ip}")

    # ── SNMP community string ──────────────────────────────────────────────
    if proto == "SNMP":
        runs = _re.findall(rb"[\x20-\x7e]{3,}", payload)
        for run in runs:
            s = run.decode("ascii","replace")
            # Skip obvious non-community OID/version strings
            if s not in ("GET","SET","public","private") and not s.startswith("1.3") and len(s) <= 40:
                hit("SNMP Community String", s, f"{src_ip} -> {dst_ip}:{dp}")
                break

    # ── NetBIOS — decode Level-2 half-ASCII encoding ──────────────────────
    if proto == "NetBIOS":
        # NetBIOS suffix byte meanings (RFC 1001/1002)
        _NB_SUFFIX = {
            0x00: "Workstation",  0x03: "Messenger",
            0x06: "RAS Server",   0x1B: "Domain Master Browser",
            0x1C: "Domain Controllers", 0x1D: "Local Master Browser",
            0x1E: "Browser Election",   0x20: "File Server",
            0x21: "RAS Client",
        }

        def _decode_nbt_encoded(s):
            """
            Decode a Level-2 NetBIOS half-ASCII encoded name (A-P chars).
            Returns (name_str, suffix_description) or (None, None) if not valid.
            Each pair of A-P bytes encodes one byte: ((A-0x41)<<4)|(B-0x41).
            Name is 15 chars + 1 suffix byte = 32 encoded chars (16 pairs).
            """
            b = s if isinstance(s, (bytes, bytearray)) else s.encode('ascii','replace')
            # Must be even length, all bytes 0x41-0x50 (A-P)
            if len(b) < 4 or len(b) % 2 != 0:
                return None, None
            if not all(0x41 <= x <= 0x50 for x in b[:32]):
                return None, None
            chars = []
            for i in range(0, min(len(b), 32), 2):
                c = ((b[i] - 0x41) << 4) | (b[i+1] - 0x41)
                chars.append(c)
            if len(chars) >= 16:
                suffix = chars[15]
                name = ''.join(chr(c) for c in chars[:15]
                               if 0x20 <= c < 0x7f).rstrip()
                suf_desc = _NB_SUFFIX.get(suffix, f"type-{suffix:#04x}")
            else:
                suffix = None
                name = ''.join(chr(c) for c in chars
                               if 0x20 <= c < 0x7f).rstrip()
                suf_desc = None
            # Filter out all-space, wildcard, and empty names
            if not name or name in ("*", "__") or name.isspace():
                return None, None
            return name, suf_desc

        seen_names = set()
        # Try NBNS wire-format decode first (UDP 137 registrations/responses)
        decoded_names = _decode_nbns_payload(payload)
        for name in decoded_names:
            if name not in seen_names:
                seen_names.add(name)
                hit("NetBIOS Name", name, f"{src_ip} -> {dst_ip}")

        # Scan payload for 32-char (full) or shorter A-P runs = NBT session names
        # These appear in NBT Session Request packets (TCP 139) and SMB negotiations
        for run in _re.findall(rb"[A-Pa-p]{16,34}", payload):
            run_s = run.decode('ascii').upper()
            # Only attempt if run length is plausibly a name (even length, 16-32 chars)
            if len(run_s) % 2 != 0:
                run_s = run_s[:len(run_s)-1]  # trim to even
            name, suf = _decode_nbt_encoded(run_s)
            if name and name not in seen_names and name not in ("WORKGROUP",):
                seen_names.add(name)
                ctx = f"{src_ip} -> {dst_ip}"
                if suf:
                    ctx += f" [{suf}]"
                hit("NetBIOS Hostname", name, ctx)

        # Also capture workgroup / domain names (they ARE useful intel)
        for run in _re.findall(rb"[A-Pa-p]{16,34}", payload):
            run_s = run.decode('ascii').upper()
            if len(run_s) % 2 != 0:
                run_s = run_s[:len(run_s)-1]
            name, suf = _decode_nbt_encoded(run_s)
            if name == "WORKGROUP" and name not in seen_names:
                seen_names.add(name)
                hit("NetBIOS Domain/Workgroup", name, f"{src_ip} -> {dst_ip}")

        # If nothing decoded, capture meaningful SMB strings (not raw encoded garbage)
        if not seen_names:
            # Only grab runs that look like real text (contain letters + digits/symbols)
            # and are NOT all A-P chars (which would be undecodeable garbage)
            for run in _re.findall(rb"[\x20-\x7e]{6,}", payload):
                s_str = run.decode("ascii", "replace").strip()
                if (s_str
                        and not all(0x41 <= ord(c) <= 0x50 for c in s_str if c.isalpha())
                        and any(c.isalnum() for c in s_str)
                        and "SMB" not in s_str[:4]):
                    hit("NetBIOS Session Data", s_str[:120],
                        f"{src_ip} -> {dst_ip}")
                    break  # one is enough

    # ── Generic API key / secret patterns (any cleartext protocol) ─────────
    if proto in CLEARTEXT_PROTOS and text:
        patterns = [
            (r"(?i)(api[_-]?key|apikey|x-api-key)[\s:=]+([A-Za-z0-9_\-]{16,80})", "API Key"),
            (r"(?i)(access_token|auth_token)[\s:=]+([A-Za-z0-9_.\-]{16,150})",     "Auth Token"),
            (r"(?i)(secret|private_key)[\s:=]+([A-Za-z0-9_.\-/+]{16,80})",         "Secret/Key"),
            (r"(AKIA[0-9A-Z]{16})",                                                   "AWS Access Key"),
            (r"(-----BEGIN (?:RSA |EC |OPENSSH )?PRIVATE KEY-----)",                  "Private Key"),
        ]
        for pattern, label in patterns:
            for m in _re.findall(pattern, text):
                val = m[-1] if isinstance(m, tuple) else m
                hit(label, val, f"{proto} {src_ip} -> {dst_ip}")

    return found




# ─────────────────────────────────────────────────────────────────────────────
# Banner / resource / version-string extractor
# ─────────────────────────────────────────────────────────────────────────────

def extract_banners(packets):
    """
    Single pass over all packets extracting:

    Banners / version strings
      • HTTP Server:, X-Powered-By:, X-Generator:, Via: response headers
      • FTP 220 greeting line
      • SMTP 220 greeting line
      • SSH version string (SSH-2.0-OpenSSH_8.2p1 …)
      • Any printable version pattern in Telnet streams

    Commonly requested resources
      • HTTP GET/POST/PUT/DELETE request lines + Host header
      • DNS queries (what domains the network is looking up)
      • HTTP User-Agent strings (reveals client OS/browser/software versions)
      • NTP server IPs
      • DHCP requested server IPs

    Returns list of dicts:
      { category, server_ip, client_ip, port, protocol,
        banner_type, value, raw_context }
    """
    hits = []
    seen = set()   # deduplicate

    def hit(category, server_ip, client_ip, port, protocol, banner_type, value, context=""):
        value = str(value).strip()[:300]
        if not value:
            return
        key = (category, server_ip, banner_type, value)
        if key in seen:
            return
        seen.add(key)
        hits.append(dict(
            category    = category,
            server_ip   = server_ip,
            client_ip   = client_ip,
            port        = int(port) if port else 0,
            protocol    = protocol,
            banner_type = banner_type,
            value       = value,
            context     = str(context).strip()[:200],
        ))

    for p in packets:
        proto   = p.get("proto", "")
        payload = p.get("app_payload", b"")
        src_ip  = p.get("src_ip", "")
        dst_ip  = p.get("dst_ip", "")
        sp      = p.get("src_port", 0)
        dp      = p.get("dst_port", 0)

        if not payload:
            continue

        try:
            text = payload.decode("utf-8", errors="replace")
        except Exception:
            text = ""

        lines = text.splitlines()
        first = lines[0].strip() if lines else ""

        # ── HTTP responses (server is the src when dp is ephemeral / sp is 80/443) ──
        if proto in ("HTTP", "HTTP-alt"):
            # Determine direction: response starts with HTTP/1.x
            if first.startswith("HTTP/"):
                server_ip, client_ip, port = src_ip, dst_ip, sp
                for line in lines:
                    l = line.strip()
                    low = l.lower()
                    if low.startswith("server:"):
                        val = l.split(":",1)[1].strip()
                        hit("Banner", server_ip, client_ip, dp or sp, proto,
                            "HTTP Server Header", val, first)
                    elif low.startswith("x-powered-by:"):
                        val = l.split(":",1)[1].strip()
                        hit("Banner", server_ip, client_ip, dp or sp, proto,
                            "X-Powered-By", val, first)
                    elif low.startswith("x-generator:"):
                        val = l.split(":",1)[1].strip()
                        hit("Banner", server_ip, client_ip, dp or sp, proto,
                            "X-Generator", val, first)
                    elif low.startswith("via:"):
                        val = l.split(":",1)[1].strip()
                        hit("Banner", server_ip, client_ip, dp or sp, proto,
                            "Via (Proxy)", val, first)
                    elif low.startswith("x-aspnet-version:") or low.startswith("x-aspnetmvc-version:"):
                        val = l.split(":",1)[1].strip()
                        hit("Banner", server_ip, client_ip, dp or sp, proto,
                            "ASP.NET Version", val, first)
                    elif low.startswith("x-runtime:"):
                        val = l.split(":",1)[1].strip()
                        hit("Banner", server_ip, client_ip, dp or sp, proto,
                            "Runtime Version", val, first)

            # Requests — resource enumeration
            elif any(first.startswith(m) for m in ("GET ","POST ","PUT ","DELETE ","HEAD ","OPTIONS ","PATCH ")):
                server_ip, client_ip, port = dst_ip, src_ip, dp
                method_path = first.rsplit(" HTTP/",1)[0] if " HTTP/" in first else first
                # Extract host for full URL reconstruction
                host = ""
                user_agent = ""
                for line in lines:
                    l = line.strip()
                    if l.lower().startswith("host:"):
                        host = l.split(":",1)[1].strip()
                    elif l.lower().startswith("user-agent:"):
                        user_agent = l.split(":",1)[1].strip()
                full_resource = f"http://{host}{method_path.split(' ',1)[-1]}" if host else method_path
                hit("Resource", server_ip, client_ip, dp, proto,
                    "HTTP Request", full_resource, f"Method: {method_path.split()[0]}")
                if user_agent:
                    hit("Client Software", client_ip, server_ip, dp, proto,
                        "HTTP User-Agent", user_agent, f"→ {server_ip}:{dp}")

        # ── FTP banners ────────────────────────────────────────────────────────
        elif proto in ("FTP", "FTP-data"):
            # 220 = service ready greeting (the banner)
            for line in lines:
                l = line.strip()
                if l.startswith("220"):
                    hit("Banner", src_ip, dst_ip, sp, proto,
                        "FTP Banner", l[3:].strip() or l, "FTP 220 greeting")
                elif l.startswith("215"):  # SYST response
                    hit("Banner", src_ip, dst_ip, sp, proto,
                        "FTP System Type", l[3:].strip(), "SYST response")

        # ── SMTP banners ────────────────────────────────────────────────────────
        elif proto == "SMTP":
            for line in lines:
                l = line.strip()
                if l.startswith("220"):
                    hit("Banner", src_ip, dst_ip, sp, proto,
                        "SMTP Banner", l[3:].strip() or l, "SMTP 220 greeting")
                elif l.upper().startswith("EHLO") or l.upper().startswith("HELO"):
                    domain = l.split(None,1)[-1] if len(l.split()) > 1 else ""
                    if domain:
                        hit("Resource", dst_ip, src_ip, dp, proto,
                            "SMTP EHLO Domain", domain, "Client announced domain")

        # ── SSH version string ──────────────────────────────────────────────────
        # SSH banner is sent in cleartext before encryption negotiation
        elif proto in ("SSH", "TCP"):
            if text.startswith("SSH-"):
                banner_line = first.strip()
                hit("Banner", src_ip, dst_ip, sp or dp, "SSH",
                    "SSH Version String", banner_line, f"{src_ip}→{dst_ip}")

        # ── Telnet — grab any version/banner patterns ──────────────────────────
        elif proto == "Telnet":
            printable = "".join(c for c in text if c.isprintable() or c in "\r\n\t")
            # Look for version patterns
            for m in _re.findall(
                    r"(?i)(version\s+[\d.]+|v[\d]+\.[\d]+[\.\d]*|release\s+[\d.]+)", printable):
                hit("Banner", src_ip, dst_ip, sp or dp, proto,
                    "Telnet Version String", m if isinstance(m,str) else m[0],
                    printable[:100])
            # Any obvious login/welcome banner lines
            for line in printable.splitlines():
                l = line.strip()
                if any(kw in l.lower() for kw in ("welcome","unauthorized","login banner","authorized users only","cisco","juniper","warning:")):
                    hit("Banner", src_ip, dst_ip, sp or dp, proto,
                        "Telnet Login Banner", l[:200], "")

        # ── DNS queries — what the network is resolving ────────────────────────
        elif proto == "DNS" and payload:
            try:
                if len(payload) >= 12:
                    flags    = struct.unpack(">H", payload[2:4])[0]
                    is_query = not ((flags >> 15) & 1)
                    qd_count = struct.unpack(">H", payload[4:6])[0]
                    if is_query and qd_count > 0:
                        offset = 12
                        qname, _ = _dns_read_name(payload, offset)
                        if qname and "." in qname:
                            hit("Resource", dst_ip, src_ip, dp, proto,
                                "DNS Query", qname, f"Queried by {src_ip}")
            except Exception:
                pass

        # ── SNMP — system description OID (sysDescr) ──────────────────────────
        elif proto == "SNMP":
            # sysDescr (1.3.6.1.2.1.1.1.0) responses often contain OS/device strings
            printable_runs = _re.findall(rb"[ -~]{8,}", payload)
            for run in printable_runs:
                s = run.decode("ascii", "replace")
                # Filter to version-like strings
                if any(kw in s.lower() for kw in ("linux","windows","cisco","juniper","version","release","snmp","net-snmp")):
                    hit("Banner", src_ip, dst_ip, sp or dp, proto,
                        "SNMP sysDescr", s[:150], f"{src_ip}→{dst_ip}")
                    break

        # ── NTP — server IPs (what time servers is the network using) ──────────
        elif proto == "NTP":
            hit("Resource", dst_ip, src_ip, dp, proto,
                "NTP Server", dst_ip, f"NTP query from {src_ip}")

        # ── DHCP — server identifier and requested options ─────────────────────
        elif proto == "DHCP" and payload and len(payload) >= 240:
            try:
                siaddr = socket.inet_ntoa(payload[20:24])  # server IP
                i = 240
                while i < len(payload):
                    opt = payload[i]; i += 1
                    if opt == 255: break
                    if opt == 0:  continue
                    if i >= len(payload): break
                    ln = payload[i]; i += 1
                    val = payload[i:i+ln]; i += ln
                    if opt == 54 and ln == 4:  # DHCP Server Identifier
                        dhcp_srv = socket.inet_ntoa(val)
                        hit("Resource", dhcp_srv, src_ip, 67, proto,
                            "DHCP Server", dhcp_srv, f"DHCP server for {src_ip}")
                    elif opt == 60:  # Vendor Class Identifier
                        vci = val.decode("ascii","replace").strip("\x00")
                        hit("Banner", src_ip, dst_ip, dp, proto,
                            "DHCP Vendor Class", vci, f"Client: {src_ip}")
            except Exception:
                pass

    return hits


# ─────────────────────────────────────────────────────────────────────────────
# TLS / SSL handshake parser and session extractor
# Parses ClientHello, ServerHello, Certificate, and Alert records from raw TCP
# payloads — all in the clear before encryption begins.
# ─────────────────────────────────────────────────────────────────────────────

import datetime as _dt

import struct, socket, datetime
from collections import defaultdict

# ─────────────────────────────────────────────────────────────────────────────
# TLS constants
# ─────────────────────────────────────────────────────────────────────────────

TLS_VERSIONS = {
    0x0300: "SSLv3",
    0x0301: "TLS 1.0",
    0x0302: "TLS 1.1",
    0x0303: "TLS 1.2",
    0x0304: "TLS 1.3",
}

# Cipher suites — subset covering weak/interesting ones + common ones
CIPHER_SUITES = {
    0x0000: "TLS_NULL_WITH_NULL_NULL",
    0x0001: "TLS_RSA_WITH_NULL_MD5",
    0x0002: "TLS_RSA_WITH_NULL_SHA",
    0x0004: "TLS_RSA_WITH_RC4_128_MD5",
    0x0005: "TLS_RSA_WITH_RC4_128_SHA",
    0x000A: "TLS_RSA_WITH_3DES_EDE_CBC_SHA",
    0x002F: "TLS_RSA_WITH_AES_128_CBC_SHA",
    0x0033: "TLS_DHE_RSA_WITH_AES_128_CBC_SHA",
    0x0035: "TLS_RSA_WITH_AES_256_CBC_SHA",
    0x0039: "TLS_DHE_RSA_WITH_AES_256_CBC_SHA",
    0x003C: "TLS_RSA_WITH_AES_128_CBC_SHA256",
    0x003D: "TLS_RSA_WITH_AES_256_CBC_SHA256",
    0x009C: "TLS_RSA_WITH_AES_128_GCM_SHA256",
    0x009D: "TLS_RSA_WITH_AES_256_GCM_SHA384",
    0x009E: "TLS_DHE_RSA_WITH_AES_128_GCM_SHA256",
    0x009F: "TLS_DHE_RSA_WITH_AES_256_GCM_SHA384",
    0xC009: "TLS_ECDHE_ECDSA_WITH_AES_128_CBC_SHA",
    0xC00A: "TLS_ECDHE_ECDSA_WITH_AES_256_CBC_SHA",
    0xC013: "TLS_ECDHE_RSA_WITH_AES_128_CBC_SHA",
    0xC014: "TLS_ECDHE_RSA_WITH_AES_256_CBC_SHA",
    0xC02B: "TLS_ECDHE_ECDSA_WITH_AES_128_GCM_SHA256",
    0xC02C: "TLS_ECDHE_ECDSA_WITH_AES_256_GCM_SHA384",
    0xC02F: "TLS_ECDHE_RSA_WITH_AES_128_GCM_SHA256",
    0xC030: "TLS_ECDHE_RSA_WITH_AES_256_GCM_SHA384",
    0xCCA8: "TLS_ECDHE_RSA_WITH_CHACHA20_POLY1305_SHA256",
    0xCCA9: "TLS_ECDHE_ECDSA_WITH_CHACHA20_POLY1305_SHA256",
    0xCCAA: "TLS_DHE_RSA_WITH_CHACHA20_POLY1305_SHA256",
    # TLS 1.3 suites
    0x1301: "TLS_AES_128_GCM_SHA256",
    0x1302: "TLS_AES_256_GCM_SHA384",
    0x1303: "TLS_CHACHA20_POLY1305_SHA256",
    # SCSV
    0x00FF: "TLS_EMPTY_RENEGOTIATION_INFO_SCSV",
    0x5600: "TLS_FALLBACK_SCSV",
}

WEAK_CIPHERS = {
    "NULL", "RC4", "DES", "EXPORT", "anon", "MD5",
    "RC2", "IDEA", "SEED", "CAMELLIA_128_CBC", "3DES"
}

ALERT_DESCS = {
    0:"close_notify", 10:"unexpected_message", 20:"bad_record_mac",
    21:"decryption_failed", 22:"record_overflow", 30:"decompression_failure",
    40:"handshake_failure", 41:"no_certificate", 42:"bad_certificate",
    43:"unsupported_certificate", 44:"certificate_revoked",
    45:"certificate_expired", 46:"certificate_unknown",
    47:"illegal_parameter", 48:"unknown_ca", 49:"access_denied",
    50:"decode_error", 51:"decrypt_error", 60:"export_restriction",
    70:"protocol_version", 71:"insufficient_security", 80:"internal_error",
    86:"inappropriate_fallback", 90:"user_canceled",
    100:"no_renegotiation", 110:"unsupported_extension",
    112:"unrecognized_name", 113:"bad_certificate_status_response",
    115:"unknown_psk_identity", 116:"certificate_required", 120:"no_application_protocol",
}

TLS_HS_TYPES = {
    0:"HelloRequest", 1:"ClientHello", 2:"ServerHello",
    4:"NewSessionTicket", 8:"EncryptedExtensions",
    11:"Certificate", 12:"ServerKeyExchange",
    13:"CertificateRequest", 14:"ServerHelloDone",
    15:"CertificateVerify", 16:"ClientKeyExchange", 20:"Finished",
}

# Extension types
EXT_SNI       = 0x0000
EXT_ALPN      = 0x0010
EXT_SUPPORTED = 0x002B   # supported_versions (TLS 1.3)
EXT_SIG_ALGS  = 0x000D

# ─────────────────────────────────────────────────────────────────────────────
# ASN.1 / X.509 minimal parser (no external libs)
# ─────────────────────────────────────────────────────────────────────────────

def _asn1_read_len(data, off):
    """Read ASN.1 length at offset. Returns (length, new_offset)."""
    if off >= len(data):
        return 0, off
    b = data[off]; off += 1
    if b < 0x80:
        return b, off
    n = b & 0x7F
    if off + n > len(data):
        return 0, off + n
    ln = int.from_bytes(data[off:off+n], "big")
    return ln, off + n

def _asn1_next(data, off):
    """Read next ASN.1 TLV. Returns (tag, value_bytes, next_offset)."""
    if off + 2 > len(data):
        return 0, b"", off
    tag = data[off]; off += 1
    ln, off = _asn1_read_len(data, off)
    end = off + ln
    val = data[off:end]
    return tag, val, end

def _asn1_seq_children(data):
    """Iterate children of an ASN.1 SEQUENCE/SET body."""
    off = 0
    while off < len(data):
        tag, val, off = _asn1_next(data, off)
        if tag == 0:
            break
        yield tag, val

def _oid_to_str(data):
    """Decode ASN.1 OID bytes to dotted string."""
    if not data:
        return ""
    try:
        oid = [data[0] // 40, data[0] % 40]
        i, cur = 1, 0
        while i < len(data):
            b = data[i]; i += 1
            cur = (cur << 7) | (b & 0x7F)
            if not (b & 0x80):
                oid.append(cur); cur = 0
        return ".".join(str(x) for x in oid)
    except Exception:
        return ""

# Common OID → friendly name
OID_NAMES = {
    "2.5.4.3":  "CN", "2.5.4.6":  "C", "2.5.4.7":  "L",
    "2.5.4.8":  "ST","2.5.4.10": "O", "2.5.4.11": "OU",
    "2.5.4.12": "title", "2.5.29.17": "SAN",
    "1.2.840.113549.1.1.1":  "rsaEncryption",
    "1.2.840.113549.1.1.11": "sha256WithRSAEncryption",
    "1.2.840.10040.4.1": "dsa",
    "1.2.840.10045.2.1": "ecPublicKey",
    "1.3.132.0.34": "secp384r1",
    "1.3.132.0.35": "secp521r1",
    "1.2.840.10045.3.1.7": "secp256r1 (P-256)",
    "1.2.840.10045.3.1.34": "secp384r1 (P-384)",
    "2.5.29.19": "basicConstraints",
    "2.5.29.15": "keyUsage",
    "2.5.29.37": "extKeyUsage",
}

def _parse_dn(data):
    """Parse X.509 Distinguished Name, return dict of RDN components."""
    result = {}
    for tag, rdn_set in _asn1_seq_children(data):
        for _, atv in _asn1_seq_children(rdn_set):
            tag2, oid_bytes, rest = _asn1_next(atv, 0)
            oid = _oid_to_str(oid_bytes)
            name = OID_NAMES.get(oid, oid)
            tag3, val_bytes, _ = _asn1_next(atv, rest - len(atv) + len(oid_bytes) + 2)
            # Reparse properly
            off2 = 0
            _, oid_raw, off2 = _asn1_next(atv, 0)
            oid_str = _oid_to_str(oid_raw)
            fname   = OID_NAMES.get(oid_str, oid_str)
            _, str_raw, _ = _asn1_next(atv, off2)
            try:
                val = str_raw.decode("utf-8", errors="replace").strip()
            except Exception:
                val = str_raw.hex()
            result[fname] = val
    return result

def _parse_utctime(data):
    """Parse ASN.1 UTCTime or GeneralizedTime → datetime or None."""
    try:
        s = data.decode("ascii").strip("\x00")
        if len(s) == 13:   # YYMMDDHHmmssZ
            return datetime.datetime.strptime(s, "%y%m%d%H%M%SZ")
        elif len(s) == 15: # YYYYMMDDHHmmssZ
            return datetime.datetime.strptime(s, "%Y%m%d%H%M%SZ")
    except Exception:
        pass
    return None

def _parse_validity(data):
    """Parse Validity SEQUENCE → (not_before, not_after) as datetime."""
    nb = na = None
    off = 0
    for _ in range(2):
        tag, val, off = _asn1_next(data, off)
        dt = _parse_utctime(val)
        if nb is None: nb = dt
        else:          na = dt
    return nb, na

def _parse_pubkey_info(data):
    """Parse SubjectPublicKeyInfo → (key_type_str, key_size_bits)."""
    try:
        off = 0
        tag, alg_seq, off = _asn1_next(data, 0)   # AlgorithmIdentifier SEQUENCE
        _, oid_bytes, _ = _asn1_next(alg_seq, 0)
        oid = _oid_to_str(oid_bytes)
        alg_name = OID_NAMES.get(oid, oid)

        tag2, bitstring, off2 = _asn1_next(data, off)   # BIT STRING
        key_bytes = bitstring[1:]   # strip unused-bits byte

        if "rsa" in alg_name.lower():
            # RSA key is a SEQUENCE of (n, e)
            _, rsa_seq, _ = _asn1_next(key_bytes, 0)
            _, n_bytes, _ = _asn1_next(rsa_seq, 0)
            key_size = (len(n_bytes) - (1 if n_bytes[0] == 0 else 0)) * 8
            return f"RSA", key_size
        elif "ec" in alg_name.lower():
            # ECDSA — size from curve OID
            _, params, _ = _asn1_next(alg_seq, len(oid_bytes) + 2)
            curve_oid = _oid_to_str(params) if params else ""
            curve_name = OID_NAMES.get(curve_oid, curve_oid)
            bits = 256 if "256" in curve_name else (384 if "384" in curve_name else (521 if "521" in curve_name else 0))
            return f"ECDSA ({curve_name})", bits
        return alg_name, 0
    except Exception:
        return "Unknown", 0

def _parse_extensions(data):
    """Parse certificate extensions body, extract SAN."""
    sans = []
    try:
        off = 0
        while off < len(data):
            tag, ext_seq, off = _asn1_next(data, off)
            if tag == 0: break
            eoff = 0
            _, oid_bytes, eoff = _asn1_next(ext_seq, eoff)
            oid = _oid_to_str(oid_bytes)
            if oid == "2.5.29.17":   # SAN
                # next may be critical bool, then octet string
                tag2, val2, eoff2 = _asn1_next(ext_seq, eoff)
                if tag2 == 0x01:     # boolean (critical flag)
                    tag2, val2, eoff2 = _asn1_next(ext_seq, eoff2)
                # val2 is OCTET STRING containing the SAN SEQUENCE
                _, san_seq, _ = _asn1_next(val2, 0)
                soff = 0
                while soff < len(san_seq):
                    stag, sval, soff = _asn1_next(san_seq, soff)
                    if stag == 0x82:   # dNSName
                        try: sans.append(sval.decode("ascii","replace"))
                        except: pass
                    elif stag == 0x87 and len(sval) == 4:   # iPAddress v4
                        try: sans.append(socket.inet_ntoa(sval))
                        except: pass
    except Exception:
        pass
    return sans

def _parse_certificate(cert_bytes):
    """
    Parse a DER-encoded X.509 certificate.
    Returns dict with subject, issuer, sans, validity, key_type, key_bits.
    """
    result = {
        "subject": "", "issuer": "", "sans": [],
        "not_before": None, "not_after": None,
        "key_type": "Unknown", "key_bits": 0,
        "serial": "",
    }
    try:
        tag, tbs_outer, _ = _asn1_next(cert_bytes, 0)   # Certificate SEQUENCE
        if tag != 0x30:
            return result
        off = 0
        # TBSCertificate
        tag2, tbs, off = _asn1_next(tbs_outer, off)
        # Parse TBSCertificate fields
        tbs_off = 0
        # Optional version [0] EXPLICIT
        tag3, v0, tbs_off = _asn1_next(tbs, tbs_off)
        if tag3 == 0xA0:   # version
            tag3, v0, tbs_off = _asn1_next(tbs, tbs_off)
        # serialNumber INTEGER
        if tag3 == 0x02:
            result["serial"] = v0.hex()[:40]
            tag3, v0, tbs_off = _asn1_next(tbs, tbs_off)
        # signature AlgorithmIdentifier
        tag3, v0, tbs_off = _asn1_next(tbs, tbs_off)
        # issuer Name
        tag3, issuer_bytes, tbs_off = _asn1_next(tbs, tbs_off)
        issuer_dn = _parse_dn(issuer_bytes)
        result["issuer"] = issuer_dn.get("CN","") or issuer_dn.get("O","") or str(issuer_dn)
        # validity
        tag3, validity_bytes, tbs_off = _asn1_next(tbs, tbs_off)
        nb, na = _parse_validity(validity_bytes)
        result["not_before"] = nb
        result["not_after"]  = na
        # subject Name
        tag3, subject_bytes, tbs_off = _asn1_next(tbs, tbs_off)
        subject_dn = _parse_dn(subject_bytes)
        result["subject"] = subject_dn.get("CN","") or subject_dn.get("O","") or str(subject_dn)
        result["subject_dn"] = subject_dn
        # SubjectPublicKeyInfo
        tag3, spki_bytes, tbs_off = _asn1_next(tbs, tbs_off)
        kt, kb = _parse_pubkey_info(spki_bytes)
        result["key_type"] = kt
        result["key_bits"] = kb
        # Extensions (optional, [3])
        while tbs_off < len(tbs):
            tag3, ext_outer, tbs_off = _asn1_next(tbs, tbs_off)
            if tag3 == 0xA3:    # [3] extensions
                tag4, exts_seq, _ = _asn1_next(ext_outer, 0)
                sans = _parse_extensions(exts_seq)
                result["sans"] = sans
    except Exception:
        pass
    return result

# ─────────────────────────────────────────────────────────────────────────────
# TLS record parser
# ─────────────────────────────────────────────────────────────────────────────

def _parse_tls_records(payload):
    """
    Yield TLS records from a raw TCP app_payload.
    Each record: (content_type, version, data_bytes)
    content_type: 20=ChangeCipherSpec, 21=Alert, 22=Handshake, 23=AppData
    """
    off = 0
    while off + 5 <= len(payload):
        ct  = payload[off]
        ver = struct.unpack(">H", payload[off+1:off+3])[0]
        ln  = struct.unpack(">H", payload[off+3:off+5])[0]
        off += 5
        if ct not in (20, 21, 22, 23):
            break   # not TLS
        if off + ln > len(payload):
            break
        yield ct, ver, payload[off:off+ln]
        off += ln

def _parse_client_hello(data):
    """Parse ClientHello handshake body. Returns dict of extracted fields."""
    result = {
        "client_version": 0, "session_id": "",
        "cipher_suites": [], "sni": "", "alpn": [], "tls13_offered": False,
    }
    try:
        off = 0
        result["client_version"] = struct.unpack(">H", data[off:off+2])[0]; off += 2
        off += 32   # random
        sid_len = data[off]; off += 1
        result["session_id"] = data[off:off+sid_len].hex(); off += sid_len
        cs_len = struct.unpack(">H", data[off:off+2])[0]; off += 2
        for i in range(0, cs_len, 2):
            cs = struct.unpack(">H", data[off+i:off+i+2])[0]
            result["cipher_suites"].append(cs)
        off += cs_len
        comp_len = data[off]; off += 1
        off += comp_len   # compression methods
        if off + 2 > len(data):
            return result
        ext_total = struct.unpack(">H", data[off:off+2])[0]; off += 2
        ext_end = off + ext_total
        while off + 4 <= ext_end:
            ext_type = struct.unpack(">H", data[off:off+2])[0]
            ext_len  = struct.unpack(">H", data[off+2:off+4])[0]
            ext_data = data[off+4:off+4+ext_len]; off += 4 + ext_len
            if ext_type == EXT_SNI and len(ext_data) >= 5:
                sni_list_len = struct.unpack(">H", ext_data[0:2])[0]
                sni_type = ext_data[2]
                sni_name_len = struct.unpack(">H", ext_data[3:5])[0]
                result["sni"] = ext_data[5:5+sni_name_len].decode("ascii","replace")
            elif ext_type == EXT_ALPN and len(ext_data) >= 4:
                proto_list_len = struct.unpack(">H", ext_data[0:2])[0]
                poff = 2
                while poff < 2 + proto_list_len:
                    plen = ext_data[poff]; poff += 1
                    result["alpn"].append(ext_data[poff:poff+plen].decode("ascii","replace"))
                    poff += plen
            elif ext_type == EXT_SUPPORTED:
                # supported_versions — if 0x0304 present, TLS 1.3 is offered
                vlist_len = ext_data[0] if ext_data else 0
                for i in range(1, 1+vlist_len, 2):
                    if ext_data[i:i+2] == b"\x03\x04":
                        result["tls13_offered"] = True
    except Exception:
        pass
    return result

def _parse_server_hello(data):
    """Parse ServerHello handshake body."""
    result = {"server_version": 0, "cipher_suite": 0, "session_id": "", "tls13_negotiated": False}
    try:
        off = 0
        result["server_version"] = struct.unpack(">H", data[off:off+2])[0]; off += 2
        off += 32   # random
        sid_len = data[off]; off += 1
        result["session_id"] = data[off:off+sid_len].hex(); off += sid_len
        result["cipher_suite"] = struct.unpack(">H", data[off:off+2])[0]; off += 2
        off += 1   # compression method
        if off + 2 <= len(data):
            ext_total = struct.unpack(">H", data[off:off+2])[0]; off += 2
            ext_end = off + ext_total
            while off + 4 <= ext_end:
                ext_type = struct.unpack(">H", data[off:off+2])[0]
                ext_len  = struct.unpack(">H", data[off+2:off+4])[0]
                ext_data = data[off+4:off+4+ext_len]; off += 4 + ext_len
                if ext_type == EXT_SUPPORTED and len(ext_data) == 2:
                    ver = struct.unpack(">H", ext_data)[0]
                    if ver == 0x0304:
                        result["tls13_negotiated"] = True
                        result["server_version"] = 0x0304
    except Exception:
        pass
    return result

def _parse_certificates(data):
    """Parse Certificate handshake body. Returns list of parsed cert dicts."""
    certs = []
    try:
        off = 0
        total_len = struct.unpack(">I", b"\x00" + data[0:3])[0]; off += 3
        end = min(off + total_len, len(data))
        while off + 3 <= end:
            cert_len = struct.unpack(">I", b"\x00" + data[off:off+3])[0]; off += 3
            cert_bytes = data[off:off+cert_len]; off += cert_len
            parsed = _parse_certificate(cert_bytes)
            certs.append(parsed)
    except Exception:
        pass
    return certs

# ─────────────────────────────────────────────────────────────────────────────
# Session correlator — main entry point
# ─────────────────────────────────────────────────────────────────────────────

def extract_tls_sessions(packets):
    """
    Walk all packets, find TLS handshakes, parse them, and return
    a list of per-session dicts suitable for the Excel sheet.

    Session key: (client_ip, server_ip, server_port)
    — we use server_port because the client port changes per connection
      but the server port identifies the service.
    """
    # Accumulate handshake fragments per TCP stream
    # key = (src_ip, dst_ip, sport, dport)  (canonical: lower IP first)
    streams = defaultdict(bytearray)   # raw payload accumulation
    sessions = {}    # (client_ip, server_ip, srv_port) -> session dict

    def _skey(p):
        """Canonical stream key: always client→server direction."""
        return (p["src_ip"], p["dst_ip"], p["src_port"], p["dst_port"])

    TLS_PORTS = {443, 8443, 4443, 993, 995, 465, 587, 636, 3269, 8883, 5671, 5672}

    for p in packets:
        proto = p.get("proto","")
        payload = p.get("app_payload", b"")
        if not payload or len(payload) < 6:
            continue

        sp, dp = p.get("src_port",0), p.get("dst_port",0)
        src, dst = p.get("src_ip",""), p.get("dst_ip","")

        # Only process TCP on TLS ports, OR packets that look like TLS
        is_tls_port = sp in TLS_PORTS or dp in TLS_PORTS
        looks_like_tls = (len(payload) >= 5 and
                          payload[0] in (20,21,22,23) and
                          payload[1] == 0x03 and
                          payload[2] in (0x00,0x01,0x02,0x03,0x04))
        if not (is_tls_port or looks_like_tls):
            continue

        # Determine session key: client is whoever initiated (higher port usually)
        if dp in TLS_PORTS or dp < sp:
            client_ip, server_ip, srv_port = src, dst, dp
        else:
            client_ip, server_ip, srv_port = dst, src, sp

        sess_key = (client_ip, server_ip, srv_port)
        if sess_key not in sessions:
            sessions[sess_key] = {
                "client_ip": client_ip,
                "server_ip": server_ip,
                "server_port": srv_port,
                "sni": "",
                "client_version_offered": "",
                "tls_version": "",
                "cipher_suite_id": 0,
                "cipher_suite": "",
                "alpn": "",
                "cert_subject": "",
                "cert_issuer": "",
                "cert_sans": "",
                "cert_not_before": "",
                "cert_not_after": "",
                "cert_key_type": "",
                "cert_key_bits": 0,
                "cert_expired": False,
                "cert_expiring_soon": False,
                "weak_cipher": False,
                "weak_version": False,
                "alerts": [],
                "handshake_complete": False,
                "issues": [],
            }
        sess = sessions[sess_key]

        # Parse TLS records from this packet's payload
        try:
            for ct, ver, rec_data in _parse_tls_records(payload):
                if ct == 22:   # Handshake
                    hs_off = 0
                    while hs_off + 4 <= len(rec_data):
                        hs_type = rec_data[hs_off]
                        hs_len  = struct.unpack(">I", b"\x00" + rec_data[hs_off+1:hs_off+4])[0]
                        hs_data = rec_data[hs_off+4:hs_off+4+hs_len]
                        hs_off += 4 + hs_len

                        if hs_type == 1:   # ClientHello
                            ch = _parse_client_hello(hs_data)
                            if not sess["sni"]:
                                sess["sni"] = ch.get("sni","")
                            if not sess["client_version_offered"]:
                                cv = ch.get("client_version",0)
                                if ch.get("tls13_offered"):
                                    sess["client_version_offered"] = "TLS 1.3 (offered)"
                                else:
                                    sess["client_version_offered"] = TLS_VERSIONS.get(cv, f"0x{cv:04x}")
                            if not sess["alpn"]:
                                sess["alpn"] = ", ".join(ch.get("alpn",[]))

                        elif hs_type == 2:   # ServerHello
                            sh = _parse_server_hello(hs_data)
                            sv = sh.get("server_version",0)
                            if sh.get("tls13_negotiated"):
                                sess["tls_version"] = "TLS 1.3"
                            else:
                                sess["tls_version"] = TLS_VERSIONS.get(sv, f"0x{sv:04x}")
                            cs_id = sh.get("cipher_suite",0)
                            sess["cipher_suite_id"] = cs_id
                            sess["cipher_suite"] = CIPHER_SUITES.get(cs_id, f"0x{cs_id:04x}")

                        elif hs_type == 11 and not sess["cert_subject"]:   # Certificate
                            certs = _parse_certificates(hs_data)
                            if certs:
                                leaf = certs[0]
                                sess["cert_subject"]   = leaf.get("subject","")
                                sess["cert_issuer"]    = leaf.get("issuer","")
                                sess["cert_sans"]      = ", ".join(leaf.get("sans",[]))[:200]
                                sess["cert_key_type"]  = leaf.get("key_type","")
                                sess["cert_key_bits"]  = leaf.get("key_bits",0)
                                nb = leaf.get("not_before")
                                na = leaf.get("not_after")
                                if nb: sess["cert_not_before"] = nb.strftime("%Y-%m-%d")
                                if na:
                                    sess["cert_not_after"] = na.strftime("%Y-%m-%d")
                                    now = datetime.datetime.utcnow()
                                    sess["cert_expired"]       = na < now
                                    sess["cert_expiring_soon"] = (na - now).days < 30 and not sess["cert_expired"]

                        elif hs_type == 20:   # Finished
                            sess["handshake_complete"] = True

                elif ct == 21 and len(rec_data) >= 2:   # Alert
                    level = {1:"Warning",2:"Fatal"}.get(rec_data[0],"?")
                    desc  = ALERT_DESCS.get(rec_data[1], f"code {rec_data[1]}")
                    alert_str = f"{level}: {desc}"
                    if alert_str not in sess["alerts"]:
                        sess["alerts"].append(alert_str)

        except Exception:
            continue

    # ── Post-process: flag issues ─────────────────────────────────────────────
    now = datetime.datetime.utcnow()
    for sess in sessions.values():
        issues = []
        cs = sess["cipher_suite"]
        ver = sess["tls_version"]

        # Weak cipher suite
        if any(w in cs for w in WEAK_CIPHERS):
            sess["weak_cipher"] = True
            issues.append(f"Weak cipher: {cs}")

        # NULL / anon
        if "NULL" in cs or "anon" in cs.lower():
            issues.append("NULL or anonymous cipher — no encryption/auth")

        # Weak TLS version
        if ver in ("SSLv3","TLS 1.0","TLS 1.1"):
            sess["weak_version"] = True
            issues.append(f"Deprecated protocol: {ver}")

        # Cert expired
        if sess["cert_expired"]:
            issues.append("Certificate EXPIRED")

        # Cert expiring soon
        if sess["cert_expiring_soon"]:
            issues.append("Certificate expiring within 30 days")

        # Weak key
        kt = sess["cert_key_type"]
        kb = sess["cert_key_bits"]
        if "RSA" in kt and kb and kb < 2048:
            issues.append(f"Weak RSA key: {kb}-bit (minimum 2048)")
        if "RSA" in kt and kb and kb < 4096:
            if kb < 2048:
                pass  # already flagged
        if "ECDSA" in kt and kb and kb < 256:
            issues.append(f"Weak EC key: {kb}-bit")

        # Alerts
        for alert in sess["alerts"]:
            if "Fatal" in alert:
                issues.append(f"TLS alert — {alert}")

        # No SNI (possible direct IP connection or misconfigured client)
        # Trigger if: negotiated TLS version known (ServerHello seen) and no SNI
        if not sess["sni"] and ver and (sess["handshake_complete"] or sess.get("cipher_suite")):
            issues.append("No SNI — direct IP or misconfigured client")

        sess["issues"] = issues

    return list(sessions.values())


# Test export


# ─────────────────────────────────────────────────────────────────────────────
# 802.11 Wi-Fi frame parser — SSIDs, clients, APs, probe requests
# ─────────────────────────────────────────────────────────────────────────────

def extract_wifi_events(packets):
    """
    Parse 802.11 management frames (link-type 127 = RADIOTAP, or raw 802.11).
    Extracts:
      • Beacon frames      → AP BSSID, SSID, channel, capabilities, rates
      • Probe Requests     → client MAC, requested SSID (or wildcard)
      • Probe Responses    → AP BSSID, SSID, responding to client
      • Association Req/Resp → client↔AP association
      • Deauthentication   → reason code (useful: detect deauth attacks)
      • Authentication     → open/shared key sequence

    Returns dict with keys:
      "aps"      → list of AP dicts  {bssid, ssid, channel, enc, rates, clients}
      "clients"  → list of client dicts  {mac, probed_ssids, associated_bssid}
      "events"   → list of event dicts   {type, src_mac, dst_mac, bssid, ssid, ...}
    """
    # Frame type/subtype constants
    MGMT   = 0x00
    SUBTYPE = {
        0x00: "Association Request",
        0x01: "Association Response",
        0x02: "Reassociation Request",
        0x03: "Reassociation Response",
        0x04: "Probe Request",
        0x05: "Probe Response",
        0x08: "Beacon",
        0x0A: "Disassociation",
        0x0B: "Authentication",
        0x0C: "Deauthentication",
    }

    DEAUTH_REASONS = {
        1: "Unspecified", 2: "Previous auth no longer valid",
        3: "Leaving BSS", 4: "Inactivity", 5: "AP overloaded",
        6: "Class 2 frame from non-auth STA",
        7: "Class 3 frame from non-assoc STA",
        8: "Leaving BSS (re-assoc)", 9: "STA not authenticated",
        14: "MIC failure (TKIP)", 15: "4-way handshake timeout",
        16: "Group key handshake timeout", 17: "IE mismatch",
        23: "IEEE 802.1X auth failed",
    }

    def _mac_str(b):
        if len(b) < 6: return "??:??:??:??:??:??"
        return ":".join(f"{x:02x}" for x in b[:6])

    def _parse_ie(payload, offset):
        """
        Parse 802.11 Information Elements starting at offset.
        Returns dict: {ssid, channel, rates, enc, ht_cap, rsn, wpa}
        """
        result = {"ssid": "", "channel": 0, "rates": [], "enc": "Open",
                  "rsn": False, "wpa": False, "wps": False}
        while offset + 2 <= len(payload):
            ie_id  = payload[offset]
            ie_len = payload[offset + 1]
            offset += 2
            if offset + ie_len > len(payload):
                break
            ie_data = payload[offset:offset + ie_len]
            offset += ie_len

            if ie_id == 0:                  # SSID
                try:
                    result["ssid"] = ie_data.decode("utf-8", errors="replace")
                except Exception:
                    result["ssid"] = ie_data.hex()

            elif ie_id == 3 and ie_len >= 1:  # DS Parameter Set (channel)
                result["channel"] = ie_data[0]

            elif ie_id == 1 or ie_id == 50:   # Supported/Extended Rates
                rates = []
                for b in ie_data:
                    rate = (b & 0x7F) * 0.5
                    rates.append(f"{rate:.0f}")
                result["rates"].extend(rates)

            elif ie_id == 48:               # RSN (WPA2/WPA3)
                # RSN IE layout (IEEE 802.11-2020):
                #   [0:2]  Version (2)
                #   [2:6]  Group Cipher Suite (4)
                #   [6:8]  Pairwise Cipher Count (2)
                #   [8:8+cnt*4]  Pairwise Cipher Suite List
                #   [8+cnt*4:8+cnt*4+2]  AKM Suite Count (2)
                #   then AKM Suite List
                result["rsn"] = True
                result["enc"] = "WPA2"
                if ie_len >= 8:
                    pcnt = struct.unpack("<H", ie_data[6:8])[0]
                    akm_off = 8 + pcnt * 4          # start of AKM count field
                    if akm_off + 2 <= ie_len:
                        akm_cnt = struct.unpack("<H", ie_data[akm_off:akm_off+2])[0]
                        for i in range(min(akm_cnt, 8)):  # cap to avoid runaway
                            off2 = akm_off + 2 + i * 4
                            if off2 + 4 <= ie_len:
                                akm_type = ie_data[off2 + 3]
                                if akm_type in (8, 9, 18):   # SAE / OWE = WPA3
                                    result["enc"] = "WPA3"
                                elif akm_type in (1, 2) and result["enc"] != "WPA3":
                                    result["enc"] = "WPA2"
                                elif akm_type in (3, 4):     # FT variants
                                    if result["enc"] != "WPA3":
                                        result["enc"] += "+FT"

            elif ie_id == 221 and ie_data[:4] == bytes([0x00, 0x50, 0xF2, 0x01]):
                result["wpa"] = True  # WPA1 vendor IE
                if result["enc"] == "Open":
                    result["enc"] = "WPA"

            elif ie_id == 221 and ie_data[:4] == bytes([0x00, 0x50, 0xF2, 0x04]):
                result["wps"] = True   # WPS vendor IE

        result["rates"] = sorted(set(result["rates"]), key=float)
        return result

    aps      = {}   # bssid → ap dict
    clients  = {}   # mac → client dict
    events   = []

    for p in packets:
        raw = p.get("_raw_frame", b"")
        if not raw:
            continue

        # Handle Radiotap header (link-type 127)
        radiotap_len = 0
        if len(raw) >= 4 and struct.unpack("<H", raw[2:4])[0] >= 8:
            # Radiotap: version=0, pad=0, len at offset 2
            ver = raw[0]
            if ver == 0:
                try:
                    radiotap_len = struct.unpack("<H", raw[2:4])[0]
                except Exception:
                    pass

        dot11 = raw[radiotap_len:]
        if len(dot11) < 10:
            continue

        fc     = struct.unpack("<H", dot11[0:2])[0]
        ftype  = (fc >> 2) & 0x03
        fsub   = (fc >> 4) & 0x0F

        if ftype != MGMT:
            continue   # only management frames for now

        subtype_name = SUBTYPE.get(fsub, f"Mgmt-{fsub:#x}")

        # Extract addresses (offsets 4, 10, 16 in 802.11 header)
        if len(dot11) < 24:
            continue
        duration = struct.unpack("<H", dot11[2:4])[0]
        addr1 = _mac_str(dot11[4:10])    # destination
        addr2 = _mac_str(dot11[10:16])   # source / transmitter
        addr3 = _mac_str(dot11[16:22])   # BSSID
        seq   = struct.unpack("<H", dot11[22:24])[0]
        body  = dot11[24:]

        ts_us = p.get("ts_us", 0)

        if fsub == 0x08 or fsub == 0x05:  # Beacon or Probe Response
            if len(body) < 12:
                continue
            # Fixed parameters: Timestamp(8), Beacon Interval(2), Capability(2)
            cap_info = struct.unpack("<H", body[10:12])[0]
            ie_info  = _parse_ie(body, 12)
            bssid    = addr2   # transmitter = AP in beacon
            ssid     = ie_info["ssid"]
            channel  = ie_info["channel"]
            enc      = ie_info["enc"]
            wps      = ie_info["wps"]

            if bssid not in aps:
                aps[bssid] = {
                    "bssid": bssid, "ssid": ssid, "channel": channel,
                    "enc": enc, "wps": wps,
                    "rates": ", ".join(ie_info["rates"]) + " Mbps",
                    "clients": set(), "probe_responses": 0, "beacons": 0,
                }
            ap = aps[bssid]
            if ssid and not ap["ssid"]: ap["ssid"] = ssid
            if channel:                 ap["channel"] = channel
            if enc != "Open":           ap["enc"] = enc
            if wps:                     ap["wps"] = True
            if fsub == 0x08:            ap["beacons"] += 1
            else:                       ap["probe_responses"] += 1

            events.append(dict(
                frame_type  = subtype_name,
                src_mac     = bssid,
                dst_mac     = addr1,
                bssid       = addr3,
                ssid        = ssid,
                channel     = channel,
                enc         = enc,
                detail      = f"ch={channel} enc={enc}" + (" WPS!" if wps else ""),
                ts_us       = ts_us,
            ))

        elif fsub == 0x04:  # Probe Request
            ie_info = _parse_ie(body, 0)
            ssid    = ie_info["ssid"]   # empty = wildcard
            client_mac = addr2

            if client_mac not in clients:
                clients[client_mac] = {
                    "mac": client_mac, "probed_ssids": set(),
                    "associated_bssid": "", "assoc_ssid": "",
                }
            cl = clients[client_mac]
            if ssid:
                cl["probed_ssids"].add(ssid)

            events.append(dict(
                frame_type  = "Probe Request",
                src_mac     = client_mac,
                dst_mac     = "ff:ff:ff:ff:ff:ff",
                bssid       = addr3,
                ssid        = ssid or "(wildcard)",
                channel     = 0,
                enc         = "",
                detail      = f"Probing: {ssid or 'any'!r}",
                ts_us       = ts_us,
            ))

        elif fsub in (0x00, 0x02):  # Association / Reassociation Request
            ie_info    = _parse_ie(body, 4)   # skip cap+listen interval
            client_mac = addr2
            bssid      = addr3
            ssid       = ie_info["ssid"]

            if client_mac not in clients:
                clients[client_mac] = {"mac": client_mac, "probed_ssids": set(),
                                        "associated_bssid": "", "assoc_ssid": ""}
            clients[client_mac]["associated_bssid"] = bssid
            clients[client_mac]["assoc_ssid"]       = ssid

            if bssid in aps:
                aps[bssid]["clients"].add(client_mac)

            events.append(dict(
                frame_type  = "Association Request",
                src_mac     = client_mac,
                dst_mac     = addr1,
                bssid       = bssid,
                ssid        = ssid,
                channel     = 0,
                enc         = "",
                detail      = f"{client_mac} → {ssid!r}",
                ts_us       = ts_us,
            ))

        elif fsub == 0x0C:  # Deauthentication
            reason_code = struct.unpack("<H", body[0:2])[0] if len(body) >= 2 else 0
            reason_str  = DEAUTH_REASONS.get(reason_code, f"Reason {reason_code}")
            events.append(dict(
                frame_type  = "Deauthentication",
                src_mac     = addr2,
                dst_mac     = addr1,
                bssid       = addr3,
                ssid        = "",
                channel     = 0,
                enc         = "",
                detail      = f"Reason: {reason_str}",
                ts_us       = ts_us,
            ))

        elif fsub == 0x0A:  # Disassociation
            reason_code = struct.unpack("<H", body[0:2])[0] if len(body) >= 2 else 0
            reason_str  = DEAUTH_REASONS.get(reason_code, f"Reason {reason_code}")
            events.append(dict(
                frame_type  = "Disassociation",
                src_mac     = addr2,
                dst_mac     = addr1,
                bssid       = addr3,
                ssid        = "",
                channel     = 0,
                enc         = "",
                detail      = f"Reason: {reason_str}",
                ts_us       = ts_us,
            ))

    # Convert client set→list for JSON-ability
    for ap in aps.values():
        ap["clients"] = sorted(ap["clients"])

    return {
        "aps":     sorted(aps.values(),     key=lambda x: x["ssid"]),
        "clients": sorted(clients.values(), key=lambda x: x["mac"]),
        "events":  events,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Graph / enrichment builder
# ─────────────────────────────────────────────────────────────────────────────

def build_graph(packets, min_packets=1, collapse_external=False, extra_hostnames=None):
    """
    Returns nodes, edges, rows, findings
    """
    nodes = defaultdict(lambda: dict(
        count=0, bytes=0, is_private=False, role="client",
        macs=set(), mac_to_ips=defaultdict(set),
        subnet="", hostname="",
        protocols=set(), open_ports=set(),
        ttls=[], win_sizes=[],
        os_guess="Unknown",
        flags=set(),           # pentest flags
        first_seen=None, last_seen=None,
    ))
    edges = defaultdict(lambda: dict(
        count=0, bytes=0, protocols=set(), ports=set(), resources=set(),
        timestamps=[],
    ))
    rows          = []   # raw per-packet rows for Excel connections sheet
    findings      = []   # pentest findings rows
    cleartext_hits = []  # extracted credentials / sensitive data

    FAKE = {"ff:ff:ff:ff:ff:ff","00:00:00:00:00:00",""}

    # ARP table: ip -> set of MACs seen (anomaly detection)
    arp_table = defaultdict(set)
    # Track connections per IP pair over time (beaconing detection)
    conn_times = defaultdict(list)

    for p in packets:
        src, dst = p["src_ip"], p["dst_ip"]
        if collapse_external:
            src = _collapse(src); dst = _collapse(dst)
        if src == dst:
            continue

        smac = p.get("src_mac","")
        dmac = p.get("dst_mac","")

        # ARP anomaly tracking
        if p.get("arp_sender_mac"):
            arp_table[p["arp_sender_ip"]].add(p["arp_sender_mac"])

        if smac and smac not in FAKE:
            nodes[src]["macs"].add(smac)
            nodes[src]["mac_to_ips"][smac].add(src)
        if dmac and dmac not in FAKE:
            nodes[dst]["macs"].add(dmac)
            nodes[dst]["mac_to_ips"][dmac].add(dst)

        nodes[src]["count"] += 1; nodes[src]["bytes"] += p["length"]
        nodes[dst]["count"] += 1; nodes[dst]["bytes"] += p["length"]

        ts = p.get("ts_us", 0)
        for ip in (src, dst):
            if nodes[ip]["first_seen"] is None or ts < nodes[ip]["first_seen"]:
                nodes[ip]["first_seen"] = ts
            if nodes[ip]["last_seen"] is None or ts > nodes[ip]["last_seen"]:
                nodes[ip]["last_seen"] = ts

        if p.get("ttl"):
            nodes[src]["ttls"].append(p["ttl"])
        if p.get("win_size"):
            nodes[src]["win_sizes"].append(p["win_size"])

        proto = p["proto"]
        port  = p["dst_port"] or p["src_port"]

        # Open ports: a host is "listening" if it receives connections on known service ports
        if port and port < 1024:
            nodes[dst]["open_ports"].add(port)
        nodes[dst]["protocols"].add(proto)
        nodes[src]["protocols"].add(proto)

        key = (src, dst)
        edges[key]["count"]    += 1
        edges[key]["bytes"]    += p["length"]
        edges[key]["protocols"].add(proto)
        if port:       edges[key]["ports"].add(port)
        if p.get("resource"): edges[key]["resources"].add(p["resource"])
        conn_times[key].append(ts)

        rows.append(dict(
            src_ip=src, dst_ip=dst,
            src_mac=smac, dst_mac=dmac,
            proto=proto, port=port,
            resource=p.get("resource",""),
            ttl=p.get("ttl"),
        ))

        # Cleartext credential / sensitive data extraction
        hits = extract_cleartext(p)
        cleartext_hits.extend(hits)

    # ── Filter ────────────────────────────────────────────────────────────────
    edges = {k:v for k,v in edges.items() if v["count"] >= min_packets}
    active = {ip for pair in edges for ip in pair}
    nodes  = {k:v for k,v in nodes.items() if k in active}

    # ── Infer "locally-observed" IPs from ARP / MAC presence ────────────────
    # An IP seen in ARP sender/target or with a known MAC is definitely
    # on a local network segment, even if it's not RFC 1918.
    # This handles organisations using non-RFC1918 address blocks internally
    # (e.g. 20.x.x.x, 100.64.x.x CGNAT, 172.15.x.x etc.)
    # ── Infer locally-attached IPs from definitive L2 signals only ──────────
    # ARP sender IPs are ALWAYS on the local segment (ARP is never routed).
    # DHCP assigned addresses are also always local.
    # We deliberately do NOT use "has a MAC address" for regular TCP/UDP frames:
    # behind NAT, every packet has the gateway's MAC — using that would wrongly
    # classify 8.8.8.8 as "local" when it's just the gateway's upstream link.
    locally_observed_ips = set()
    for p in packets:
        # ARP sender/target are L2-local by definition
        arp_s = p.get("arp_sender_ip")
        if arp_s:
            locally_observed_ips.add(arp_s)
        if p.get("proto") == "ARP":
            arp_t = p.get("dst_ip","")
            if arp_t and arp_t not in ("0.0.0.0", "255.255.255.255"):
                locally_observed_ips.add(arp_t)
        # DHCP assigned IP (yiaddr) is always local
        if p.get("proto") == "DHCP":
            try:
                payload = p.get("app_payload", b"")
                if len(payload) >= 20:
                    for off in (12, 16):   # ciaddr + yiaddr
                        ip4 = socket.inet_ntoa(payload[off:off+4])
                        if ip4 not in ("0.0.0.0",):
                            locally_observed_ips.add(ip4)
            except Exception:
                pass
    locally_observed_ips.discard("")
    locally_observed_ips.discard("0.0.0.0")
    locally_observed_ips.discard("255.255.255.255")

    # ── Build the "always internal" network list ──────────────────────────────
    # These ranges are never globally routable and must always be internal,
    # regardless of whether we see ARP traffic for them.
    _ALWAYS_INTERNAL_NETS = [
        ipaddress.ip_network("10.0.0.0/8"),          # RFC 1918
        ipaddress.ip_network("172.16.0.0/12"),        # RFC 1918
        ipaddress.ip_network("192.168.0.0/16"),       # RFC 1918
        ipaddress.ip_network("169.254.0.0/16"),       # Link-local (RFC 3927)
        ipaddress.ip_network("127.0.0.0/8"),          # Loopback
        ipaddress.ip_network("100.64.0.0/10"),        # CGNAT shared (RFC 6598)
        ipaddress.ip_network("192.0.0.0/24"),         # IANA special (RFC 5736)
        ipaddress.ip_network("198.18.0.0/15"),        # Benchmarking (RFC 2544)
        ipaddress.ip_network("198.51.100.0/24"),      # Documentation (RFC 5737)
        ipaddress.ip_network("203.0.113.0/24"),       # Documentation (RFC 5737)
        ipaddress.ip_network("192.0.2.0/24"),         # Documentation (RFC 5737)
        ipaddress.ip_network("240.0.0.0/4"),          # Reserved (RFC 1112)
    ]
    # Merge in any caller-supplied networks (--internal-networks option)
    _extra_nets = (extra_hostnames or {}).get("__internal_networks__", [])
    _ALL_INTERNAL_NETS = _ALWAYS_INTERNAL_NETS + list(_extra_nets)

    def _is_internal_ip(ip_str):
        """Return True if this IP belongs to a non-globally-routable / internal range."""
        try:
            addr2 = ipaddress.ip_address(ip_str.split("/")[0])
            if addr2.is_private or addr2.is_link_local or addr2.is_loopback:
                return True
            if addr2.is_multicast:
                return False
            for net in _ALL_INTERNAL_NETS:
                if addr2 in net:
                    return True
            if ip_str in locally_observed_ips:
                return True
            return False
        except ValueError:
            return False

    # ── Annotate nodes ────────────────────────────────────────────────────────
    for ip, info in nodes.items():
        try:
            addr = ipaddress.ip_address(ip.split("/")[0])
            is_priv = _is_internal_ip(ip)
            if ip == "255.255.255.255" or addr.is_multicast:
                is_priv = False
            info["is_private"] = is_priv
            info["subnet"] = (str(ipaddress.ip_network(ip + "/24", strict=False))
                              if is_priv else "external")
        except ValueError:
            info["is_private"] = False; info["subnet"] = "external"

        info["role"] = _guess_role(ip, edges)

        # OS guess from most common TTL
        if info["ttls"]:
            common_ttl = max(set(info["ttls"]), key=info["ttls"].count)
            info["os_guess"] = _os_from_ttl(common_ttl)
            info["ttl_val"]  = common_ttl
        else:
            info["os_guess"] = "Unknown"
            info["ttl_val"]  = None

        # Pentest flags on node
        if info["protocols"] & CLEARTEXT_PROTOS:
            info["flags"].add("⚠ Cleartext protocol")
        if info["protocols"] & LATERAL_PROTOS and info["is_private"]:
            info["flags"].add("🔴 Lateral movement proto")

    # ── Passive hostname resolution from packet data ─────────────────────────
    passive_hostnames = resolve_hostnames_from_packets(packets)
    # Apply passive names first (lowest priority)
    for ip, hn in passive_hostnames.items():
        if ip in nodes and not nodes[ip]["hostname"]:
            nodes[ip]["hostname"] = hn
    # File-supplied hostnames override passive ones (highest priority)
    if extra_hostnames:
        for ip, hn in extra_hostnames.items():
            if ip in nodes:
                nodes[ip]["hostname"] = hn

    # ── Pentest Findings ──────────────────────────────────────────────────────

    # 1. Cleartext credential protocols
    for (src, dst), info in edges.items():
        ct = info["protocols"] & CLEARTEXT_PROTOS
        if ct:
            for proto in ct:
                findings.append(dict(
                    severity="HIGH",
                    category="Cleartext Protocol",
                    src=src, dst=dst,
                    detail=f"{proto} — credentials/data sent in cleartext",
                    recommendation="Upgrade to encrypted equivalent (SSH, HTTPS, LDAPS, IMAPS…)",
                ))

    # 2. Suspicious ports
    for (src, dst), info in edges.items():
        for port in info["ports"]:
            if port in SUSPICIOUS_PORTS:
                findings.append(dict(
                    severity="HIGH",
                    category="Suspicious Port",
                    src=src, dst=dst,
                    detail=f"Port {port} — {SUSPICIOUS_PORTS[port]}",
                    recommendation="Investigate — possible backdoor, C2, or misconfiguration",
                ))

    # 3. ARP anomalies (same IP, multiple MACs)
    for ip, macs in arp_table.items():
        if len(macs) > 1:
            findings.append(dict(
                severity="CRITICAL",
                category="ARP Anomaly / Possible MITM",
                src=ip, dst="N/A",
                detail=f"IP {ip} seen with MACs: {', '.join(sorted(macs))}",
                recommendation="Investigate for ARP spoofing / MITM attack",
            ))

    # 4. Unusual outbound (internal→external on non-standard ports)
    standard_out = {80,443,53,123,25,465,587,993,995,143,110,22}
    for (src, dst), info in edges.items():
        try:
            src_addr = ipaddress.ip_address(src.split("/")[0])
            dst_addr = ipaddress.ip_address(dst.split("/")[0])
        except ValueError:
            continue
        if src_addr.is_private and not dst_addr.is_private:
            odd_ports = info["ports"] - standard_out
            if odd_ports:
                findings.append(dict(
                    severity="MEDIUM",
                    category="Unusual Outbound",
                    src=src, dst=dst,
                    detail=f"Internal→External on non-standard port(s): {sorted(odd_ports)}",
                    recommendation="Verify legitimate — possible exfil or C2 beaconing",
                ))

    # 5. Beaconing detection (very regular inter-arrival times)
    for (src, dst), times in conn_times.items():
        if len(times) < 6:
            continue
        times_s = sorted(times)
        intervals = [times_s[i+1]-times_s[i] for i in range(len(times_s)-1)]
        intervals = [x for x in intervals if x > 0]
        if not intervals:
            continue
        mean = sum(intervals)/len(intervals)
        if mean == 0:
            continue
        variance = sum((x-mean)**2 for x in intervals)/len(intervals)
        cv = (variance**0.5) / mean  # coefficient of variation
        if cv < 0.15 and len(times) >= 8:  # very regular
            findings.append(dict(
                severity="MEDIUM",
                category="Potential Beaconing",
                src=src, dst=dst,
                detail=(f"{len(times)} connections, avg interval "
                        f"{mean/1e6:.1f}s, CoV={cv:.3f} (very regular)"),
                recommendation="Investigate for C2 callback / malware beaconing",
            ))

    # 6. Internal SMB/lateral movement between workstations
    for (src, dst), info in edges.items():
        if info["protocols"] & {"SMB","RDP","VNC"}:
            try:
                s = ipaddress.ip_address(src.split("/")[0])
                d = ipaddress.ip_address(dst.split("/")[0])
            except ValueError:
                continue
            if s.is_private and d.is_private:
                role_d = nodes.get(dst,{}).get("role","")
                if role_d == "client":   # workstation→workstation
                    findings.append(dict(
                        severity="MEDIUM",
                        category="Lateral Movement Indicator",
                        src=src, dst=dst,
                        detail=f"{', '.join(info['protocols'] & {'SMB','RDP','VNC'})} to a client workstation",
                        recommendation="Verify — unusual for workstations to accept SMB/RDP from peers",
                    ))

    # 7. SNMPv1/v2 (community strings in cleartext at scale)
    snmp_hosts = set()
    for (src, dst), info in edges.items():
        if "SNMP" in info["protocols"]:
            snmp_hosts.add(dst)
    if snmp_hosts:
        findings.append(dict(
            severity="MEDIUM",
            category="SNMP Cleartext",
            src="Multiple", dst=", ".join(sorted(snmp_hosts)[:5]),
            detail=f"SNMPv1/v2 traffic detected ({len(snmp_hosts)} hosts). Community strings in cleartext.",
            recommendation="Migrate to SNMPv3 with authentication and encryption",
        ))

    return nodes, edges, rows, findings, cleartext_hits, dict(arp_table)


def _collapse(ip):
    try:
        addr = ipaddress.ip_address(ip)
        if not addr.is_private:
            return str(ipaddress.ip_network(ip+"/24", strict=False))
    except ValueError:
        pass
    return ip


SERVER_PROTOS = {"HTTP","HTTPS","SSH","SMTP","DNS","MySQL","PostgreSQL","MSSQL",
                 "SMB","LDAP","RDP","HTTP-alt","HTTPS-alt","FTP","VNC",
                 "Redis","MongoDB","IMAP","POP3","SNMP","Syslog","NetBIOS"}


def _guess_role(ip, edges):
    peers = set()
    for s, d in edges:
        if s == ip: peers.add(d)
        elif d == ip: peers.add(s)
    for (s, d), info in edges.items():
        if d == ip and (info["protocols"] & SERVER_PROTOS):
            return "server"
    return "server" if len(peers)>=6 else ("host" if len(peers)>=3 else "client")


def _load_hostname_file(path):
    result = {}
    if not path or not os.path.exists(path): return result
    with open(path) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"): continue
            parts = line.split(None, 1)
            if len(parts) == 2: result[parts[0]] = parts[1]
    return result



# ─────────────────────────────────────────────────────────────────────────────
# Passive hostname resolver  (DNS / DHCP / mDNS / NetBIOS / HTTP Host)
# ─────────────────────────────────────────────────────────────────────────────

def resolve_hostnames_from_packets(packets):
    """
    Single-pass scan to build ip→hostname from data inside the capture.
    Sources (priority, lower = more trusted):
      1. DNS A/AAAA responses
      2. DHCP Option 12 (client self-reported hostname)
      3. mDNS A/AAAA responses
      4. NetBIOS Name Service registrations
      5. HTTP Host header (lowest — labels the destination only)
    """
    pending = {}   # ip -> (hostname, priority)

    def _set(ip, name, priority):
        name = name.rstrip(".").strip()
        if not name or not ip or ip in ("0.0.0.0", "255.255.255.255"):
            return
        existing = pending.get(ip)
        if existing is None or priority < existing[1]:
            pending[ip] = (name, priority)

    for p in packets:
        proto   = p.get("proto", "")
        payload = p.get("app_payload", b"")
        src_ip  = p.get("src_ip", "")
        dst_ip  = p.get("dst_ip", "")
        dp      = p.get("dst_port", 0)

        if proto in ("DNS", "mDNS") and payload:
            _parse_dns(payload, _set)

        if proto == "DHCP" and payload:
            _parse_dhcp(payload, _set)

        if proto == "NetBIOS" and payload:
            _parse_nbns(payload, src_ip, _set)

        if proto in ("HTTP", "HTTP-alt") and payload:
            try:
                text = payload.decode("ascii", errors="ignore")
                for line in text.split("\r\n"):
                    if line.lower().startswith("host:"):
                        host = line.split(":", 1)[1].strip().split(":")[0]
                        if host and not host.replace(".", "").isdigit():
                            _set(dst_ip, host, 5)
                        break
            except Exception:
                pass

    return {ip: name for ip, (name, _) in pending.items()}


def _parse_dns(data, setter):
    """
    Parse DNS / mDNS wire format.
    Extracts hostname→IP mappings (A, AAAA), PTR records, SRV, TXT.
    Calls setter(ip, name, priority) for every resolved mapping.
    Also handles mDNS reverse PTR (x.x.x.x.in-addr.arpa → hostname).
    """
    try:
        if len(data) < 12:
            return
        flags    = struct.unpack(">H", data[2:4])[0]
        is_resp  = (flags >> 15) & 1
        qd_count = struct.unpack(">H", data[4:6])[0]
        an_count = struct.unpack(">H", data[6:8])[0]
        ar_count = struct.unpack(">H", data[10:12])[0]   # additional records

        # Process ALL record sections: answers + additional records
        # (mDNS puts A records in the Additional section after PTR answers)
        total_answers = an_count + ar_count

        if not is_resp and an_count == 0 and ar_count == 0:
            return   # pure query with no piggybacked answers — skip for hostname resolution

        offset = 12
        # Skip questions
        for _ in range(qd_count):
            offset = _dns_skip_name(data, offset)
            if offset + 4 > len(data): return
            offset += 4   # QTYPE + QCLASS

        def _process_section(count):
            nonlocal offset
            for _ in range(count):
                if offset + 2 > len(data): break
                name, offset = _dns_read_name(data, offset)
                if offset + 10 > len(data): break
                rtype = struct.unpack(">H", data[offset:offset+2])[0]
                # Skip class (2), TTL (4)
                rdlen = struct.unpack(">H", data[offset+8:offset+10])[0]
                offset += 10
                if offset + rdlen > len(data): break
                rdata = data[offset:offset+rdlen]
                offset += rdlen

                if rtype == 1 and rdlen == 4:       # A record
                    try:
                        ip = socket.inet_ntoa(rdata)
                        setter(ip, name.rstrip("."), 1)
                    except Exception: pass

                elif rtype == 28 and rdlen == 16:   # AAAA record
                    try:
                        ip = socket.inet_ntop(socket.AF_INET6, rdata)
                        setter(ip, name.rstrip("."), 1)
                    except Exception: pass

                elif rtype == 12:                   # PTR record
                    try:
                        ptr_name, _ = _dns_read_name(data, offset - rdlen)
                        # Reverse PTR: x.x.x.x.in-addr.arpa → hostname
                        if ".in-addr.arpa" in name.lower():
                            # Decode reversed IP: 4.3.2.1.in-addr.arpa → 1.2.3.4
                            parts = name.lower().replace(".in-addr.arpa","").split(".")
                            if len(parts) == 4:
                                fwd_ip = ".".join(reversed(parts))
                                setter(fwd_ip, ptr_name.rstrip("."), 2)
                        elif ".ip6.arpa" not in name.lower():
                            # mDNS service PTR: _http._tcp.local → service name
                            # The PTR target is the device hostname
                            pass   # handled by SRV below
                    except Exception: pass

                elif rtype == 33:                   # SRV record
                    try:
                        if rdlen >= 6:
                            srv_target, _ = _dns_read_name(data, offset - rdlen + 6)
                            # SRV target is the canonical hostname — no IP yet
                            pass
                    except Exception: pass

        _process_section(an_count)
        _process_section(ar_count)

    except Exception:
        pass


def _extract_dns_events(packets):
    """
    Extract every DNS and mDNS query/response event from a packet list.
    Returns list of dicts:
      { client_ip, server_ip, direction, proto, query_name, qtype,
        answer_ip, answer_name, ttl, is_response, flags_desc, ts_us }

    Record types decoded: A(1), AAAA(28), CNAME(5), PTR(12), MX(15), NS(2),
                          SRV(33), TXT(16), SOA(6), ANY(255)
    """
    RTYPES = {
        1:"A", 2:"NS", 5:"CNAME", 6:"SOA", 12:"PTR", 15:"MX",
        16:"TXT", 28:"AAAA", 33:"SRV", 41:"OPT", 255:"ANY",
        65:"HTTPS", 64:"SVCB",
    }

    events = []

    for p in packets:
        proto   = p.get("proto","")
        payload = p.get("app_payload", b"")
        if proto not in ("DNS","mDNS") or not payload or len(payload) < 12:
            continue

        src_ip = p.get("src_ip","")
        dst_ip = p.get("dst_ip","")
        sp     = p.get("src_port",0)
        dp     = p.get("dst_port",0)
        ts_us  = p.get("ts_us",0)

        # mDNS: src=client, dst=224.0.0.251. For regular DNS:
        # query: client→53, response: 53→client
        is_mdns = (proto == "mDNS")

        try:
            txid     = struct.unpack(">H", payload[0:2])[0]
            flags    = struct.unpack(">H", payload[2:4])[0]
            is_resp  = (flags >> 15) & 1
            qd_count = struct.unpack(">H", payload[4:6])[0]
            an_count = struct.unpack(">H", payload[6:8])[0]
            ns_count = struct.unpack(">H", payload[8:10])[0]
            ar_count = struct.unpack(">H", payload[10:12])[0]

            # Decode flags
            rcode    = flags & 0x000F
            opcode   = (flags >> 11) & 0x000F
            aa       = (flags >> 10) & 1
            tc       = (flags >> 9) & 1
            rd       = (flags >> 8) & 1
            ra       = (flags >> 7) & 1

            flags_parts = []
            if is_resp:   flags_parts.append("Response")
            else:         flags_parts.append("Query")
            if aa:        flags_parts.append("AA")
            if rd:        flags_parts.append("RD")
            if ra:        flags_parts.append("RA")
            if tc:        flags_parts.append("TC")
            rcodes = {0:"NOERROR",1:"FORMERR",2:"SERVFAIL",3:"NXDOMAIN",
                      4:"NOTIMP",5:"REFUSED"}
            rcode_str = rcodes.get(rcode, f"RCODE{rcode}")
            if rcode: flags_parts.append(rcode_str)
            flags_desc = " | ".join(flags_parts)

            client_ip = src_ip if not is_resp else dst_ip
            server_ip = dst_ip if not is_resp else src_ip
            if is_mdns:
                client_ip = src_ip
                server_ip = "224.0.0.251 (mDNS)"

            offset = 12

            # ── Parse questions ────────────────────────────────────────────
            questions = []
            for _ in range(qd_count):
                if offset >= len(payload): break
                qname, offset = _dns_read_name(payload, offset)
                if offset + 4 > len(payload): break
                qt  = struct.unpack(">H", payload[offset:offset+2])[0]
                qc  = struct.unpack(">H", payload[offset+2:offset+4])[0]
                offset += 4
                qt_str = RTYPES.get(qt, str(qt))
                questions.append((qname.rstrip("."), qt_str))

            # ── Parse answer records ───────────────────────────────────────
            def parse_rrs(count):
                nonlocal offset
                records = []
                for _ in range(count):
                    if offset + 2 > len(payload): break
                    rname, offset = _dns_read_name(payload, offset)
                    if offset + 10 > len(payload): break
                    rtype  = struct.unpack(">H", payload[offset:offset+2])[0]
                    rclass = struct.unpack(">H", payload[offset+2:offset+4])[0]
                    rttl   = struct.unpack(">I", payload[offset+4:offset+8])[0]
                    rdlen  = struct.unpack(">H", payload[offset+8:offset+10])[0]
                    offset += 10
                    if offset + rdlen > len(payload): break
                    rdata  = payload[offset:offset+rdlen]
                    offset += rdlen

                    rtype_str = RTYPES.get(rtype, str(rtype))
                    answer_ip   = ""
                    answer_name = ""
                    answer_val  = ""

                    if rtype == 1 and rdlen == 4:    # A
                        try: answer_ip = socket.inet_ntoa(rdata)
                        except: pass
                        answer_val = answer_ip

                    elif rtype == 28 and rdlen == 16: # AAAA
                        try: answer_ip = socket.inet_ntop(socket.AF_INET6, rdata)
                        except: pass
                        answer_val = answer_ip

                    elif rtype in (5, 12, 2):         # CNAME, PTR, NS
                        try:
                            n, _ = _dns_read_name(payload, offset - rdlen)
                            answer_name = n.rstrip(".")
                            answer_val  = answer_name
                        except: pass

                    elif rtype == 15 and rdlen >= 3:  # MX
                        pref = struct.unpack(">H", rdata[:2])[0]
                        try:
                            exch, _ = _dns_read_name(payload, offset - rdlen + 2)
                            answer_val = f"[pref={pref}] {exch.rstrip('.')}"
                            answer_name = exch.rstrip(".")
                        except: pass

                    elif rtype == 16:                 # TXT
                        parts = []
                        pos2 = 0
                        while pos2 < len(rdata):
                            slen = rdata[pos2]; pos2 += 1
                            parts.append(rdata[pos2:pos2+slen].decode("utf-8","replace"))
                            pos2 += slen
                        answer_val = "; ".join(parts)[:200]

                    elif rtype == 33 and rdlen >= 6:  # SRV
                        pri = struct.unpack(">H", rdata[0:2])[0]
                        wt  = struct.unpack(">H", rdata[2:4])[0]
                        pt  = struct.unpack(">H", rdata[4:6])[0]
                        try:
                            tgt, _ = _dns_read_name(payload, offset - rdlen + 6)
                            answer_val = f"{tgt.rstrip('.')}:{pt} (pri={pri})"
                        except: pass

                    elif rtype == 41:                 # OPT (EDNS0) — skip silently
                        continue

                    records.append({
                        "rname": rname.rstrip("."),
                        "rtype": rtype_str,
                        "ttl":   rttl,
                        "answer_ip":   answer_ip,
                        "answer_name": answer_name,
                        "answer_val":  answer_val,
                    })
                return records

            answers    = parse_rrs(an_count)
            auth_rrs   = parse_rrs(ns_count)
            addl_rrs   = parse_rrs(ar_count)

            # ── Emit one event row per question×answer pair ───────────────
            if not questions:
                # Unsolicited announcement (mDNS) — no question, just answers
                for ans in answers + addl_rrs:
                    if ans["rtype"] in ("A","AAAA","PTR","CNAME"):
                        events.append(dict(
                            client_ip   = client_ip,
                            server_ip   = server_ip,
                            proto       = proto,
                            query_name  = ans["rname"],
                            qtype       = ans["rtype"],
                            answer_ip   = ans["answer_ip"],
                            answer_name = ans["answer_name"],
                            answer_val  = ans["answer_val"],
                            ttl         = ans["ttl"],
                            rcode       = rcode_str,
                            is_response = bool(is_resp),
                            flags_desc  = flags_desc,
                            ts_us       = ts_us,
                        ))
            else:
                for qname, qt_str in questions:
                    # Find matching answers (same name or any answer if no match)
                    matching = [a for a in answers if a["rname"].lower() == qname.lower()
                                or a["rname"].lower().endswith("." + qname.lower())]
                    if not matching:
                        matching = answers or [None]
                    for ans in matching:
                        events.append(dict(
                            client_ip   = client_ip,
                            server_ip   = server_ip,
                            proto       = proto,
                            query_name  = qname,
                            qtype       = qt_str,
                            answer_ip   = ans["answer_ip"]   if ans else "",
                            answer_name = ans["answer_name"] if ans else "",
                            answer_val  = ans["answer_val"]  if ans else "",
                            ttl         = ans["ttl"]         if ans else 0,
                            rcode       = rcode_str,
                            is_response = bool(is_resp),
                            flags_desc  = flags_desc,
                            ts_us       = ts_us,
                        ))

        except Exception:
            continue

    return events


def _dns_read_name(data, offset):
    labels = []; jumped = False; orig = offset
    visited = set()
    try:
        while offset < len(data):
            if offset in visited: break
            visited.add(offset)
            ln = data[offset]
            if ln == 0:
                if not jumped: orig = offset + 1
                break
            elif (ln & 0xC0) == 0xC0:
                if offset + 1 >= len(data): break
                ptr = ((ln & 0x3F) << 8) | data[offset+1]
                if not jumped: orig = offset + 2
                offset = ptr; jumped = True
            else:
                offset += 1
                labels.append(data[offset:offset+ln].decode("ascii","replace"))
                offset += ln
    except Exception:
        pass
    return ".".join(labels), orig


def _dns_skip_name(data, offset):
    try:
        while offset < len(data):
            ln = data[offset]
            if ln == 0: return offset + 1
            elif (ln & 0xC0) == 0xC0: return offset + 2
            else: offset += ln + 1
    except Exception:
        pass
    return offset


def _parse_dhcp(data, setter):
    """Extract DHCP Option 12 (Hostname) and correlate with yiaddr/ciaddr."""
    try:
        if len(data) < 240: return
        yiaddr = socket.inet_ntoa(data[16:20])
        ciaddr = socket.inet_ntoa(data[12:16])
        i = 240; hostname = None
        while i < len(data):
            opt = data[i]; i += 1
            if opt == 255: break
            if opt == 0: continue
            if i >= len(data): break
            ln = data[i]; i += 1
            val = data[i:i+ln]; i += ln
            if opt == 12:
                hostname = val.decode("ascii","replace").strip("\x00").strip()
        if hostname:
            for ip in (yiaddr, ciaddr):
                if ip and ip not in ("0.0.0.0","255.255.255.255"):
                    setter(ip, hostname, 2)
    except Exception:
        pass


def _parse_nbns(data, src_ip, setter):
    """
    Extract NetBIOS name from NBNS packets (both registrations and responses).

    NBNS wire format (RFC 1002 Section 4.2):
      Header:  12 bytes (TXID, FLAGS, QDCOUNT, ANCOUNT, NSCOUNT, ARCOUNT)
      Questions: QDCOUNT × (QNAME + QTYPE[2] + QCLASS[2])
      Answers:   ANCOUNT × (NAME + TYPE[2] + CLASS[2] + TTL[4] + RDLEN[2] + RDATA)

    QNAME/NAME encoding: a single label of length 0x20 (32 bytes) containing
    the Level-2 half-ASCII encoded 16-character NetBIOS name, followed by
    a null terminator byte (0x00).

    For registrations (QD=1, AN=0): name is in the question QNAME.
    For responses    (QD=0, AN=1): name is in the answer NAME.
    """
    try:
        if len(data) < 12:
            return
        qd_count = struct.unpack(">H", data[4:6])[0]
        an_count = struct.unpack(">H", data[6:8])[0]
        offset   = 12

        def _read_nbns_name(buf, off):
            """
            Read a NetBIOS encoded name at buf[off].
            Returns (decoded_name_str, new_offset) or ("", new_offset_past_name).
            """
            if off >= len(buf):
                return "", off
            ln = buf[off]; off += 1
            if ln == 0x20:                  # standard 32-byte NBNS label
                if off + 32 > len(buf):
                    return "", off + 32
                raw  = buf[off:off+32]
                name = _decode_nbns_name(raw).strip()
                off += 32
                # null terminator
                if off < len(buf) and buf[off] == 0:
                    off += 1
                return name, off
            elif ln == 0:                   # already at null terminator
                return "", off
            else:
                # skip non-standard label length gracefully
                off += min(ln, len(buf) - off)
                if off < len(buf) and buf[off] == 0:
                    off += 1
                return "", off

        # ── Questions ────────────────────────────────────────────────────────
        for _ in range(qd_count):
            if offset >= len(data):
                break
            name, offset = _read_nbns_name(data, offset)
            offset += 4   # QTYPE + QCLASS
            if name and name not in ("*", "__MSBROWSE__", ""):
                setter(src_ip, name, 3)

        # ── Answers ──────────────────────────────────────────────────────────
        for _ in range(an_count):
            if offset >= len(data):
                break
            name, offset = _read_nbns_name(data, offset)
            if offset + 10 > len(data):
                break
            rtype  = struct.unpack(">H", data[offset:offset+2])[0]
            offset += 8   # TYPE + CLASS + TTL
            rdlen  = struct.unpack(">H", data[offset:offset+2])[0]
            offset += 2
            rdata  = data[offset:offset+rdlen]
            offset += rdlen
            # Extract owner IP from NB record RDATA (2-byte flags + 4-byte IP)
            if name and rtype == 0x0020 and rdlen >= 6:
                try:
                    owner_ip = socket.inet_ntoa(rdata[2:6])
                    if owner_ip and owner_ip not in ("0.0.0.0",):
                        setter(owner_ip, name, 3)
                    else:
                        setter(src_ip, name, 3)
                except Exception:
                    setter(src_ip, name, 3)
            elif name and name not in ("*", "__MSBROWSE__", ""):
                setter(src_ip, name, 3)

    except Exception:
        pass


def _decode_nbns_name(raw):
    """
    Decode a Level-2 half-ASCII encoded NetBIOS name.
    Each original byte is split into two nibbles, each stored as (nibble + 0x41).
    So to decode: take pairs of bytes (A, B) → char = ((A-0x41)<<4) | (B-0x41)
    The name is padded to 16 chars (15 + suffix byte); strip trailing spaces.
    """
    try:
        chars = []
        for i in range(0, min(len(raw), 32), 2):
            a, b = raw[i], raw[i+1]
            # Validate: both bytes must be in range 0x41–0x50 (A–P)
            if not (0x41 <= a <= 0x50 and 0x41 <= b <= 0x50):
                break
            c = ((a - 0x41) << 4) | (b - 0x41)
            # Include all printable ASCII including space (0x20)
            if 0x20 <= c < 0x7f:
                chars.append(chr(c))
            else:
                break   # non-printable → stop (suffix byte or corruption)
        # Strip trailing spaces (padding) but keep internal spaces
        return "".join(chars).rstrip()
    except Exception:
        return ""



def _decode_nbns_payload(payload):
    """
    Decode NetBIOS Name Service packets (UDP 137) — both queries and responses.
    Returns list of human-readable NetBIOS names found, properly decoded from
    the Level-2 half-ASCII encoding (each byte pair encodes one character).
    """
    names = []
    try:
        if len(payload) < 12:
            return names
        qd_count = struct.unpack(">H", payload[4:6])[0]
        an_count = struct.unpack(">H", payload[6:8])[0]
        total    = qd_count + an_count
        offset   = 12
        for _ in range(total):
            if offset >= len(payload):
                break
            # Length byte of the encoded name label (should be 0x20 = 32)
            ln = payload[offset]
            offset += 1
            if ln == 0x20 and offset + 32 <= len(payload):
                raw  = payload[offset:offset+32]
                name = _decode_nbns_name(raw).strip()
                if name and name not in ("*", ""):
                    names.append(name)
                offset += 32
                # null terminator
                if offset < len(payload) and payload[offset] == 0:
                    offset += 1
                # skip QTYPE + QCLASS (4 bytes) or RTYPE+RCLASS+TTL+RDLEN (10 bytes)
                offset += 4
            else:
                break
    except Exception:
        pass
    return names


# ─────────────────────────────────────────────────────────────────────────────
# Traceroute / ICMP TTL-exceeded hop reconstructor
# ─────────────────────────────────────────────────────────────────────────────

def extract_traceroutes(packets):
    """
    Reconstruct traceroute paths from ICMP TTL-exceeded (type 11) replies.

    A traceroute works by sending probes with increasing TTL values.
    Each router that drops a packet (TTL=0) sends back an ICMP type 11
    "Time Exceeded" message.  The inner IP header tells us the original
    destination; the outer IP source is the hop.

    Returns list of traceroute dicts:
        {src, dst, hops: [{hop_n, router_ip, rtt_ms}]}
    """
    # Collect ICMP type-11 replies: outer_src = router, inner = original probe
    # We detect them by looking at ICMP packets where proto is ICMP and
    # the payload starts with an IP header (inner encapsulated packet).
    hop_events = []   # (origin_src, final_dst, hop_router, approx_ttl_level)

    for p in packets:
        if p.get("proto") != "ICMP":
            continue
        payload = p.get("app_payload", b"")
        # ICMP header: type(1) code(1) checksum(2) rest(4) = 8 bytes
        # For type 11 (TTL exceeded), the payload is the original IP header + 8 bytes
        if len(payload) < 8:
            continue
        icmp_type = payload[0]
        if icmp_type != 11:   # Time Exceeded
            continue
        inner = payload[8:]   # original IP packet (at least 20 bytes header)
        if len(inner) < 20:
            continue
        try:
            inner_ihl  = (inner[0] & 0x0F) * 4
            inner_src  = socket.inet_ntoa(inner[12:16])
            inner_dst  = socket.inet_ntoa(inner[16:20])
            router_ip  = p["src_ip"]   # the router sending the TTL-exceeded back
            hop_events.append((inner_src, inner_dst, router_ip))
        except Exception:
            continue

    if not hop_events:
        return []

    # Group by (origin → destination) pair
    paths = defaultdict(list)
    for origin, dest, router in hop_events:
        paths[(origin, dest)].append(router)

    traces = []
    for (origin, dest), routers in paths.items():
        # Deduplicate while preserving order
        seen = []
        for r in routers:
            if r not in seen:
                seen.append(r)
        hops = [{"hop_n": i+1, "router_ip": r} for i, r in enumerate(seen)]
        traces.append({"src": origin, "dst": dest, "hops": hops})

    return traces


# ─────────────────────────────────────────────────────────────────────────────
# Gateway / router detection from ARP + routing heuristics
# ─────────────────────────────────────────────────────────────────────────────

def detect_gateways(packets, nodes):
    """
    Identify default gateways by looking at:
      1. ARP requests sent to the router from each subnet — the router IP
         is the most-queried ARP target that is NOT a host sending traffic
      2. Hosts that forward packets for many other subnets (high TTL, many peers)

    Returns dict: {subnet_cidr -> gateway_ip}
    """
    # Count ARP targets per subnet
    arp_targets = defaultdict(lambda: defaultdict(int))  # subnet -> ip -> count

    for p in packets:
        if p.get("proto") != "ARP":
            continue
        src_ip = p.get("src_ip","")
        dst_ip = p.get("dst_ip","")
        if not src_ip or not dst_ip:
            continue
        try:
            src_addr = ipaddress.ip_address(src_ip)
            if not src_addr.is_private:
                continue
            net = str(ipaddress.ip_network(src_ip + "/24", strict=False))
            arp_targets[net][dst_ip] += 1
        except Exception:
            continue

    gateways = {}
    for subnet, targets in arp_targets.items():
        if not targets:
            continue
        # Most ARP'd target that is itself in the same subnet is likely the gateway
        candidates = sorted(targets.items(), key=lambda x: -x[1])
        for ip, count in candidates:
            try:
                addr = ipaddress.ip_address(ip)
                if addr.is_private and count >= 2:
                    gateways[subnet] = ip
                    break
            except Exception:
                continue

    # Fallback: nodes with role "server" or "host" that have the lowest
    # host octet (e.g. .1) are likely gateways/routers
    for ip, info in nodes.items():
        try:
            addr   = ipaddress.ip_address(ip)
            subnet = info["subnet"]
            if subnet == "external" or subnet in gateways:
                continue
            if addr.is_private and (addr.packed[-1] in (1, 254)):
                gateways[subnet] = ip
        except Exception:
            continue

    return gateways


# ─────────────────────────────────────────────────────────────────────────────
# Diagram styles & layout
# ─────────────────────────────────────────────────────────────────────────────

SHAPE_SERVER = "shape=mxgraph.cisco.servers.standard_server;"
SHAPE_PC     = "shape=mxgraph.cisco.computers_and_peripherals.pc;"
SHAPE_CLOUD  = "shape=mxgraph.cisco.storage.cloud;"

ROLE_FILL_STROKE = {
    "server":   ("#dae8fc","#6c8ebf"),
    "host":     ("#d5e8d4","#82b366"),
    "client":   ("#fff2cc","#d6b656"),
    "external": ("#ffe6cc","#d79b00"),
}
FLAG_FILL_STROKE = ("#f8cecc","#b85450")  # red tint for flagged nodes

SUBNET_PALETTES = [
    ("#e8f5e9","#2e7d32"),
    ("#e3f2fd","#1565c0"),
    ("#fce4ec","#c62828"),
    ("#f3e5f5","#6a1b9a"),
    ("#fff8e1","#f57f17"),
    ("#e0f7fa","#00695c"),
    ("#fbe9e7","#bf360c"),
    ("#ede7f6","#4527a0"),
]

# ── Layout constants ─────────────────────────────────────────────────────────
# The node is rendered as a Cisco icon (NODE_W × NODE_H) with a text label
# below it.  draw.io's verticalLabelPosition=bottom places the label OUTSIDE
# the icon geometry, so we must account for label height in row spacing.
#
# STRIDE_X = horizontal distance between node top-left corners
# STRIDE_Y = vertical distance between node top-left corners
#          = NODE_H + LABEL_RESERVE + inter-node gap
NODE_W          = 72    # icon width  (also used as label width)
NODE_H          = 60    # icon height only — geometry passed to draw.io
LABEL_RESERVE   = 72    # vertical space reserved for label text (4 lines + padding)
INTER_GAP_X     = 28    # horizontal gap between adjacent node label areas
INTER_GAP_Y     = 18    # vertical gap between bottom of one label and top of next icon
STRIDE_X        = NODE_W + INTER_GAP_X          # 100
STRIDE_Y        = NODE_H + LABEL_RESERVE + INTER_GAP_Y   # 150
CONT_PAD        = 44    # padding inside container border
CONT_TITLE      = 34    # container header height
CONT_GAP_X      = 60    # gap between container boxes horizontally
CONT_GAP_Y      = 50    # gap between container boxes vertically
LEGEND_W        = 185
PAGE_X          = 215
PAGE_Y          = 80
NODES_PER_ROW   = 4
CONTS_PER_ROW   = 3


def layout(nodes):
    subnets = defaultdict(list)
    for ip, info in nodes.items():
        subnets[info["subnet"]].append(ip)

    def skey(s):
        if s == "external": return (1,"")
        try: return (0, str(ipaddress.ip_network(s)))
        except: return (0,s)

    sorted_subs = sorted(subnets, key=skey)
    containers  = []
    node_pos    = {}
    cx, cy      = PAGE_X, PAGE_Y
    row_max_h   = 0
    col         = 0

    for si, sub in enumerate(sorted_subs):
        ips = subnets[sub]

        def isort(ip):
            r = {"server":0,"host":1,"client":2}.get(nodes[ip]["role"],3)
            try: return (r, int(ipaddress.ip_address(ip.split("/")[0])))
            except: return (r, ip)
        ips.sort(key=isort)

        n     = len(ips)
        ncols = min(n, NODES_PER_ROW)
        nrows = math.ceil(n / NODES_PER_ROW)

        # Inner dimensions: each node occupies STRIDE_X × STRIDE_Y
        # but the last node in each row/col doesn't need the trailing gap
        inner_w = ncols * STRIDE_X - INTER_GAP_X
        inner_h = nrows * STRIDE_Y - INTER_GAP_Y
        cw = inner_w + CONT_PAD * 2
        ch = inner_h + CONT_PAD * 2 + CONT_TITLE

        for i, ip in enumerate(ips):
            r2, c2 = divmod(i, NODES_PER_ROW)
            # nx/ny = top-left corner of the icon within the container
            nx = cx + CONT_PAD + c2 * STRIDE_X
            ny = cy + CONT_TITLE + CONT_PAD + r2 * STRIDE_Y
            node_pos[ip] = (int(nx), int(ny))

        containers.append(dict(subnet=sub, palette_idx=si,
                               x=cx, y=cy, w=cw, h=ch, ips=ips))
        row_max_h = max(row_max_h, ch)
        col += 1
        if col >= CONTS_PER_ROW:
            cx = PAGE_X; cy += row_max_h + CONT_GAP_Y; row_max_h = 0; col = 0
        else:
            cx += cw + CONT_GAP_X

    return node_pos, containers


# ─────────────────────────────────────────────────────────────────────────────
# XML helpers
# ─────────────────────────────────────────────────────────────────────────────

def _cell(gp, **attrs):
    c = ET.SubElement(gp, "mxCell")
    for k,v in attrs.items(): c.set(k, str(v))
    return c

def _geo(cell, x=0, y=0, w=10, h=10, relative=None):
    kw = {"x":str(int(x)),"y":str(int(y)),
          "width":str(int(w)),"height":str(int(h)),"as":"geometry"}
    if relative is not None: kw["relative"] = str(relative)
    ET.SubElement(cell, "mxGeometry", **kw)


# ─────────────────────────────────────────────────────────────────────────────

def generate_l2_drawio(wifi_data, nodes, arp_table, title="L2 / Wi-Fi Topology"):
    """
    Produce a Layer-2 / Wi-Fi topology draw.io diagram.

    Wi-Fi section  (from 802.11 management frames):
      - One swimlane per AP: SSID / BSSID / channel / encryption label
      - AP node with Cisco wireless icon, colour-coded by encryption strength
        Open=red, WPA1=amber, WPA2=green, WPA3=blue
      - Associated client nodes inside the AP lane, connected by lines
      - WPS badge on APs advertising Wi-Fi Protected Setup
      - Deauthentication event count + reason shown as a warning banner
      - Channel labelled with band (2.4 / 5 / 6 GHz)

    Unassociated clients section:
      - Clients seen probing but not associated to any AP

    Wired L2 section  (from ARP / MAC table):
      - One swimlane per /24 subnet
      - Each host: IP / hostname / MAC(s)
      - ARP anomaly (multiple MACs for same IP) flagged in red

    Legend at top-left explains all colours.
    """

    # ── XML-safe string helper ────────────────────────────────────────────────
    # Strip characters that are illegal in XML 1.0 (anything < 0x09,
    # 0x0B, 0x0C, 0x0E-0x1F).  We do NOT use a regex with \x00 in the
    # source because that literal can be corrupted by binary editors;
    # instead we build the translation table programmatically.
    _illegal_xml = str.maketrans(
        "", "",
        "".join(chr(i) for i in
                list(range(0, 9)) + [11, 12] + list(range(14, 32)))
    )

    def xs(s):
        """Return s safe for embedding in draw.io XML cell values (html=1)."""
        if not isinstance(s, str):
            s = str(s)
        s = s.translate(_illegal_xml)
        # ET handles attribute escaping, but for HTML embedded in value=""
        # we need to escape & < > ourselves.
        return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    # ── Cisco shape strings ───────────────────────────────────────────────────
    SHAPE_AP     = "shape=mxgraph.cisco.wireless.wireless_access_point;"
    SHAPE_PC     = "shape=mxgraph.cisco.computers_and_peripherals.pc;"
    SHAPE_LAPTOP = "shape=mxgraph.cisco.computers_and_peripherals.workstation;"
    SHAPE_SERVER = "shape=mxgraph.cisco.servers.standard_server;"

    # ── Colour scheme (fill, stroke) ──────────────────────────────────────────
    ENC_COLOURS = {
        "Open":    ("#f8cecc", "#b85450"),   # red   — no encryption
        "WPA":     ("#ffe6cc", "#d79b00"),   # amber — deprecated
        "WPA2":    ("#d5e8d4", "#82b366"),   # green — current standard
        "WPA3":    ("#dae8fc", "#6c8ebf"),   # blue  — latest (SAE/OWE)
        "WPA2+FT": ("#d5e8d4", "#82b366"),
        "WPA3+FT": ("#dae8fc", "#6c8ebf"),
    }
    ENC_DEFAULT    = ("#f5f5f5", "#666666")
    CLI_ASSOC_FILL = "#fff2cc";  CLI_ASSOC_STR = "#d6b656"
    CLI_UA_FILL    = "#f5f5f5";  CLI_UA_STR    = "#aaaaaa"
    WIRED_FILL     = "#e3f2fd";  WIRED_STR     = "#1565c0"
    ANOMALY_FILL   = "#f8cecc";  ANOMALY_STR   = "#b85450"

    # ── Layout constants ──────────────────────────────────────────────────────
    PX = 30         # page left margin
    PY = 80         # first content Y (below title)
    AP_W    = 330   # AP swimlane width
    AP_H0   = 140   # base height before clients
    CLI_H   = 44    # per-client row height
    CLI_PAD = 6
    AP_GAPX = 28
    AP_GAPY = 50
    APS_ROW = 4
    AP_ICO_W = 64; AP_ICO_H = 52
    CLI_ICO_W = 48; CLI_ICO_H = 36
    WIRED_W  = 270
    WIRED_H0 = 36   # header height
    HOST_H   = 48
    HOST_PAD = 6
    SEG_GAPX = 26
    SEG_GAPY = 40
    SEGS_ROW = 4

    aps     = wifi_data.get("aps",     [])
    clients = wifi_data.get("clients", [])
    events  = wifi_data.get("events",  [])

    # Precompute: MAC → (ip, hostname) from the nodes dict
    mac_to_ip = {}
    mac_to_hn = {}
    for ip, info in nodes.items():
        for mac in info.get("macs", set()):
            mac_to_ip[mac] = ip
            if info.get("hostname"):
                mac_to_hn[mac] = info["hostname"]

    # Deauth / disassoc events grouped by BSSID
    deauths_by_bssid = defaultdict(list)
    for ev in events:
        if ev.get("frame_type") in ("Deauthentication", "Disassociation"):
            b = ev.get("bssid", "") or ev.get("dst_mac", "")
            if b:
                deauths_by_bssid[b].append(ev)

    # ── XML scaffold ─────────────────────────────────────────────────────────
    root = ET.Element("mxGraphModel",
        dx="1422", dy="762", grid="1", gridSize="10", guides="1",
        tooltips="1", connect="1", arrows="1", fold="1", page="1",
        pageScale="1", pageWidth="1654", pageHeight="1169",
        math="0", shadow="0")
    g = ET.SubElement(root, "root")
    ET.SubElement(g, "mxCell", id="0")
    ET.SubElement(g, "mxCell", id="1", parent="0")

    _cid = [0]
    def cid():
        _cid[0] += 1
        return f"l2c{_cid[0]}"

    # ── Title ─────────────────────────────────────────────────────────────────
    tc = _cell(g, id=cid(),
               value=f"<b>&#128246; {xs(title)}</b>",
               style=("text;html=1;strokeColor=none;fillColor=none;"
                      "align=left;verticalAlign=middle;whiteSpace=wrap;"
                      "rounded=0;fontSize=18;fontColor=#1A2744;"),
               vertex="1", parent="1")
    _geo(tc, x=PX, y=16, w=900, h=40)

    # ── Legend ────────────────────────────────────────────────────────────────
    legend_rows = [
        ("Open network — no encryption",     "#f8cecc", "#b85450"),
        ("WPA (deprecated TKIP/CCMP)",        "#ffe6cc", "#d79b00"),
        ("WPA2 — current standard",           "#d5e8d4", "#82b366"),
        ("WPA3 — latest (SAE / OWE)",         "#dae8fc", "#6c8ebf"),
        ("Associated client",                 "#fff2cc", "#d6b656"),
        ("Unassociated / probing client",     "#f5f5f5", "#aaaaaa"),
        ("Wired host (from ARP table)",       "#e3f2fd", "#1565c0"),
        ("ARP anomaly — possible MITM",       "#f8cecc", "#b85450"),
    ]
    ROW_H   = 22
    LEG_H   = 28 + len(legend_rows) * ROW_H + 8
    leg_id  = cid()
    lc = _cell(g, id=leg_id, value="<b>Legend</b>",
               style=("swimlane;startSize=24;fillColor=#f5f5f5;"
                      "strokeColor=#aaaaaa;fontColor=#333;fontSize=10;"
                      "fontStyle=1;rounded=1;arcSize=3;html=1;"),
               vertex="1", parent="1")
    _geo(lc, x=PX, y=PY, w=240, h=LEG_H)
    for li, (lbl, fill, stroke) in enumerate(legend_rows):
        sw = _cell(g, id=cid(), value="",
                   style=(f"rounded=1;arcSize=50;fillColor={fill};"
                          f"strokeColor={stroke};strokeWidth=1.5;html=1;"),
                   vertex="1", parent=leg_id)
        _geo(sw, x=8, y=28 + li * ROW_H + 4, w=16, h=14)
        lt = _cell(g, id=cid(), value=xs(lbl),
                   style=("text;html=1;strokeColor=none;fillColor=none;"
                          "align=left;verticalAlign=middle;fontSize=9;"),
                   vertex="1", parent=leg_id)
        _geo(lt, x=30, y=28 + li * ROW_H + 3, w=206, h=16)

    cur_y = PY + LEG_H + 30   # Y cursor below legend

    # ════════════════════════════════════════════════════════════════════════
    # SECTION 1 — Wi-Fi Access Points
    # ════════════════════════════════════════════════════════════════════════
    def ch_band(ch):
        """Return human-readable channel + band string."""
        try:
            c = int(ch)
            if 1 <= c <= 14:   return f"ch {c}  (2.4 GHz)"
            if 36 <= c <= 177: return f"ch {c}  (5 GHz)"
            if c > 177:        return f"ch {c}  (6 GHz)"
        except (TypeError, ValueError):
            pass
        return f"ch {ch}" if ch else "ch ?"

    if aps:
        # Section header
        sh = _cell(g, id=cid(),
                   value="<b>&#128246; Wi-Fi Access Points &amp; Associated Clients</b>",
                   style=("text;html=1;strokeColor=none;fillColor=none;"
                          "align=left;verticalAlign=middle;fontSize=13;"
                          "fontColor=#1A2744;fontStyle=1;"),
                   vertex="1", parent="1")
        _geo(sh, x=PX, y=cur_y, w=900, h=26)
        cur_y += 32

        ap_cx = PX; ap_cy = cur_y; col = 0; row_max_h = 0

        for ap in aps:
            bssid   = ap.get("bssid", "??")
            ssid    = ap.get("ssid", "") or "(hidden SSID)"
            channel = ap.get("channel", 0)
            enc     = ap.get("enc", "Open")
            wps     = ap.get("wps", False)
            beacons = ap.get("beacons", 0)
            ap_clis = ap.get("clients", [])      # list of client MACs
            deauths = deauths_by_bssid.get(bssid, [])
            n_dauth = len(deauths)

            fill, stroke = ENC_COLOURS.get(enc, ENC_DEFAULT)

            # Compute container height
            extra_h = CLI_H * len(ap_clis) + (32 if n_dauth else 0)
            cont_h  = AP_H0 + extra_h

            # Encryption icon
            enc_ico = {"Open": "&#128274;",
                       "WPA":  "&#9888;",
                       "WPA2": "&#9989;",
                       "WPA3": "&#9989;&#9989;"}.get(enc, "&#128275;")
            wps_txt   = "  &#9888; WPS!" if wps else ""
            dauth_txt = f"  &#128308; {n_dauth} deauth" if n_dauth else ""

            cont_val = (f"<b>{xs(ssid)}</b><br/>"
                        f"<font style='font-size:9px;color:#444;'>{xs(bssid)}</font><br/>"
                        f"<font style='font-size:9px;'>{enc_ico} {xs(enc)}"
                        f"{wps_txt}{dauth_txt}"
                        f"  &middot;  {xs(ch_band(channel))}</font>")

            cont_id = cid()
            cc = _cell(g, id=cont_id, value=cont_val,
                       style=(f"swimlane;startSize=54;fillColor={fill};"
                              f"strokeColor={stroke};strokeWidth=2;"
                              "fontColor=#333;fontSize=10;"
                              "rounded=1;arcSize=3;swimlaneLine=1;html=1;"
                              "spacingLeft=6;align=left;verticalAlign=top;"),
                       vertex="1", parent="1")
            _geo(cc, x=ap_cx, y=ap_cy, w=AP_W, h=cont_h)

            # AP icon node
            ap_nid = cid()
            ap_tip = (f"BSSID: {xs(bssid)}\nSSID: {xs(ssid)}\n"
                      f"Encryption: {xs(enc)}\nChannel: {channel}\n"
                      f"WPS: {'YES - risk!' if wps else 'no'}\n"
                      f"Beacons seen: {beacons}\nDeauth frames: {n_dauth}")
            an = _cell(g, id=ap_nid,
                       value=(f"<b>AP</b><br/>"
                              f"<font style='font-size:8px;'>&#9901; {beacons} beacons</font>"),
                       tooltip=ap_tip,
                       style=(f"{SHAPE_AP}fillColor={fill};strokeColor={stroke};"
                              "strokeWidth=2;verticalLabelPosition=bottom;"
                              "verticalAlign=top;labelPosition=center;"
                              "align=center;fontSize=9;html=1;"),
                       vertex="1", parent=cont_id)
            _geo(an, x=(AP_W - AP_ICO_W) // 2, y=58, w=AP_ICO_W, h=AP_ICO_H)

            # Deauth warning banner
            if n_dauth:
                reasons = sorted({ev.get("detail", "") for ev in deauths})
                rsn_str = xs(" | ".join(r for r in reasons if r)[:80])
                dw = _cell(g, id=cid(),
                           value=(f"<b>&#128308; {n_dauth} deauth / disassoc frame(s)</b>"
                                  + (f"<br/><font style='font-size:8px;'>{rsn_str}</font>"
                                     if rsn_str else "")),
                           style=("text;html=1;strokeColor=#b85450;fillColor=#f8cecc;"
                                  "align=center;verticalAlign=middle;fontSize=9;"
                                  "rounded=1;arcSize=8;"),
                           vertex="1", parent=cont_id)
                dw_y = AP_H0 - 36 + CLI_H * len(ap_clis)
                _geo(dw, x=6, y=dw_y, w=AP_W - 12, h=26)

            # Associated client nodes
            for ci2, cli_mac in enumerate(sorted(ap_clis)):
                cli_ip = mac_to_ip.get(cli_mac, "")
                cli_hn = mac_to_hn.get(cli_mac, "")
                top    = xs(cli_hn or cli_ip or cli_mac)
                parts  = []
                if cli_hn and cli_ip: parts.append(cli_ip)
                parts.append(cli_mac)
                bot = xs(" | ".join(parts)) if top != xs(cli_mac) else ""

                cli_val = (f"<b>{top}</b>"
                           + (f"<br/><font style='font-size:8px;color:#444;'>"
                              f"{bot}</font>" if bot else ""))
                cli_tip = (f"MAC: {xs(cli_mac)}\nIP: {xs(cli_ip or 'unknown')}"
                           f"\nHostname: {xs(cli_hn or 'unknown')}")
                cy_off = AP_H0 - 10 + ci2 * (CLI_H + CLI_PAD)
                cli_nid = cid()
                cn = _cell(g, id=cli_nid, value=cli_val, tooltip=cli_tip,
                           style=(f"{SHAPE_PC}fillColor={CLI_ASSOC_FILL};"
                                  f"strokeColor={CLI_ASSOC_STR};"
                                  "verticalLabelPosition=right;verticalAlign=middle;"
                                  "labelPosition=right;align=left;"
                                  "fontSize=9;html=1;"),
                           vertex="1", parent=cont_id)
                _geo(cn, x=CLI_PAD, y=cy_off, w=CLI_ICO_W, h=CLI_ICO_H)

                # Edge: AP icon → client
                ec = _cell(g, id=cid(), value="",
                           style=("endArrow=none;startArrow=none;"
                                  "strokeColor=#bbbbbb;strokeWidth=1;"
                                  "opacity=60;dashed=0;rounded=1;html=1;"),
                           edge="1", source=ap_nid, target=cli_nid,
                           parent=cont_id)
                ET.SubElement(ec, "mxGeometry", relative="1", **{"as": "geometry"})

            # Advance layout
            row_max_h = max(row_max_h, cont_h)
            col += 1
            if col >= APS_ROW:
                ap_cx = PX; ap_cy += row_max_h + AP_GAPY
                row_max_h = 0; col = 0
            else:
                ap_cx += AP_W + AP_GAPX

        cur_y = ap_cy + row_max_h + AP_GAPY + 10

    else:
        # Placeholder when no Wi-Fi frames in capture
        ph = _cell(g, id=cid(),
                   value=("<i>No 802.11 management frames found in this capture.<br/>"
                          "Capture in monitor mode:  "
                          "airmon-ng start wlan0 &amp;&amp; "
                          "tcpdump -i wlan0mon -w wifi.pcap</i>"),
                   style=("text;html=1;strokeColor=#aaaaaa;fillColor=#f9f9f9;"
                          "align=center;verticalAlign=middle;fontSize=10;"
                          "rounded=1;arcSize=5;"),
                   vertex="1", parent="1")
        _geo(ph, x=PX, y=cur_y + 32, w=680, h=52)
        cur_y += 100

    # ════════════════════════════════════════════════════════════════════════
    # SECTION 2 — Unassociated / probing clients
    # ════════════════════════════════════════════════════════════════════════
    ap_macs = {mac for ap in aps for mac in ap.get("clients", [])}
    unassoc  = [c for c in clients if c.get("mac") and c["mac"] not in ap_macs]

    if unassoc:
        ush = _cell(g, id=cid(),
                    value="<b>&#128270; Unassociated / Probing Clients</b>",
                    style=("text;html=1;strokeColor=none;fillColor=none;"
                           "align=left;verticalAlign=middle;fontSize=13;"
                           "fontColor=#1A2744;fontStyle=1;"),
                    vertex="1", parent="1")
        _geo(ush, x=PX, y=cur_y, w=900, h=26)
        cur_y += 32

        ua_cont_h = 32 + len(unassoc) * (CLI_ICO_H + CLI_PAD)
        ua_id = cid()
        uc = _cell(g, id=ua_id,
                   value="<b>&#128270; Probing — no association seen</b>",
                   style=("swimlane;startSize=28;fillColor=#f5f5f5;"
                          f"strokeColor={CLI_UA_STR};strokeWidth=1.5;"
                          "fontColor=#444;fontSize=10;fontStyle=1;"
                          "rounded=1;arcSize=3;html=1;"),
                   vertex="1", parent="1")
        max_ua_w = min(AP_W * APS_ROW + AP_GAPX * (APS_ROW - 1), 1200)
        _geo(uc, x=PX, y=cur_y, w=max_ua_w, h=ua_cont_h)

        for ui, cli in enumerate(sorted(unassoc, key=lambda c: c.get("mac", ""))):
            mac    = cli.get("mac", "")
            cli_ip = mac_to_ip.get(mac, "")
            cli_hn = mac_to_hn.get(mac, "")
            # Only include non-empty, non-null SSIDs in probed list
            probed = sorted(
                xs(s) for s in cli.get("probed_ssids", set())
                if s and s.strip()
            )
            top = xs(cli_hn or cli_ip or mac)
            bot_parts = []
            if cli_hn or cli_ip:
                bot_parts.append(mac)
            if probed:
                probe_str = ", ".join(probed[:3])
                if len(probed) > 3:
                    probe_str += f" +{len(probed)-3}"
                bot_parts.append(f"Probing: {probe_str}")
            bot = xs(" | ".join(bot_parts))

            uval = (f"<b>{top}</b>"
                    + (f"<br/><font style='font-size:8px;color:#555;'>"
                       f"{bot}</font>" if bot else ""))
            utip = (f"MAC: {xs(mac)}\nIP: {xs(cli_ip or 'unknown')}"
                    f"\nHostname: {xs(cli_hn or 'unknown')}"
                    f"\nProbed SSIDs: {xs(', '.join(probed) or 'wildcard only')}")

            ui_node = _cell(g, id=cid(), value=uval, tooltip=utip,
                            style=(f"{SHAPE_LAPTOP}fillColor={CLI_UA_FILL};"
                                   f"strokeColor={CLI_UA_STR};"
                                   "verticalLabelPosition=right;verticalAlign=middle;"
                                   "labelPosition=right;align=left;"
                                   "fontSize=9;html=1;"),
                            vertex="1", parent=ua_id)
            _geo(ui_node, x=CLI_PAD, y=32 + ui * (CLI_ICO_H + CLI_PAD),
                 w=CLI_ICO_W, h=CLI_ICO_H)

        cur_y += ua_cont_h + AP_GAPY

    # ════════════════════════════════════════════════════════════════════════
    # SECTION 3 — Wired L2 segments from ARP / node MAC data
    # ════════════════════════════════════════════════════════════════════════
    wired_by_subnet = defaultdict(list)
    for ip, info in nodes.items():
        if not info.get("is_private"):
            continue
        macs = sorted(info.get("macs", set()))
        if not macs:
            continue
        subnet  = info.get("subnet", "unknown")
        anomaly = len(arp_table.get(ip, set())) > 1
        wired_by_subnet[subnet].append({
            "ip":       ip,
            "macs":     macs,
            "hostname": info.get("hostname", ""),
            "os_guess": info.get("os_guess", "Unknown"),
            "role":     info.get("role", "client"),
            "anomaly":  anomaly,
        })

    if wired_by_subnet:
        wsh = _cell(g, id=cid(),
                    value="<b>&#128279; Wired L2 Segments (ARP / MAC Table)</b>",
                    style=("text;html=1;strokeColor=none;fillColor=none;"
                           "align=left;verticalAlign=middle;fontSize=13;"
                           "fontColor=#1A2744;fontStyle=1;"),
                    vertex="1", parent="1")
        _geo(wsh, x=PX, y=cur_y, w=900, h=26)
        cur_y += 32

        seg_cx = PX; seg_cy = cur_y; scol = 0; srow_max = 0

        for subnet, hosts in sorted(wired_by_subnet.items()):
            n_hosts  = len(hosts)
            seg_h    = WIRED_H0 + n_hosts * (HOST_H + HOST_PAD) + HOST_PAD
            seg_id   = cid()
            sc = _cell(g, id=seg_id,
                       value=f"<b>Subnet {xs(subnet)}</b>",
                       style=(f"swimlane;startSize=28;fillColor={WIRED_FILL};"
                              f"strokeColor={WIRED_STR};strokeWidth=1.5;"
                              f"fontColor={WIRED_STR};fontSize=10;fontStyle=1;"
                              "rounded=1;arcSize=3;html=1;"),
                       vertex="1", parent="1")
            _geo(sc, x=seg_cx, y=seg_cy, w=WIRED_W, h=seg_h)

            for hi, host in enumerate(sorted(hosts, key=lambda h: h["ip"])):
                ip      = host["ip"]
                macs    = host["macs"]
                hn      = host["hostname"]
                os_g    = host["os_guess"]
                anomaly = host["anomaly"]

                top = xs(hn or ip)
                bot_parts = []
                if hn:
                    bot_parts.append(ip)
                for m in macs[:2]:
                    bot_parts.append(m)
                if len(macs) > 2:
                    bot_parts.append(f"+{len(macs)-2}")
                bot = xs(" | ".join(bot_parts))

                hfill  = ANOMALY_FILL if anomaly else WIRED_FILL
                hstroke = ANOMALY_STR if anomaly else WIRED_STR
                anom_txt = ("<br/><font style='font-size:8px;color:#b85450;'>"
                            "&#9888; ARP anomaly!</font>") if anomaly else ""

                hval = (f"<b>{'&#9888; ' if anomaly else ''}{top}</b>"
                        + anom_txt
                        + (f"<br/><font style='font-size:8px;color:#444;'>"
                           f"{bot}</font>" if bot else ""))
                htip = (f"IP: {xs(ip)}\nMAC(s): {xs(', '.join(macs))}"
                        f"\nHostname: {xs(hn or 'unknown')}\nOS: {xs(os_g)}"
                        + ("\n&#9888; ARP ANOMALY — possible MITM!" if anomaly else ""))

                hshape = (SHAPE_SERVER if host["role"] == "server" else SHAPE_PC)
                hn_node = _cell(g, id=cid(), value=hval, tooltip=htip,
                                style=(f"{hshape}fillColor={hfill};"
                                       f"strokeColor={hstroke};"
                                       "verticalLabelPosition=right;verticalAlign=middle;"
                                       "labelPosition=right;align=left;"
                                       "fontSize=9;html=1;"),
                                vertex="1", parent=seg_id)
                _geo(hn_node, x=HOST_PAD,
                     y=WIRED_H0 + hi * (HOST_H + HOST_PAD),
                     w=44, h=38)

            srow_max = max(srow_max, seg_h)
            scol += 1
            if scol >= SEGS_ROW:
                seg_cx = PX; seg_cy += srow_max + SEG_GAPY
                srow_max = 0; scol = 0
            else:
                seg_cx += WIRED_W + SEG_GAPX

    # ── Render to XML ─────────────────────────────────────────────────────────
    raw_xml = ET.tostring(root, encoding="unicode")
    # Final safety pass: strip any XML-illegal control characters that
    # may have survived from packet data (e.g. null bytes in malformed SSIDs).
    raw_xml = raw_xml.translate(_illegal_xml)
    return minidom.parseString(raw_xml).toprettyxml(indent="  ")


# ─────────────────────────────────────────────────────────────────────────────
# draw.io generator
# ─────────────────────────────────────────────────────────────────────────────

def generate_drawio(nodes, findings, gateways, traceroutes, title="Network Diagram"):
    """
    Host inventory diagram with:
      - Gateway lines (thin grey, gateways only)
      - Traceroute section below main diagram
    """
    node_pos, containers = layout(nodes)
    # Work out how tall the main diagram area is so we can place traceroute below
    max_y = max((c["y"] + c["h"] for c in containers), default=PAGE_Y) if containers else PAGE_Y

    # Build finding set for quick lookup
    flagged_ips = set()
    for f in findings:
        if f["severity"] in ("HIGH","CRITICAL"):
            flagged_ips.add(f["src"])
            if f["dst"] != "N/A": flagged_ips.add(f["dst"])

    root = ET.Element("mxGraphModel",
        dx="1422", dy="762", grid="1", gridSize="10", guides="1",
        tooltips="1", connect="1", arrows="1", fold="1", page="1",
        pageScale="1", pageWidth="1654", pageHeight="1169",
        math="0", shadow="0")
    g = ET.SubElement(root, "root")
    ET.SubElement(g, "mxCell", id="0")
    ET.SubElement(g, "mxCell", id="1", parent="0")

    # Title
    tc = _cell(g, id="title",
               value=f"<b>{title}</b>",
               style="text;html=1;strokeColor=none;fillColor=none;"
                     "align=center;verticalAlign=middle;whiteSpace=wrap;"
                     "rounded=0;fontSize=20;",
               vertex="1", parent="1")
    _geo(tc, x=PAGE_X, y=22, w=900, h=42)

    _add_legend(g, findings)

    # Subnet containers
    cont_ids = {}
    for ci, cont in enumerate(containers):
        cid = f"cont_{ci}"
        cont_ids[cont["subnet"]] = cid
        if cont["subnet"] == "external":
            fill, stroke = "#fff3e0","#e65100"; fc="#bf360c"
            lbl = "&#x2601;  External / Internet"
        else:
            fill, stroke = SUBNET_PALETTES[cont["palette_idx"] % len(SUBNET_PALETTES)]
            fc = stroke
            lbl = f"Subnet: {cont['subnet']}"

        cc = _cell(g, id=cid, value=f"<b>{lbl}</b>",
                   style=(f"swimlane;startSize={CONT_TITLE};fillColor={fill};"
                          f"strokeColor={stroke};fontColor={fc};fontSize=11;"
                          "fontStyle=1;rounded=1;arcSize=3;swimlaneLine=1;"
                          "align=left;spacingLeft=8;html=1;"),
                   vertex="1", parent="1")
        _geo(cc, x=cont["x"], y=cont["y"], w=cont["w"], h=cont["h"])

    # Nodes — NO edges drawn at all
    for ni, (ip, info) in enumerate(nodes.items()):
        nid = f"n{ni}"
        ax, ay  = node_pos[ip]
        sub     = info["subnet"]
        cont    = next(c for c in containers if c["subnet"] == sub)
        rx, ry  = ax - cont["x"], ay - cont["y"]
        parent  = cont_ids[sub]

        shape = (SHAPE_CLOUD  if not info["is_private"] else
                 SHAPE_SERVER if info["role"] == "server" else SHAPE_PC)

        # Red fill if flagged
        if ip in flagged_ips:
            fill, stroke = FLAG_FILL_STROKE
        else:
            role_k = info["role"] if info["is_private"] else "external"
            fill, stroke = ROLE_FILL_STROKE.get(role_k, ("#fff","#999"))

        hostname = info.get("hostname","")
        macs     = sorted(info["macs"])
        mac_str  = " | ".join(macs) if macs else "MAC: unknown"
        os_str   = info.get("os_guess","Unknown")
        flags    = info.get("flags", set())
        flag_str = "  ".join(sorted(flags)) if flags else ""

        # Label: hostname (bold) / IP / MAC / OS guess
        if hostname:
            label = (f"<b>{hostname}</b><br/>"
                     f"<font style='font-size:9px;'>{ip}</font><br/>"
                     f"<font style='font-size:8px;color:#444;'>{mac_str}</font><br/>"
                     f"<font style='font-size:8px;color:#666;'>OS: {os_str}</font>")
        else:
            label = (f"<b>{ip}</b><br/>"
                     f"<font style='font-size:8px;color:#444;'>{mac_str}</font><br/>"
                     f"<font style='font-size:8px;color:#666;'>OS: {os_str}</font>")

        protos  = sorted(info["protocols"])
        ports   = sorted(info["open_ports"])
        tooltip = (f"Protocols: {', '.join(protos)}\n"
                   f"Open ports (passive): {', '.join(str(p) for p in ports) or 'none seen'}\n"
                   f"OS guess: {os_str}\n"
                   f"Pkts: {info['count']:,}  Bytes: {info['bytes']:,}\n"
                   + (f"⚠ FLAGS: {flag_str}" if flag_str else ""))

        nc = _cell(g, id=nid, value=label, tooltip=tooltip,
                   style=(f"{shape}fillColor={fill};strokeColor={stroke};"
                          "verticalLabelPosition=bottom;verticalAlign=top;"
                          "labelPosition=center;align=center;"
                          "labelBackgroundColor=none;labelBorderColor=none;"
                          "fontSize=9;whiteSpace=wrap;html=1;"),
                   vertex="1", parent=parent)
        # Icon geometry: NODE_W × NODE_H only.
        # The label renders below this box via verticalLabelPosition=bottom.
        # Row spacing (STRIDE_Y) already reserves LABEL_RESERVE px for it,
        # so labels never reach the next row's icons.
        _geo(nc, x=rx, y=ry, w=NODE_W, h=NODE_H)

    # ── Gateway lines: thin grey lines between gateway and every node in subnet ──
    gw_node_ids = {}   # ip -> id  (for gateways that ARE nodes)
    for ip, nid_candidate in zip(
            [ip for ip in nodes],
            [f"n{ni}" for ni in range(len(nodes))]):
        if ip in gateways.values():
            gw_node_ids[ip] = nid_candidate

    # Build reverse lookup: nid for each ip
    node_id_map = {ip: f"n{ni}" for ni, ip in enumerate(nodes)}

    edge_idx = 0
    for subnet, gw_ip in gateways.items():
        if gw_ip not in node_id_map:
            continue
        gw_id = node_id_map[gw_ip]
        # Connect gateway to every other node in the same subnet
        for ip, info in nodes.items():
            if ip == gw_ip:
                continue
            if info["subnet"] != subnet:
                continue
            if ip not in node_id_map:
                continue
            ec = _cell(g, id=f"gwe{edge_idx}", value="",
                       style=("endArrow=none;startArrow=none;"
                              "strokeColor=#bbbbbb;strokeWidth=1;opacity=50;"
                              "dashed=1;dashPattern=4 4;"
                              "rounded=1;html=1;"),
                       edge="1", source=gw_id, target=node_id_map[ip],
                       parent="1")
            ET.SubElement(ec, "mxGeometry", relative="1", **{"as":"geometry"})
            edge_idx += 1

    # ── Traceroute section ─────────────────────────────────────────────────────
    if traceroutes:
        tr_y = max_y + 80   # start below main diagram
        _draw_traceroute_section(g, traceroutes, node_id_map, tr_y)

    raw = ET.tostring(root, encoding="unicode")
    return minidom.parseString(raw).toprettyxml(indent="  ")


def _draw_traceroute_section(g, traceroutes, node_id_map, start_y):
    """
    Draw traceroute hop chains below the main diagram.
    Each trace is rendered as a horizontal chain of hop nodes
    connected by labelled arrows, grouped in a swimlane container.
    """
    HOP_W     = 120
    HOP_H     = 50
    HOP_GAP   = 60
    SECT_PAD  = 20
    SECT_TITLE= 30
    ROW_H     = HOP_H + SECT_PAD * 2 + SECT_TITLE + 20
    cy        = int(start_y)
    SHAPE_ROUTER = "shape=mxgraph.cisco.routers.router;"

    # Section header label
    hdr = _cell(g, id="tr_hdr",
                value="<b>&#128246; Traceroute Paths (reconstructed from ICMP TTL-exceeded)</b>",
                style=("text;html=1;strokeColor=none;fillColor=none;"
                       "align=left;verticalAlign=middle;fontSize=13;"
                       "fontColor=#333;fontStyle=1;"),
                vertex="1", parent="1")
    _geo(hdr, x=PAGE_X, y=cy, w=900, h=28)
    cy += 36

    for ti, trace in enumerate(traceroutes):
        hops     = trace["hops"]
        n_hops   = len(hops)
        if n_hops == 0:
            continue

        src_label = trace.get("src_hostname") or trace["src"]
        dst_label = trace.get("dst_hostname") or trace["dst"]
        title_lbl = (f"<b>Trace {ti+1}:</b>  {src_label}  →  {dst_label}  "
                     f"<font style='font-size:9px;color:#666;'>({n_hops} hop{"s" if n_hops!=1 else ""})</font>")

        # Container width: origin + hops + destination
        total_nodes = 1 + n_hops + 1
        cw = SECT_PAD * 2 + total_nodes * HOP_W + (total_nodes - 1) * HOP_GAP
        ch = SECT_TITLE + SECT_PAD * 2 + HOP_H

        cid = f"tr_cont_{ti}"
        cc = _cell(g, id=cid, value=title_lbl,
                   style=(f"swimlane;startSize={SECT_TITLE};"
                          "fillColor=#f0f4f8;strokeColor=#607d8b;"
                          "fontColor=#37474f;fontSize=10;"
                          "fontStyle=0;rounded=1;arcSize=3;html=1;"),
                   vertex="1", parent="1")
        _geo(cc, x=PAGE_X, y=cy, w=cw, h=ch)

        # Helper: draw a hop node inside the container
        def hop_node(node_id, col_idx, label, shape, fill, stroke, tooltip=""):
            nx = SECT_PAD + col_idx * (HOP_W + HOP_GAP)
            ny = SECT_TITLE + SECT_PAD
            nc = _cell(g, id=node_id, value=label, tooltip=tooltip,
                       style=(f"{shape}fillColor={fill};strokeColor={stroke};"
                              "verticalLabelPosition=bottom;verticalAlign=top;"
                              "labelPosition=center;align=center;fontSize=9;html=1;"),
                       vertex="1", parent=cid)
            _geo(nc, x=nx, y=ny, w=HOP_W, h=HOP_H)
            return node_id

        # Helper: draw an arrow between two hop nodes
        def hop_edge(eid, src_id, tgt_id, lbl=""):
            ec = _cell(g, id=eid, value=lbl,
                       style=("endArrow=block;endFill=1;"
                              "strokeColor=#607d8b;strokeWidth=1.5;"
                              "fontSize=8;fontColor=#607d8b;"
                              "rounded=1;html=1;"),
                       edge="1", source=src_id, target=tgt_id,
                       parent=cid)
            ET.SubElement(ec, "mxGeometry", relative="1", **{"as":"geometry"})

        # Origin node (the client that ran the trace)
        origin_id = f"tr{ti}_origin"
        origin_lbl = (f"<b>{src_label}</b><br/>"
                      f"<font style='font-size:8px;color:#555;'>{trace['src']}</font>")
        hop_node(origin_id, 0, origin_lbl,
                 "shape=mxgraph.cisco.computers_and_peripherals.pc;",
                 "#fff2cc","#d6b656", f"Traceroute origin: {trace['src']}")

        prev_id = origin_id
        for hi, hop in enumerate(hops):
            hop_ip  = hop["router_ip"]
            hop_hn  = hop.get("hostname", "")
            hop_lbl = (f"<b>Hop {hop['hop_n']}</b><br/>"
                       f"<font style='font-size:8px;'>{hop_hn or hop_ip}</font><br/>"
                       f"<font style='font-size:7px;color:#888;'>{hop_ip if hop_hn else ''}</font>")
            nid = f"tr{ti}_hop{hi}"
            # Check if this hop IP is a known node — if so link style differs
            is_known = hop_ip in node_id_map
            fill   = "#dae8fc" if is_known else "#f5f5f5"
            stroke = "#6c8ebf" if is_known else "#aaaaaa"
            known_note = "\nKnown node in diagram" if is_known else ""
            hop_node(nid, hi+1, hop_lbl, SHAPE_ROUTER, fill, stroke,
                     tooltip=f"Router hop {hop['hop_n']}: {hop_ip}{known_note}")
            hop_edge(f"tr{ti}_e{hi}", prev_id, nid)
            prev_id = nid

        # Destination node
        dst_id  = f"tr{ti}_dst"
        dst_lbl = (f"<b>{dst_label}</b><br/>"
                   f"<font style='font-size:8px;color:#555;'>{trace['dst']}</font>")
        hop_node(dst_id, n_hops+1, dst_lbl,
                 "shape=mxgraph.cisco.storage.cloud;",
                 "#ffe6cc","#d79b00", f"Trace destination: {trace['dst']}")
        hop_edge(f"tr{ti}_efinal", prev_id, dst_id, "")

        cy += ch + 16   # gap between traces



def _add_legend(g, findings):
    sev_counts = defaultdict(int)
    for f in findings: sev_counts[f["severity"]] += 1

    roles = [("server","Server"),("host","Host"),
             ("client","Client"),("external","External")]
    height = 30 + len(roles)*24 + 20 + (60 if findings else 0)

    bg = _cell(g, id="legend_bg", value="<b>Legend</b>",
               style=("rounded=1;html=1;fillColor=#f5f5f5;strokeColor=#666;"
                      "fontColor=#333;align=center;verticalAlign=top;"
                      "spacingTop=6;fontSize=11;"),
               vertex="1", parent="1")
    _geo(bg, x=20, y=100, w=LEGEND_W, h=height)

    for ri, (role, lbl) in enumerate(roles):
        y = 100 + 30 + ri*24
        fill, stroke = ROLE_FILL_STROKE[role]
        rs = _cell(g, id=f"rs{ri}", value="",
                   style=(f"shape=mxgraph.cisco.computers_and_peripherals.pc;"
                          f"fillColor={fill};strokeColor={stroke};"),
                   vertex="1", parent="1")
        _geo(rs, x=28, y=y, w=20, h=20)
        rll = _cell(g, id=f"rll{ri}", value=lbl,
                    style="text;html=1;strokeColor=none;fillColor=none;"
                          "align=left;verticalAlign=middle;fontSize=10;",
                    vertex="1", parent="1")
        _geo(rll, x=54, y=y, w=148, h=20)

    if findings:
        base = 100 + 30 + len(roles)*24 + 14
        # Red flag node indicator
        rs2 = _cell(g, id="rs_flag", value="",
                    style=(f"rounded=1;html=1;"
                           f"fillColor={FLAG_FILL_STROKE[0]};strokeColor={FLAG_FILL_STROKE[1]};"),
                    vertex="1", parent="1")
        _geo(rs2, x=28, y=base, w=20, h=14)
        total = len(findings)
        crit  = sev_counts.get("CRITICAL",0)
        high  = sev_counts.get("HIGH",0)
        rll2 = _cell(g, id="rll_flag",
                     value=f"⚠ Pentest finding ({total} total, {crit} CRIT, {high} HIGH)",
                     style="text;html=1;strokeColor=none;fillColor=none;"
                           "align=left;verticalAlign=middle;fontSize=10;fontColor=#b85450;",
                     vertex="1", parent="1")
        _geo(rll2, x=54, y=base, w=148, h=20)
        note = _cell(g, id="rll_note",
                     value="Hover nodes &amp; see Excel<br/>Pentest Findings sheet",
                     style="text;html=1;strokeColor=none;fillColor=none;"
                           "align=left;verticalAlign=middle;fontSize=9;fontColor=#888;",
                     vertex="1", parent="1")
        _geo(note, x=24, y=base+22, w=165, h=28)


# ─────────────────────────────────────────────────────────────────────────────
# Excel export
# ─────────────────────────────────────────────────────────────────────────────

def generate_xlsx(rows, nodes, edges, findings, cleartext_hits, banner_hits, tls_sessions, dns_events, wifi_data, output_path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[!] openpyxl not installed — skipping XLSX.  pip install openpyxl")
        return False

    wb = Workbook()

    # ── Common styles ──────────────────────────────────────────────────────────
    def hdr_style(fill_hex="1F4E79"):
        return dict(
            font  =Font(name="Arial", bold=True, color="FFFFFF", size=10),
            fill  =PatternFill("solid", fgColor=fill_hex),
            align =Alignment(horizontal="center", vertical="center", wrap_text=True),
        )

    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    bf   = Font(name="Arial", size=9)
    ALT_FILL = PatternFill("solid", fgColor="EBF3FB")
    WHT_FILL = PatternFill("solid", fgColor="FFFFFF")
    ctr  = Alignment(horizontal="center", vertical="center")
    lft  = Alignment(horizontal="left",   vertical="center")

    def apply_hdr(ws, headers, fill_hex="1F4E79"):
        ws.row_dimensions[1].height = 30
        hs = hdr_style(fill_hex)
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = hs["font"]; c.fill = hs["fill"]
            c.alignment = hs["align"]; c.border = bdr

    def _sanitise(v):
        """Strip illegal XML/openpyxl characters (control chars, null bytes)."""
        if not isinstance(v, str):
            return v
        # Remove null bytes and other control characters that openpyxl rejects
        import re as _re
        return _re.sub('[\\x00-\\x08\\x0b\\x0c\\x0e-\\x1f]', '', v)

    def body_cell(ws, row, col, val, alt_row=False, centre=False):
        c = ws.cell(row=row, column=col, value=_sanitise(val))
        c.font = bf
        c.fill = ALT_FILL if alt_row else WHT_FILL
        c.border = bdr
        c.alignment = ctr if centre else lft
        return c

    def col_w(ws, widths):
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    def add_autofilter(ws, ncols, hdr_row=1):
        """Add Excel AutoFilter drop-downs to the header row."""
        ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(ncols)}{hdr_row}"

    # ── Sheet 1: Connections ──────────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "Connections"
    HDR1 = ["Hostname (src)","MAC Address (src)","Source IP",
            "Hostname (dst)","MAC Address (dst)","Destination IP",
            "Protocol","Port","Resource (if HTTP/HTTPS)",
            "Assessed Server Role","Packets"]
    apply_hdr(ws1, HDR1)
    ws1.freeze_panes = "A2"
    add_autofilter(ws1, len(HDR1))

    def hn(ip):  return nodes.get(ip,{}).get("hostname","")
    def fm(ip):
        macs = sorted(nodes.get(ip,{}).get("macs",set()))
        return macs[0] if macs else ""
    def rl(ip):  return nodes.get(ip,{}).get("role","unknown")

    agg = defaultdict(lambda: dict(count=0, resources=set()))
    for r in rows:
        k = (r["src_ip"], r["dst_ip"], r["proto"], r["port"],
             r.get("src_mac",""), r.get("dst_mac",""))
        agg[k]["count"] += 1
        if r.get("resource"): agg[k]["resources"].add(r["resource"])

    for ri, ((src_ip,dst_ip,proto,port,smac,dmac), info) in \
            enumerate(sorted(agg.items()), 2):
        ws1.row_dimensions[ri].height = 15
        alt = (ri % 2 == 0)
        vals = [hn(src_ip), smac or fm(src_ip), src_ip,
                hn(dst_ip), dmac or fm(dst_ip), dst_ip,
                proto, port or "", "; ".join(sorted(info["resources"])),
                rl(dst_ip), info["count"]]
        for ci, v in enumerate(vals, 1):
            body_cell(ws1, ri, ci, v, alt, centre=(ci in (7,8,11)))

    col_w(ws1, [18,18,15,18,18,15,13,7,28,18,10])

    # ── Sheet 2: Node Summary ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Node Summary")
    HDR2 = ["IP Address","Hostname","MAC Address(es)","Subnet","Role",
            "OS Guess (TTL)","TTL Value","Is Private",
            "Protocols Observed","Passive Open Ports",
            "Pentest Flags","Packet Count","Bytes"]
    apply_hdr(ws2, HDR2, "1F4E79")
    ws2.freeze_panes = "A2"
    add_autofilter(ws2, len(HDR2))

    def node_sort(item):
        ip, info = item
        r = {"server":0,"host":1,"client":2,"external":3}.get(info["role"],4)
        try: return (r, int(ipaddress.ip_address(ip.split("/")[0])))
        except: return (r, ip)

    for ri, (ip, info) in enumerate(sorted(nodes.items(), key=node_sort), 2):
        ws2.row_dimensions[ri].height = 15
        alt = (ri%2==0)
        vals = [
            ip,
            info.get("hostname",""),
            ", ".join(sorted(info["macs"])),
            info["subnet"],
            info["role"],
            info.get("os_guess","Unknown"),
            info.get("ttl_val",""),
            "Yes" if info["is_private"] else "No",
            ", ".join(sorted(info["protocols"])),
            ", ".join(str(p) for p in sorted(info["open_ports"])),
            "; ".join(sorted(info.get("flags",set()))),
            info["count"],
            info["bytes"],
        ]
        for ci, v in enumerate(vals, 1):
            c = body_cell(ws2, ri, ci, v, alt, centre=(ci in (6,7,8,12,13)))
            # Highlight flagged rows
            if info.get("flags"):
                if ci == 11:
                    c.font = Font(name="Arial", size=9, color="B85450", bold=True)

    col_w(ws2, [15,18,30,18,10,15,9,10,45,22,28,12,12])

    # ── Sheet 3: Pentest Findings ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Pentest Findings")
    HDR3 = ["Severity","Category","Source","Destination","Detail","Recommendation"]
    apply_hdr(ws3, HDR3, "8B0000")
    ws3.freeze_panes = "A2"
    add_autofilter(ws3, len(HDR3))

    SEV_COLOURS = {"CRITICAL":"FF0000","HIGH":"FF6600",
                   "MEDIUM":"FFB300","LOW":"00AA00","INFO":"888888"}

    sev_order = {"CRITICAL":0,"HIGH":1,"MEDIUM":2,"LOW":3,"INFO":4}
    sorted_findings = sorted(findings, key=lambda f: sev_order.get(f["severity"],5))

    for ri, f in enumerate(sorted_findings, 2):
        ws3.row_dimensions[ri].height = 28
        vals = [f["severity"], f["category"], f["src"], f["dst"],
                f["detail"], f["recommendation"]]
        for ci, v in enumerate(vals, 1):
            c = body_cell(ws3, ri, ci, v, False, centre=(ci==1))
            c.alignment = Alignment(horizontal="center" if ci==1 else "left",
                                    vertical="center", wrap_text=True)
            if ci == 1:
                clr = SEV_COLOURS.get(f["severity"],"888888")
                c.font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
                c.fill  = PatternFill("solid", fgColor=clr)
                c.border = bdr

    col_w(ws3, [12,28,16,16,50,42])

    # ── Sheet 4: Protocol Summary ─────────────────────────────────────────────
    ws4 = wb.create_sheet("Protocol Summary")
    HDR4 = ["Protocol","Connections","Unique Sources","Unique Destinations",
            "Cleartext?","Lateral Movement Risk?"]
    apply_hdr(ws4, HDR4, "1F4E79")
    ws4.freeze_panes = "A2"
    add_autofilter(ws4, len(HDR4))

    proto_stats = defaultdict(lambda: dict(count=0, srcs=set(), dsts=set()))
    for (src_ip,dst_ip,proto,port,smac,dmac), info in agg.items():
        proto_stats[proto]["count"] += info["count"]
        proto_stats[proto]["srcs"].add(src_ip)
        proto_stats[proto]["dsts"].add(dst_ip)

    for ri, (proto, ps) in enumerate(
            sorted(proto_stats.items(), key=lambda x: -x[1]["count"]), 2):
        alt = (ri%2==0)
        vals = [proto, ps["count"], len(ps["srcs"]), len(ps["dsts"]),
                "YES ⚠" if proto in CLEARTEXT_PROTOS else "No",
                "YES 🔴" if proto in LATERAL_PROTOS else "No"]
        for ci, v in enumerate(vals, 1):
            c = body_cell(ws4, ri, ci, v, alt, centre=True)
            if ci==5 and v.startswith("YES"):
                c.font = Font(name="Arial", size=9, bold=True, color="B85450")
            if ci==6 and v.startswith("YES"):
                c.font = Font(name="Arial", size=9, bold=True, color="CC0000")

    col_w(ws4, [16,14,16,20,14,22])

    # ── Sheet 5: Port Inventory ───────────────────────────────────────────────
    ws5 = wb.create_sheet("Port Inventory")
    HDR5 = ["IP","Hostname","Role","OS Guess","Subnet",
            "Passive Open Ports","Well-Known Service Names","Suspicious Ports"]
    apply_hdr(ws5, HDR5, "1F4E79")
    ws5.freeze_panes = "A2"
    add_autofilter(ws5, len(HDR5))

    for ri, (ip, info) in enumerate(sorted(nodes.items(), key=node_sort), 2):
        ports    = sorted(info["open_ports"])
        svc_names = [WELL_KNOWN.get(("TCP",p)) or WELL_KNOWN.get(("UDP",p)) or ""
                     for p in ports]
        sus_ports = [str(p) for p in ports if p in SUSPICIOUS_PORTS]
        alt = (ri%2==0)
        vals = [ip, info.get("hostname",""), info["role"],
                info.get("os_guess","Unknown"), info["subnet"],
                ", ".join(str(p) for p in ports),
                ", ".join(s for s in svc_names if s),
                ", ".join(sus_ports) or ""]
        for ci, v in enumerate(vals, 1):
            c = body_cell(ws5, ri, ci, v, alt, centre=(ci in (3,4)))
            if ci==8 and v:
                c.font = Font(name="Arial", size=9, bold=True, color="B85450")

    col_w(ws5, [15,18,10,14,18,28,30,20])


    # ── Sheet 6: Cleartext Intercepts ────────────────────────────────────────
    ws6 = wb.create_sheet("Cleartext Intercepts")
    HDR6 = ["Protocol","Data Type","Extracted Value","Context / Raw",
            "Source IP","Destination IP","Src Port","Dst Port"]
    apply_hdr(ws6, HDR6, "4A0000")
    ws6.freeze_panes = "A2"
    add_autofilter(ws6, len(HDR6))

    # De-duplicate: same (proto, type, value, src, dst) seen in multiple packets
    seen_ct = set()
    deduped = []
    for h in cleartext_hits:
        key = (h["protocol"], h["type"], h["value"], h["src_ip"], h["dst_ip"])
        if key not in seen_ct:
            seen_ct.add(key)
            deduped.append(h)

    TYPE_SEVERITY = {
        "FTP Password":"CRITICAL", "FTP Username":"HIGH",
        "HTTP Basic Auth (decoded)":"CRITICAL", "HTTP Basic Auth (raw b64)":"HIGH",
        "HTTP Bearer Token":"HIGH", "HTTP Cookie":"MEDIUM",
        "HTTP POST Credential":"CRITICAL", "HTTP JSON Credential":"CRITICAL",
        "Telnet Keystrokes/Data":"HIGH", "Telnet Data":"HIGH",
        "SMTP Auth Command":"HIGH", "SMTP Auth (b64 decoded)":"CRITICAL",
        "POP3 Password":"CRITICAL", "POP3 Username":"HIGH",
        "IMAP Login":"CRITICAL",
        "LDAP Bind Data":"HIGH",
        "SNMP Community String":"MEDIUM",
        "AWS Access Key":"CRITICAL", "Private Key":"CRITICAL",
        "API Key":"HIGH", "Auth Token":"HIGH", "Secret/Key":"HIGH",
    }

    SEV_CLR = {
        "CRITICAL": ("FF0000","FFFFFF"),
        "HIGH":     ("FF6600","FFFFFF"),
        "MEDIUM":   ("FFB300","000000"),
        "LOW":      ("00AA00","FFFFFF"),
    }

    # Sort: CRITICAL first, then by protocol
    def ct_sort(h):
        sev = TYPE_SEVERITY.get(h["type"],"LOW")
        return ({"CRITICAL":0,"HIGH":1,"MEDIUM":2,"LOW":3}.get(sev,4), h["protocol"])

    deduped.sort(key=ct_sort)

    if not deduped:
        ws6.cell(row=2, column=1, value="No cleartext credentials or sensitive data detected.")
    else:
        for ri, h in enumerate(deduped, 2):
            ws6.row_dimensions[ri].height = 20
            sev  = TYPE_SEVERITY.get(h["type"], "LOW")
            bg, fg = SEV_CLR.get(sev, ("FFFFFF","000000"))
            alt_row = (ri % 2 == 0)

            vals = [h["protocol"], h["type"], h["value"], h["context"],
                    h["src_ip"], h["dst_ip"], h["src_port"] or "", h["dst_port"] or ""]
            for ci, v in enumerate(vals, 1):
                c = body_cell(ws6, ri, ci, v, alt_row, centre=(ci in (1,7,8)))
                # Colour the Data Type cell by severity
                if ci == 2:
                    c.font  = Font(name="Arial", size=9, bold=True, color=fg)
                    c.fill  = PatternFill("solid", fgColor=bg)
                    c.border = bdr
                # Highlight the value cell for critical items
                if ci == 3 and sev == "CRITICAL":
                    c.font = Font(name="Arial", size=9, bold=True, color="8B0000")

    col_w(ws6, [14, 28, 45, 45, 15, 15, 9, 9])

    # Add a summary note at the top
    ws6.insert_rows(1)
    ws6.row_dimensions[1].height = 40
    summary = ws6.cell(row=1, column=1,
        value=(f"CLEARTEXT INTERCEPTS — {len(deduped)} unique items captured  "
               f"| CRITICAL: {sum(1 for h in deduped if TYPE_SEVERITY.get(h['type'])=='CRITICAL')}  "
               f"HIGH: {sum(1 for h in deduped if TYPE_SEVERITY.get(h['type'])=='HIGH')}  "
               f"MEDIUM: {sum(1 for h in deduped if TYPE_SEVERITY.get(h['type'])=='MEDIUM')}"))
    summary.font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    summary.fill  = PatternFill("solid", fgColor="4A0000")
    summary.alignment = Alignment(horizontal="left", vertical="center",
                                   indent=1, wrap_text=False)
    ws6.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)



    # ── Sheet 7: Banner & Resource Intelligence ───────────────────────────────
    ws7 = wb.create_sheet("Banner Intel")

    # Sub-section colours
    CAT_COLOURS = {
        "Banner":          ("1A3A5C", "FFFFFF"),   # dark navy
        "Resource":        ("1A5C2E", "FFFFFF"),   # dark green
        "Client Software": ("5C3A1A", "FFFFFF"),   # dark brown
    }

    # Build structured sections: Banners, then Resources, then Client Software
    def _banner_sort(h):
        cat_order = {"Banner":0, "Resource":1, "Client Software":2}
        return (cat_order.get(h["category"], 9), h["protocol"], h["server_ip"])

    banner_hits_sorted = sorted(banner_hits, key=_banner_sort)

    HDR7 = ["Category", "Banner / Resource Type", "Server / Host IP",
            "Client IP", "Port", "Protocol", "Value / Detail", "Context"]
    apply_hdr(ws7, HDR7, "1A3A5C")
    ws7.freeze_panes = "A2"
    add_autofilter(ws7, len(HDR7))

    BANNER_TYPE_NOTE = {
        "HTTP Server Header":   "⚠ Version disclosure — update server header suppression",
        "X-Powered-By":         "⚠ Technology disclosure — remove in production",
        "X-Generator":          "⚠ CMS/framework version disclosed",
        "ASP.NET Version":      "⚠ .NET version disclosure — disable X-AspNet-Version header",
        "Via (Proxy)":          "ℹ Proxy/load balancer in path",
        "FTP Banner":           "⚠ Server version in banner — consider suppressing",
        "SMTP Banner":          "⚠ Mail server version disclosed in greeting",
        "SSH Version String":   "⚠ SSH version disclosed — consider 'DebianBanner no'",
        "Telnet Version String":"⚠ Device version in Telnet banner",
        "Telnet Login Banner":  "ℹ Login banner content",
        "SNMP sysDescr":        "⚠ OS/device info in SNMP — consider restricting access",
        "DHCP Vendor Class":    "ℹ Client identifies itself to DHCP server",
        "Runtime Version":      "⚠ Runtime version disclosure",
        "HTTP Request":         "ℹ Commonly requested resource",
        "HTTP User-Agent":      "ℹ Client software / OS fingerprint",
        "DNS Query":            "ℹ Domain being resolved",
        "NTP Server":           "ℹ NTP time source",
        "DHCP Server":          "ℹ DHCP server identity",
        "SMTP EHLO Domain":     "ℹ Client mail domain",
    }

    # De-duplicate banner_hits (already done inside extract_banners, but wb sheet needs it)
    seen_b = set()
    deduped_b = []
    for h in banner_hits_sorted:
        key = (h["category"], h["banner_type"], h["server_ip"], h["value"])
        if key not in seen_b:
            seen_b.add(key)
            deduped_b.append(h)

    if not deduped_b:
        ws7.cell(row=2, column=1, value="No banners or notable resources detected.")
    else:
        current_cat = None
        data_row = 2

        for h in deduped_b:
            cat = h["category"]

            # Insert a category divider row when the category changes
            if cat != current_cat:
                current_cat = cat
                bg_hex, fg_hex = CAT_COLOURS.get(cat, ("444444","FFFFFF"))
                div_cell = ws7.cell(row=data_row, column=1,
                                    value=f"── {cat.upper()} ──")
                div_cell.font  = Font(name="Arial", bold=True, size=10, color=fg_hex)
                div_cell.fill  = PatternFill("solid", fgColor=bg_hex)
                div_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                div_cell.border = bdr
                # Merge across all columns for the divider
                ws7.merge_cells(start_row=data_row, start_column=1,
                                end_row=data_row, end_column=8)
                ws7.row_dimensions[data_row].height = 18
                data_row += 1

            ws7.row_dimensions[data_row].height = 22
            note = BANNER_TYPE_NOTE.get(h["banner_type"], "")
            vals = [
                h["category"],
                h["banner_type"],
                h["server_ip"],
                h["client_ip"],
                h["port"] or "",
                h["protocol"],
                h["value"],
                note or h.get("context",""),
            ]
            for ci, v in enumerate(vals, 1):
                c = body_cell(ws7, data_row, ci, v, (data_row % 2 == 0),
                              centre=(ci in (1,5,6)))
                c.alignment = Alignment(
                    horizontal="center" if ci in (1,5,6) else "left",
                    vertical="center", wrap_text=(ci in (7,8)))
                # Colour-code the Category column
                if ci == 1:
                    bg_hex, fg_hex = CAT_COLOURS.get(cat, ("444444","FFFFFF"))
                    c.font  = Font(name="Arial", size=9, bold=True, color=fg_hex)
                    c.fill  = PatternFill("solid", fgColor=bg_hex)
                    c.border = bdr
                # Highlight ⚠ notes in the Context column
                if ci == 8 and str(v).startswith("⚠"):
                    c.font = Font(name="Arial", size=9, color="B85450", bold=True)

            data_row += 1

    col_w(ws7, [16, 26, 16, 16, 7, 12, 55, 48])

    # Summary header at top
    ws7.insert_rows(1)
    ws7.row_dimensions[1].height = 36
    n_banners   = sum(1 for h in deduped_b if h["category"] == "Banner")
    n_resources = sum(1 for h in deduped_b if h["category"] == "Resource")
    n_clients   = sum(1 for h in deduped_b if h["category"] == "Client Software")
    summ = ws7.cell(row=1, column=1,
        value=(f"BANNER & RESOURCE INTELLIGENCE  |  "
               f"Service Banners: {n_banners}   "
               f"Resources / Queries: {n_resources}   "
               f"Client Software: {n_clients}"))
    summ.font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    summ.fill  = PatternFill("solid", fgColor="1A3A5C")
    summ.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws7.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)



    # ── Sheet 8: TLS Session Analysis ────────────────────────────────────────
    ws8 = wb.create_sheet("TLS Sessions")

    HDR8 = [
        "Risk",
        "Client IP", "Server IP", "Port",
        "SNI / Hostname", "ALPN",
        "TLS Version (Negotiated)", "Cipher Suite",
        "Cert Subject", "Cert Issuer", "SANs",
        "Cert Valid From", "Cert Expiry",
        "Key Type", "Key Bits",
        "Handshake Complete", "Alerts",
        "Issues / Findings",
    ]
    apply_hdr(ws8, HDR8, "1B3A5C")
    ws8.freeze_panes = "A2"
    add_autofilter(ws8, len(HDR8))
    ws8.row_dimensions[1].height = 36

    # Severity helpers
    def _tls_risk(sess):
        if sess.get("cert_expired"):        return "CRITICAL"
        if sess.get("weak_cipher"):         return "HIGH"
        if sess.get("weak_version"):        return "HIGH"
        if sess.get("cert_expiring_soon"):  return "MEDIUM"
        if sess.get("issues"):              return "MEDIUM"
        if not sess.get("tls_version"):     return "INFO"
        return "OK"

    TLS_RISK_CLR = {
        "CRITICAL": ("C00000", "FFFFFF"),
        "HIGH":     ("C55A11", "FFFFFF"),
        "MEDIUM":   ("BF8F00", "000000"),
        "INFO":     ("595959", "FFFFFF"),
        "OK":       ("375623", "FFFFFF"),
    }

    # Sort: worst issues first, then by server IP
    def _tls_sort(s):
        order = {"CRITICAL":0,"HIGH":1,"MEDIUM":2,"INFO":3,"OK":4}
        return (order.get(_tls_risk(s), 9), s["server_ip"], s["server_port"])

    tls_sorted = sorted(tls_sessions, key=_tls_sort)

    if not tls_sorted:
        ws8.cell(row=2, column=1,
                 value="No TLS handshakes detected in capture — only encrypted app-data seen, or no TLS traffic present.")
    else:
        for ri, sess in enumerate(tls_sorted, 2):
            ws8.row_dimensions[ri].height = 22
            alt = (ri % 2 == 0)
            risk = _tls_risk(sess)
            risk_bg, risk_fg = TLS_RISK_CLR.get(risk, ("FFFFFF","000000"))

            cs = sess.get("cipher_suite","")
            ver = sess.get("tls_version","")
            expiry = sess.get("cert_not_after","")

            issues_str = "; ".join(sess.get("issues",[])) if sess.get("issues") else ""
            alerts_str = "; ".join(sess.get("alerts",[])) if sess.get("alerts") else ""

            vals = [
                risk,
                sess["client_ip"],
                sess["server_ip"],
                sess["server_port"] or "",
                sess.get("sni",""),
                sess.get("alpn",""),
                ver,
                cs,
                sess.get("cert_subject",""),
                sess.get("cert_issuer",""),
                sess.get("cert_sans",""),
                sess.get("cert_not_before",""),
                expiry,
                sess.get("cert_key_type",""),
                sess.get("cert_key_bits","") or "",
                "Yes" if sess.get("handshake_complete") else "Partial",
                alerts_str,
                issues_str,
            ]

            for ci, v in enumerate(vals, 1):
                centre_cols = {1, 4, 6, 7, 14, 15, 16}
                c = body_cell(ws8, ri, ci, v, alt, centre=(ci in centre_cols))
                c.alignment = Alignment(
                    horizontal="center" if ci in centre_cols else "left",
                    vertical="center",
                    wrap_text=(ci in {8, 11, 18}),
                )

                # Risk column — colour-coded badge
                if ci == 1:
                    c.font  = Font(name="Arial", bold=True, size=9, color=risk_fg)
                    c.fill  = PatternFill("solid", fgColor=risk_bg)
                    c.border = bdr

                # TLS version — red if deprecated
                elif ci == 7 and ver in ("SSLv3","TLS 1.0","TLS 1.1"):
                    c.font = Font(name="Arial", size=9, bold=True, color="C00000")

                # Cipher suite — orange if weak
                elif ci == 8 and sess.get("weak_cipher"):
                    c.font = Font(name="Arial", size=9, bold=True, color="C55A11")

                # Cert expiry — red if expired, amber if soon
                elif ci == 13:
                    if sess.get("cert_expired"):
                        c.font = Font(name="Arial", size=9, bold=True, color="C00000")
                    elif sess.get("cert_expiring_soon"):
                        c.font = Font(name="Arial", size=9, bold=True, color="BF8F00")

                # Key bits — red if weak
                elif ci == 15 and isinstance(v, int):
                    kt = sess.get("cert_key_type","")
                    if ("RSA" in kt and v and v < 2048) or ("ECDSA" in kt and v and v < 256):
                        c.font = Font(name="Arial", size=9, bold=True, color="C00000")

                # Issues — bold red if non-empty
                elif ci == 18 and v:
                    c.font = Font(name="Arial", size=9, bold=True, color="8B0000")

    col_w(ws8, [10, 15, 15, 6, 28, 10, 14, 38, 28, 28, 32, 13, 13, 14, 8, 10, 28, 48])

    # Summary banner at top
    ws8.insert_rows(1)
    ws8.row_dimensions[1].height = 36
    n_ok       = sum(1 for s in tls_sessions if _tls_risk(s) == "OK")
    n_critical = sum(1 for s in tls_sessions if _tls_risk(s) == "CRITICAL")
    n_high     = sum(1 for s in tls_sessions if _tls_risk(s) == "HIGH")
    n_medium   = sum(1 for s in tls_sessions if _tls_risk(s) == "MEDIUM")
    n_info     = sum(1 for s in tls_sessions if _tls_risk(s) == "INFO")
    tls_summ = ws8.cell(row=1, column=1,
        value=(f"TLS SESSION ANALYSIS  |  Total: {len(tls_sessions)}   "
               f"✓ OK: {n_ok}   "
               f"⚠ MEDIUM: {n_medium}   "
               f"▲ HIGH: {n_high}   "
               f"✖ CRITICAL: {n_critical}   "
               f"ℹ Info only: {n_info}"))
    tls_summ.font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    tls_summ.fill  = PatternFill("solid", fgColor="1B3A5C")
    tls_summ.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws8.merge_cells(start_row=1, start_column=1, end_row=1, end_column=18)



    # ── Sheet 9: DNS / mDNS Events ────────────────────────────────────────────
    ws9 = wb.create_sheet("DNS Events")

    HDR9 = ["Direction","Protocol","Client IP","Server/Resolver IP",
            "Query Name","Type","Answer IP / Value","Answer Name",
            "TTL (s)","Response Code","Flags"]
    apply_hdr(ws9, HDR9, "1A4A1A")
    ws9.freeze_panes = "A2"
    add_autofilter(ws9, len(HDR9))

    RTYPE_CLR = {"A":"D9F0D3","AAAA":"C6E2FF","CNAME":"FFF2CC",
                 "PTR":"FFE0CC","MX":"F0D9F0","TXT":"F0F0F0",
                 "SRV":"E8D5FF","NS":"D0D0D0","HTTPS":"C6F0FF"}
    RCODE_CLR = {"NXDOMAIN":"FFC0C0","SERVFAIL":"FFB366","REFUSED":"FFD0D0"}

    if not dns_events:
        ws9.cell(row=2, column=1, value="No DNS/mDNS events captured in this PCAP.")
    else:
        # Sort: queries first, then responses; within each by client IP
        def _dns_sort(e):
            return (0 if not e["is_response"] else 1,
                    e.get("client_ip",""),
                    e.get("query_name",""))
        sorted_events = sorted(dns_events, key=_dns_sort)

        for ri, ev in enumerate(sorted_events, 2):
            ws9.row_dimensions[ri].height = 20
            alt = (ri % 2 == 0)
            direction = "← Response" if ev["is_response"] else "→ Query"
            rcode = ev.get("rcode","")

            vals = [
                direction,
                ev.get("proto","DNS"),
                ev.get("client_ip",""),
                ev.get("server_ip",""),
                ev.get("query_name",""),
                ev.get("qtype",""),
                ev.get("answer_val",""),
                ev.get("answer_name",""),
                ev.get("ttl","") or "",
                rcode,
                ev.get("flags_desc",""),
            ]
            for ci, v in enumerate(vals, 1):
                c = body_cell(ws9, ri, ci, v, alt, centre=(ci in {1,2,6,9,10}))
                c.alignment = Alignment(
                    horizontal="center" if ci in {1,2,6,9,10} else "left",
                    vertical="center",
                    wrap_text=(ci in {5,7,8,11}))

                # Colour direction column
                if ci == 1:
                    clr = "1A4A1A" if "Response" in str(v) else "4A1A1A"
                    c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
                    c.fill = PatternFill("solid", fgColor=clr)

                # Colour record type
                elif ci == 6:
                    qtype_bg = RTYPE_CLR.get(str(v),"FFFFFF")
                    c.fill   = PatternFill("solid", fgColor=qtype_bg)
                    c.font   = Font(name="Arial", size=9, bold=True)

                # Red for NXDOMAIN etc
                elif ci == 10 and rcode in RCODE_CLR:
                    c.fill = PatternFill("solid", fgColor=RCODE_CLR[rcode])
                    c.font = Font(name="Arial", size=9, bold=True, color="8B0000")

    col_w(ws9, [12, 8, 16, 18, 38, 8, 32, 28, 7, 10, 30])

    # Summary banner
    ws9.insert_rows(1)
    ws9.row_dimensions[1].height = 32
    n_queries   = sum(1 for e in dns_events if not e["is_response"])
    n_responses = sum(1 for e in dns_events if e["is_response"])
    n_nxdomain  = sum(1 for e in dns_events if e.get("rcode")=="NXDOMAIN")
    n_mdns      = sum(1 for e in dns_events if e.get("proto")=="mDNS")
    unique_names = len({e["query_name"] for e in dns_events if e.get("query_name")})
    dns_summ = ws9.cell(row=1, column=1,
        value=(f"DNS & mDNS EVENT LOG  |  "
               f"Total: {len(dns_events)}   "
               f"Queries: {n_queries}   Responses: {n_responses}   "
               f"mDNS: {n_mdns}   NXDOMAIN: {n_nxdomain}   "
               f"Unique names queried: {unique_names}"))
    dns_summ.font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    dns_summ.fill  = PatternFill("solid", fgColor="1A4A1A")
    dns_summ.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws9.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)


    # ── Sheet 10: Wi-Fi Networks & Clients ────────────────────────────────────
    ws10 = wb.create_sheet("Wi-Fi Networks")
    aps_list     = wifi_data.get("aps", [])
    clients_list = wifi_data.get("clients", [])
    wifi_events  = wifi_data.get("events", [])

    if not aps_list and not clients_list:
        ws10.cell(row=1, column=1,
                  value=("No 802.11 management frames found. "
                         "To capture Wi-Fi data, use monitor mode: "
                         "airmon-ng start wlan0 && tcpdump -i wlan0mon -w wifi.pcap"))
        col_w(ws10, [80])
    else:
        # ── AP table ──────────────────────────────────────────────────────
        HDR_AP = ["BSSID (AP MAC)","SSID","Channel","Encryption",
                  "WPS","Beacons","Probe Resp","Associated Clients","Notes"]
        apply_hdr(ws10, HDR_AP, "1B2A5C")
        ws10.freeze_panes = "A2"
        add_autofilter(ws10, len(HDR_AP))

        ENC_CLR = {"Open":"C00000","WPA":"C55A11","WPA2":"375623",
                   "WPA3":"1F7A1F","WPA2+FT":"375623"}

        for ri, ap in enumerate(aps_list, 2):
            ws10.row_dimensions[ri].height = 22
            alt = (ri % 2 == 0)
            enc = ap.get("enc","Open")
            ssid = ap.get("ssid","<Hidden>") or "<Hidden>"
            notes = []
            if enc == "Open":     notes.append("⚠ Open network — no encryption")
            if ap.get("wps"):     notes.append("⚠ WPS enabled — KRACK/Pixie-Dust risk")
            if not ap.get("ssid"): notes.append("Hidden SSID")

            vals = [
                ap["bssid"],
                ssid,
                ap.get("channel","") or "",
                enc,
                "Yes" if ap.get("wps") else "No",
                ap.get("beacons",0),
                ap.get("probe_responses",0),
                ", ".join(ap.get("clients",[])) or "None seen",
                "; ".join(notes),
            ]
            for ci, v in enumerate(vals, 1):
                c = body_cell(ws10, ri, ci, v, alt, centre=(ci in {3,5,6,7}))
                c.alignment = Alignment(
                    horizontal="center" if ci in {3,5,6,7} else "left",
                    vertical="center", wrap_text=(ci in {8,9}))
                # Colour encryption column
                if ci == 4:
                    enc_bg = ENC_CLR.get(enc.split("+")[0], "595959")
                    c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
                    c.fill = PatternFill("solid", fgColor=enc_bg)
                # Red for open / WPS
                elif ci == 9 and notes:
                    c.font = Font(name="Arial", size=9, bold=True, color="8B0000")

        col_w(ws10, [20, 28, 8, 10, 6, 9, 10, 45, 45])

        # ── Clients section ───────────────────────────────────────────────
        client_start = len(aps_list) + 4
        ws10.cell(row=client_start-1, column=1, value="Wi-Fi Clients / Stations")
        hdr_row = ws10.row_dimensions[client_start-1]
        hdr_row.height = 24
        c_hdr = ws10.cell(row=client_start-1, column=1)
        c_hdr.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        c_hdr.fill = PatternFill("solid", fgColor="3A1A5C")
        ws10.merge_cells(start_row=client_start-1, start_column=1,
                         end_row=client_start-1, end_column=4)

        HDR_CL = ["Client MAC","Associated BSSID","Associated SSID","Probed SSIDs"]
        for ci, h in enumerate(HDR_CL, 1):
            c = ws10.cell(row=client_start, column=ci, value=h)
            c.font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            c.fill  = PatternFill("solid", fgColor="3A1A5C")
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = bdr

        for ri, cl in enumerate(clients_list, client_start+1):
            ws10.row_dimensions[ri].height = 20
            alt = (ri % 2 == 0)
            vals = [
                cl["mac"],
                cl.get("associated_bssid",""),
                cl.get("assoc_ssid",""),
                ", ".join(sorted(cl.get("probed_ssids", set()))) or "(wildcard only)",
            ]
            for ci, v in enumerate(vals, 1):
                body_cell(ws10, ri, ci, v, alt)

        # ── Events section ────────────────────────────────────────────────
        if wifi_events:
            evt_start = client_start + len(clients_list) + 4
            ws10.cell(row=evt_start-1, column=1, value="802.11 Management Frame Events")
            c_hdr2 = ws10.cell(row=evt_start-1, column=1)
            c_hdr2.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
            c_hdr2.fill = PatternFill("solid", fgColor="1A3A5C")
            ws10.merge_cells(start_row=evt_start-1, start_column=1,
                             end_row=evt_start-1, end_column=6)

            HDR_EV = ["Frame Type","Source MAC","Destination MAC","BSSID","SSID","Detail"]
            for ci, h in enumerate(HDR_EV, 1):
                c = ws10.cell(row=evt_start, column=ci, value=h)
                c.font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
                c.fill  = PatternFill("solid", fgColor="1A3A5C")
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.border = bdr

            EVT_CLR = {
                "Beacon":             "EBF3FB",
                "Probe Request":      "FFF2CC",
                "Probe Response":     "E2EFDA",
                "Association Request":"D9D9FF",
                "Deauthentication":   "FFD0D0",
                "Disassociation":     "FFDCC0",
            }
            # Limit to 500 events in sheet (can be huge)
            shown_events = wifi_events[:500]
            for ri, ev in enumerate(shown_events, evt_start+1):
                ws10.row_dimensions[ri].height = 18
                alt = (ri % 2 == 0)
                ft = ev.get("frame_type","")
                vals = [
                    ft,
                    ev.get("src_mac",""),
                    ev.get("dst_mac",""),
                    ev.get("bssid",""),
                    ev.get("ssid",""),
                    ev.get("detail",""),
                ]
                row_fill = EVT_CLR.get(ft, "FFFFFF")
                for ci, v in enumerate(vals, 1):
                    c = body_cell(ws10, ri, ci, v, False)
                    c.fill = PatternFill("solid", fgColor=row_fill)
                    c.alignment = Alignment(horizontal="left", vertical="center")
                    if ft in ("Deauthentication","Disassociation") and ci == 1:
                        c.font = Font(name="Arial", size=9, bold=True, color="C00000")

        # Summary banner (insert at top)
        ws10.insert_rows(1)
        ws10.row_dimensions[1].height = 32
        n_open = sum(1 for a in aps_list if a.get("enc","Open") == "Open")
        n_wps  = sum(1 for a in aps_list if a.get("wps"))
        n_deauth = sum(1 for e in wifi_events if e.get("frame_type") == "Deauthentication")
        wifi_summ = ws10.cell(row=1, column=1,
            value=(f"Wi-Fi NETWORK SURVEY  |  "
                   f"APs: {len(aps_list)}   "
                   f"Clients: {len(clients_list)}   "
                   f"Open networks: {n_open}   "
                   f"WPS enabled: {n_wps}   "
                   f"Deauth frames: {n_deauth}   "
                   f"Total mgmt events: {len(wifi_events)}"))
        wifi_summ.font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        wifi_summ.fill  = PatternFill("solid", fgColor="1B2A5C")
        wifi_summ.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws10.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)


    wb.save(output_path)
    return True


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(
        description="PCAP → draw.io + Excel  (v4.0 Pentester Edition)")
    ap.add_argument("pcap")
    ap.add_argument("-o","--output",  help=".drawio output path")
    ap.add_argument("--xlsx",         help=".xlsx output path")
    ap.add_argument("--min-packets",  type=int, default=1)
    ap.add_argument("--collapse-external", action="store_true")
    ap.add_argument("--hostname-file", metavar="FILE")
    ap.add_argument("--title", default="Network Diagram")
    ap.add_argument("--l2-output", metavar="FILE",
                    help=".drawio output path for L2/Wi-Fi topology diagram "
                         "(default: <input>_l2.drawio)")
    ap.add_argument("--internal-networks", metavar="CIDR", nargs="+",
                    help="Additional CIDR ranges to treat as internal "
                         "(e.g. 20.16.0.0/14 for corporate non-RFC1918 space)")
    args = ap.parse_args()

    base    = args.pcap.rsplit(".",1)[0]
    out_dio = args.output or base+".drawio"
    out_xl  = args.xlsx   or base+".xlsx"
    out_l2  = getattr(args, "l2_output", None) or base+"_l2.drawio"

    print(f"[*] Parsing {args.pcap} ...")
    try:
        packets = list(parse_pcap(args.pcap))
    except FileNotFoundError:
        print(f"[!] Not found: {args.pcap}", file=sys.stderr); sys.exit(1)
    except ValueError as e:
        print(f"[!] {e}",              file=sys.stderr); sys.exit(1)

    print(f"[*] {len(packets):,} packets")
    if not packets:
        print("[!] No packets — valid PCAP?", file=sys.stderr); sys.exit(1)

    extra_hn = _load_hostname_file(args.hostname_file)
    if args.internal_networks:
        import ipaddress as _ipmod
        parsed = []
        for cidr in args.internal_networks:
            try:
                parsed.append(_ipmod.ip_network(cidr, strict=False))
                print(f"[*] Treating {cidr} as internal network")
            except ValueError as e:
                print(f"[!] Bad --internal-networks value '{cidr}': {e}", file=sys.stderr)
        if parsed:
            extra_hn["__internal_networks__"] = parsed
    nodes, edges, rows, findings, cleartext_hits, arp_table = build_graph(
        packets, args.min_packets, args.collapse_external, extra_hn)

    unique_conn = len({tuple(sorted(p)) for p in edges})
    subnets     = len({n["subnet"] for n in nodes.values()})
    print(f"[*] {len(nodes)} nodes · {unique_conn} connections · {subnets} subnets")
    print(f"[*] {len(findings)} pentest findings  "
          f"({sum(1 for f in findings if f['severity']=='CRITICAL')} CRITICAL  "
          f"{sum(1 for f in findings if f['severity']=='HIGH')} HIGH  "
          f"{sum(1 for f in findings if f['severity']=='MEDIUM')} MEDIUM)")

    # Gateway detection
    gateways = detect_gateways(packets, nodes)
    print(f"[*] {len(gateways)} gateway(s) detected: " +
          ", ".join(f"{s}→{ip}" for s,ip in gateways.items()) if gateways else "[*] No gateways detected")

    # Traceroute reconstruction
    traceroutes = extract_traceroutes(packets)
    # Enrich hop hostnames from resolved node names
    passive_hn = resolve_hostnames_from_packets(packets)
    all_hn = {**passive_hn, **extra_hn}
    for tr in traceroutes:
        tr["src_hostname"] = all_hn.get(tr["src"], "")
        tr["dst_hostname"] = all_hn.get(tr["dst"], "")
        for hop in tr["hops"]:
            hop["hostname"] = all_hn.get(hop["router_ip"], "")
    if traceroutes:
        print(f"[*] {len(traceroutes)} traceroute path(s) reconstructed")
        for tr in traceroutes:
            hops_str = " → ".join(h["hostname"] or h["router_ip"] for h in tr["hops"])
            print(f"    {tr['src']} → [{hops_str}] → {tr['dst']}")

    # Banner / resource extraction
    banner_hits = extract_banners(packets)
    dns_hits    = [b for b in banner_hits if b["banner_type"] == "DNS Query"]
    svc_banners = [b for b in banner_hits if b["category"] == "Banner"]
    resources   = [b for b in banner_hits if b["category"] == "Resource" and b["banner_type"] != "DNS Query"]
    print(f"[*] {len(svc_banners)} service banners · {len(resources)} resources · {len(dns_hits)} DNS queries")

    # TLS handshake analysis
    tls_sessions = extract_tls_sessions(packets)
    n_tls_issues = sum(1 for s in tls_sessions if s.get("issues"))
    print(f"[*] {len(tls_sessions)} TLS session(s) reconstructed  "
          f"({n_tls_issues} with issues)")

    # DNS / mDNS event log
    dns_events = _extract_dns_events(packets)
    n_mdns     = sum(1 for e in dns_events if e.get("proto") == "mDNS")
    n_nxdomain = sum(1 for e in dns_events if e.get("rcode") == "NXDOMAIN")
    print(f"[*] {len(dns_events)} DNS events extracted  "
          f"({n_mdns} mDNS · {n_nxdomain} NXDOMAIN)")

    # Wi-Fi 802.11 network survey
    wifi_data   = extract_wifi_events(packets)
    n_aps       = len(wifi_data["aps"])
    n_clients   = len(wifi_data["clients"])
    n_deauths   = sum(1 for e in wifi_data["events"] if e.get("frame_type") == "Deauthentication")
    if n_aps or n_clients:
        print(f"[*] Wi-Fi survey: {n_aps} AP(s) · {n_clients} client(s) · {n_deauths} deauth frame(s)")
    else:
        print(f"[*] No 802.11 frames detected (not a Wi-Fi capture)")

    xml = generate_drawio(nodes, findings, gateways, traceroutes, title=args.title)
    with open(out_dio,"w",encoding="utf-8") as f:
        f.write(xml)
    print(f"[+] Diagram  → {out_dio}")

    l2_xml = generate_l2_drawio(wifi_data, nodes, arp_table,
                                 title=f"{args.title} — L2/Wi-Fi Topology")
    with open(out_l2,"w",encoding="utf-8") as f:
        f.write(l2_xml)
    print(f"[+] L2 Diagram → {out_l2}")

    ok = generate_xlsx(rows, nodes, edges, findings, cleartext_hits, banner_hits, tls_sessions, dns_events, wifi_data, out_xl)
    if ok:
        print(f"[+] Workbook → {out_xl}  (10 sheets: Connections, Node Summary, "
              f"Pentest Findings, Protocol Summary, Port Inventory, "
              f"Cleartext Intercepts, Banner Intel, TLS Sessions, "
              f"DNS Events, Wi-Fi Networks)")

    print()
    print("  draw.io: File → Import From → Device → .drawio file")
    print("  Excel:   Pentest Findings sheet has colour-coded severity rows")
    print("           Node Summary has OS guesses, open ports, and risk flags")


if __name__ == "__main__":
    main()
